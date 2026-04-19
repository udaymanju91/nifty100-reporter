#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Nifty 100 Daily EMA Cross Reporter
Version: 2.4.0
Released: 2026-04-16

Change log
- 3.3.8
- HTML READABILITY OVERHAUL (full pass):
  * Darker, lower-contrast bg (#0b1623) — easier on eyes than harsh navy.
  * Consistent text colour palette via CSS variables throughout.
  * Table: uppercase column headers, 2px bottom border, tighter cell padding.
  * TD improvements: Symbol bold+blue, Company ellipsis, Close right-align
    monospace, RSI colour-banded (green/orange/red), 52WH green / 52WL red,
    Conviction Score JS-coloured (green/orange/red), news smaller font.
  * Market Health pane: full inline-style dark-theme rewrite — no light
    colours (#f5f5f5, #eee, #ddd, #666, #e8eaf6) remain anywhere.
  * MH score rows: colour-coded score values, proper row separators.
  * MH header strip uses market-regime colour band.
  * MH subheader: dark bg, position size in accent gold.
  * thead sticky adjusted to top:55px matching new header height.
  * News list items: .73rem, dim colour, proper list indent.
  * Disclaimer: .70rem muted, border-top separator.
- 3.3.7
- HTML REDESIGN: full dark-theme refresh for readability.
  * Font: Inter (Google Fonts) replacing Arial; consistent size scale.
  * Viewport meta tag added (mobile-safe).
  * CSS custom properties (--bg, --text, --accent etc.) for consistency.
  * Sticky top header bar with date, version, history row count.
  * Sticky table thead (top:58px) — column labels always visible.
  * Zebra striping on table rows (nth-child even).
  * Row hover highlight (#162f58).
  * ⇅ sort-hint arrows on all sortable columns before click.
  * Horizontal scroll .tbl-wrap with styled scrollbar.
  * Collapsible sections (▲/▼ toggle button per section).
  * Signal count badge auto-populated by JS on each section.
  * Floating ↑ back-to-top button.
  * Disclaimer footer styled and separated.
  * Market Health pane: all light-theme inline colours replaced
    with dark-theme equivalents (score bar, label, value, card bg).
- 3.3.6
- CONFIRMED FIX: switched index OHLCV + VIX primary source from
  nseindia.com (permanently Akamai-blocked for requests library) to
  niftyindices.com POST API (official NSE subsidiary, no bot protection).
  Confirmed working via live test (test_nse_api_v4.py):
    endpoint : POST .../Backpage.aspx/getHistoricaldatatabletoString
    payload  : {"cinfo": json.dumps({name, startDate, endDate, indexName})}
    date fmt : yyyy-MM-dd (unambiguous, confirmed working)
    fields   : HistoricalDate, OPEN, HIGH, LOW, CLOSE
    session  : GET https://www.niftyindices.com/ to prime cookies
  VIX: same endpoint with name="India VIX"
  Fallback: yfinance ^NSEI / JUNIORBEES.NS / ^INDIAVIX (all confirmed working)
  NSE indicesHistory/vixHistory calls removed entirely.
- 3.3.5
- ROOT CAUSE FIX: NSE indicesHistory and vixHistory 404s.
  Issue 1 — Wrong URL encoding: requests.get(params=) encodes space as
  "+" (application/x-www-form-urlencoded). NSE API requires "%%20".
  Result: "NIFTY+50" was sent instead of "NIFTY%%2050" → 404 every time.
  Fix: URL built manually using urllib.parse.quote(safe="") which
  produces proper %%20 encoding. params= dict removed from NSE calls.
  Issue 2 — Wrong endpoint name: "vixhistory" (lowercase h) does not
  exist. Correct endpoint is "vixHistory" (capital H).
  Fix: URL changed to .../api/historical/vixHistory
  Issue 3 — Chunk size 90d → 60d for both index OHLCV and VIX.
  Example correct URLs now generated:
    indicesHistory?indexType=NIFTY%%2050&from=01-01-2024&to=01-03-2024
    vixHistory?from=01-01-2024&to=01-03-2024
- 3.3.4
- AUDIT FIX fetch_index_membership (found during full network-call audit):
  * Index param strings changed: "NIFTY%%2050" -> "NIFTY 50", etc.
    (hardcoded URL encoding was redundant and fragile)
  * URL build changed from f-string concat to params= dict passed into
    _http_get_with_retry — consistent with all other NSE API calls.
  * Zero impact on normal runs (cache-first; cache already exists).
    Fix ensures live re-fetch works after quarterly NSE rebalancing.
- Full network-call audit: ALL existing calls confirmed unaffected by MH changes.
  * fetch_universe, bootstrap_nse_session, delivery, backfill,
    corporate_actions, fetch_recent_news — all ✅ unchanged.
- 3.3.3
- FIX fetch_index_ohlcv:
  * NSE API now called with params= dict + explicit Referer/Accept headers
    (root cause of 404s: URL was string-concatenated, not params-encoded)
  * History start fixed to 2020-10-20 (replaces 520d rolling window)
  * yfinance interval="1d" explicitly set (was defaulting to monthly)
  * Nifty Next 50 yfinance tickers corrected to:
    NIFTYNXT50.NS → ^NIFNXT50 → JUNIORBEES.NS (tried in order)
  * ignore_tz=True added for clean date normalization
- FIX fetch_vix_history:
  * NSE vixhistory now called with params= dict + proper headers
  * yfinance fallback also uses interval="1d" + ignore_tz=True
- EMA values for indexes: computed on the fly from persisted OHLCV CSV
  (correct pattern — no separate EMA cache needed for indexes)
- 3.3.2
- FIX: MergedCell AttributeError in render_excel — column_dimensions loop
  now skips MergedCell objects (triggered by Market Health merged A1:F1).
- FIX: NSE indicesHistory API returning 404 — added yfinance fallback:
  fetch_index_ohlcv() now falls back to ^NSEI (Nifty 50) and ^CNXNXT
  (Nifty Next 50) via yfinance when NSE API returns empty/404.
  Results cached to same CSV; subsequent runs use cache, no yfinance hit.
- FIX: NSE vixhistory API returning 404 — added yfinance fallback:
  fetch_vix_history() now falls back to ^INDIAVIX via yfinance.
  Results cached to mh_vix_cache.csv.
- 3.3.1
- BUG FIX: Market Health was silently skipped on every run with
  "name index_map is not defined". Root cause: index_map is a local
  variable inside build_analysis() and was never accessible in main().
  Fix: main() now calls fetch_index_membership(s) directly (cache-first,
  zero extra network cost) into _mh_index_map before passing to
  compute_market_health(). All cache files now created on first run.
- 3.3.0
- Market Health Dashboard (zero changes to existing screening logic):
  * fetch_index_ohlcv(): cache-first NSE indicesHistory, date-chunked, 520d history
  * fetch_vix_history(): cache-first NSE vixhistory API
  * compute_trend_score(): EMA10/20/50/200 + slope bonuses + HH+HL structure
  * update_breadth_ema_cache(): incremental per-stock EMA10/20/50 cache
  * compute_breadth_score(): % stocks > EMA20 AND EMA50, both displayed
  * compute_momentum_score(): RSI-14 + MACD(12/26/9) on index OHLCV
  * compute_volatility_score(): India VIX -> score [-100,+100]
  * compute_rs_score(): NN50 vs N50 20d RS ratio
  * compute_market_health(): orchestrator, weights 35/25/20/10/10
  * render_market_health_html(): side-by-side panes prepended to HTML
  * render_market_health_excel(): Market Health tab inserted as Sheet 1
  * Score scale: -100 (strong bear) to +100 (strong bull); 5 regimes
  * Breadth: BOTH EMA20 and EMA50 % shown in pane and detail rows
  * New cache files: mh_idx_NIFTY_50.csv, mh_idx_NIFTY_NEXT_50.csv,
    mh_vix_cache.csv, mh_breadth_ema_cache.csv
- 3.2.1
- Index badge vertical-align changed middle->top so badge sits at
  top of cell, consistent with all other HTML table columns.
- 3.2.0
- Index membership column:
  * fetch_index_membership(): cache-first (data/nifty50_members.json,
    data/niftynext50_members.json); fetches NSE equity-stockIndices API
    only when cache absent. Delete cache after quarterly NSE rebalancing.
  * New "Index" col in HTML (col 1): N50 blue | NN50 orange | N100 grey.
  * CONVICTION_COL updated 16->17. PATH constants added.
  * Fully graceful: failure -> N100 badge, run continues normally.
- 3.1.9
- Corporate actions HTML rendering fix (3 bugs fixed):
  FIX1: Removed wrong v3.1.8 normalisation that converted news_html list→str.
  FIX2: CA block prepended as first list item ([ca_html] + news_html list)
        instead of broken str+list concatenation.
  FIX3: HTML renderer hardened: accepts str OR list for RecentNewsHtml;
        str is wrapped in [str] instead of being discarded.
  FIX4: fetch_corporate_actions log now includes comma-separated symbol
        names of all symbols that have upcoming corporate events.
- 3.1.8
- _build_stock_row: normalise news_html/news_text to str immediately after
  fetch_recent_news() call. fetch_recent_news() returns html_items as a list;
  the ca_html + news_html concatenation (added in v3.1.6) broke with
  TypeError: can only concatenate str (not "list") to str.
- 3.1.7
- fetch_universe() hardened: NiftyIndices CSV sometimes contains preamble
  disclaimer lines before the real header. Now scans for the first line
  containing "symbol" + comma and parses from there. Fixes:
  "Error tokenizing data: Expected 1 fields in line 4, saw 10"
- 3.1.6
- Corporate Actions: new fetch_corporate_actions() using confirmed NSE endpoint
  (?index=equities&from_date=&to_date=). Bulk call + per-symbol fallback.
  Explicit "no events" message per symbol. Colour-coded urgency in news cell.
  _http_get_with_retry: params= kwarg + auto cookie refresh on 401/403/429.
  Config: corporate_action_lookahead_days (default 14).
- 3.1.5
- Retry logic: added _http_get_with_retry() helper (5 attempts, 10s apart).
  Applied to ALL network-touching fetch points that previously had no retry:
    • fetch_universe()            — niftyindices.com CSV (5 retries, 10s)
    • bootstrap_nse_session()     — nseindia.com homepage (5 retries, 10s)
    • backfill_delivery_history() — archives.nseindia.com (5 retries, 10s)
    • fetch_recent_news()         — Google News RSS (3 retries, 5s)
    • fetch_delivery_data()       — archives.nseindia.com (already had own
                                    retry loop; inner GET now uses helper too)
- 3.1.4
- Delivery Fix 4 (root cause fix): build_analysis() was calling
  get_delivery_series() with TODAY_IST (the wall-clock datetime, e.g.
  2026-04-17 14:11 IST) instead of RESOLVED_TRADE_DATE (the actual
  analysis date, e.g. 2026-04-16). This caused the "prior rows strictly
  before trade_date" filter to include the 16-Apr cached delivery row
  in prior.tail(4), and then today_pct (also 16-Apr = 45.93) was
  appended again, producing the duplicate-last-value sparkline:
    [51.4% → 51.4% → 52.7% → 45.9% → 45.9%]
  Fix: call site now passes RESOLVED_TRADE_DATE, giving:
    [46.1% → 51.4% → 51.4% → 52.7% → 45.9%]
- 3.1.3
- Delivery Fix 3: get_delivery_series() — normalise trade_date to a plain
  date object before applying the strict-less-than filter for "prior" rows.
  Previously, when RESOLVED_TRADE_DATE (a date object) was passed as
  trade_date, calling trade_date.date() raised AttributeError inside the
  try block, which caused the function to fall through to the bare except
  and return a single-element series. In the fetch_delivery_data() path,
  trade_date arrives as a datetime (used_date), so the comparison succeeded
  but included today's cached row (16-Apr) in "prior", then appended
  today_pct again — producing the duplicate-last-value sparkline
  e.g. [51.4% → 51.4% → 52.7% → 45.9% → 45.9%] instead of
       [46.1% → 51.4% → 51.4% → 52.7% → 45.9%].
- 3.1.2
- Delivery Fix 2: _conv_delivery() sparkline now uses vals[-5:] to build
  the DelivTrend display string, ensuring today_pct is never double-counted.
- 3.1.1
- Delivery Fix 1: fetch_delivery_data() now deduplicates (Date, Symbol)
  before saving the cache; extra safety repair pass after every write;
  get_delivery_series() keeps highest DelivPct per date when duplicates exist.
- 3.1.0
- News enhancement: fetch_recent_news() now includes site-targeted queries
  for Economic Times, Moneycontrol, and Business Standard in addition to
  the existing generic Google News RSS queries. Each article now displays
  a source badge ([ET], [MC], [BS], [Mint], [FE], [Reuters] etc.) in both
  HTML report and Excel text output. Source extracted from article URL via
  _news_source_label() helper. recent_news_max_items default raised 3 → 5.
- 3.0.0
- Fix 13: EMA warm-up discard — the first `span` bars of each EMA are now
  set to NaN before use in any crossover or ribbon logic. Prevents false
  signals from mathematically unreliable warm-up values. Affects EMA 9, 10,
  20, 21, 50, 200. For Nifty 100 stocks with years of history the change in
  signal output is zero. Only newly listed / short-history stocks are affected.
- Fix 15: PARKED — HMM daily cache deferred (single daily run — no benefit).
- 2.5.0
- Fix 8: scan_pullback_to_ema() — crossover age limit (max 60 bars).
  Pullback signals against a crossover older than 60 trading bars are
  suppressed. last_cross_date is read from nifty100_signal_log.csv and
  passed per-symbol from build_analysis().
- Fix 9: REVERTED — EMA Ribbon re-entry deduplication removed.
  Ribbon "New Entry" restored to original simple logic:
  ribbon active today AND not active yesterday.
- Delivery URL fix: fetch_delivery_data() updated from
  nsearchives.nseindia.com → archives.nseindia.com.
  Confirmed working: 16-Apr-2026 file returned 2,455 EQ rows
  with DELIV_PER column intact. All column mapping and EQ
  filter logic unchanged.
- Fix 10: dow_theory_label() — volume now modifies the phase label itself.
  Bull: Markup + contracting vol  -> "Bull: Markup ⚠ Vol Contraction"
  Bear: Markdown + expanding vol  -> "Bear: Markdown ⚠ Vol Surge"
  Bear: Accumulation + contracting -> "Bear: Accumulation ✓ Vol Contraction"
  Bull: Accumulation + expanding   -> "Bull: Accumulation ✓ Vol Surge"
- Fix 11: update_signal_tracker() + SIGNAL_LOG_COLS — days_aligned and
  exit_reason columns added. days_aligned incremented each run for Active
  signals. exit_reason written on status transition (EMA alignment lost /
  stop hit / target reached / archived). CSV migration patch handles
  existing rows that lack these columns.
- Fix 12: log_new_signals() — suggested_stop/target_1/target_2 now anchored
  to the nearest parsed S/R zone level above/below the close price.
  ATR multiples (1.5x stop, 5%/10% targets) used as fallback only when
  no S/R zone is available. days_aligned and exit_reason seeded at row
  creation.
- 2.4.0
- Fix 1: detect_candlestick() — Hammer/Hanging Man/Shooting Star now use
  _slope_direction() on prior bars instead of 10-bar close compare.
  Shooting Star no longer requires a bearish body (per Nison 1991).
- Fix 3: macd() — Signal line seeded with ema() helper, consistent with
  ema12 and ema26. Eliminates warm-up mismatch between line and signal.
- Fix 4: rsi() — Warm-up guard added. Returns NaN series when
  len(close) < period * 2, preventing false signals in first 28 bars.
- Fix 6: compute_conviction_score() — Hard cap at 58.0 for stocks in
  Bear: Markdown or Bear: Distribution. Warning banner injected into
  both HTML and text breakdown fields.
- Fix 7: detect_chart_patterns() — Conflicting pattern warning appended
  when bullish and bearish patterns fire simultaneously.
- Fix 14: build_analysis() — Minimum bars guard raised 60 → 220.
  Symbols below threshold are skipped with a structured logger.warning().
- Fix 16: classify_sentiment() — India-market keyword pre-pass added.
  INDIA_BULLISH_KW (20 terms) and INDIA_BEARISH_KW (18 terms) cover
  NSE-specific events that VADER mis-classifies (circuits, SEBI probes,
  bulk/block deals, promoter activity, pledging, forensic audits).
- 2.3.1
- Signal Tracker:
  New persistent nifty100_signal_log.csv tracks every EMA 9x21 and EMA 21x50
  crossover signal with 40 fields across 7 groups: identity, signal-day OHLCV,
  technical levels, conviction scores, live tracking, signal health/status, and
  risk levels (ATR-based stop + targets).
  rotate_signal_log(): auto-rotates to zip archive in data/archives/ when file
  reaches cfg.signal_log_max_mb (default 30 MB). Active rows younger than
  cfg.signal_log_archive_days (default 90 days) are retained in the live file.
  Signal Tracker tab added to daily Excel report.
  New Config fields: signal_log_max_mb (default 30), signal_log_archive_days (default 90).
- HTML Report:
  Section order updated to:
  1. EMA 21/50 Crossover  2. EMA 9/21 Crossover  3. Pullback EMA9
  4. Pullback EMA21  5. Ribbon New Entry  6. Ribbon Cumulative
- GitHub Actions:
  Script is fully compatible. See docstring for workflow YAML snippet to
  commit data files back to repo for persistence between runs.
- 2.3.0
- Delivery factor revamp:
  - New DelivTrend engine: level-aware (low/mid/high), net-change sensitive,
    with dominant-recent-strength override and choppy/indecisive handling.
  - New DeliveryStrength composite (0-100): combines level, 5-day trend, and
    1-day change for a unified delivery quality measure.
  - New PriceConfirm scoring model (no fixed grid): Price, DeliveryStrength,
    and recent delivery change combined into a smooth 4-point scale.
- Date handling:
  - RESOLVED_TRADE_DATE introduced: if run between 08:30–15:45 IST on a
    weekday, the script analyzes the previous trading day; otherwise it
    analyzes the most recent trading day <= today. Used consistently for
    yfinance OHLCV and NSE delivery data.
- 2.2.3
- DelivTrend sparkline: the last 5 days of delivery % are now shown inline
  on the DelivTrend reason line in Conviction Explanation (both HTML + Excel),
  formatted as "D5% → D4% → D3% → D2% → Today%". Shows however many days
  are available (1–5). Net boost/penalty suffix still appended after sparkline.
  Example: Deliv Trend : 5/5 → strong accumulation ✓  [42.1% → 44.3% → 43.8% → 46.2% → 51.0%]
- 2.2.2
- Conviction Explanation cosmetic & labelling fixes:
  DelivTrend labels: "slight distribution" → "slightly declining";
    "sharp distribution" → "sharply declining" (clearer language).
  Factor order in Conviction Explanation (HTML + Excel/text) changed to:
    Price Action → Trend → Volume → Delivery → RSI → MACD → SuperTrend
    (structure first, then participation, then momentum indicators).
  Factor header in HTML report: factor name now rendered in its score colour
    (green/yellow/red) with slightly larger font and letter-spacing for
    improved visual prominence.
- 2.2.1
- Delivery data fixes (two bugs found on first production run):
  Bug 1 (wrong date): fetch_delivery_data() was requesting TODAY's file which is never
    available (NSE publishes ~17:30 IST post-market). Now walks back up to 5 calendar
    days to find the most recently published delivery file. Handles weekends, holidays,
    and pre-market runs transparently.
  Bug 2 (empty fallback on first run): fallback returned 0 symbols because
    nifty100_delivery_history.csv did not exist on first run. Added
    backfill_delivery_history() to fill the cache with up to cfg.delivery_backfill_days
    (default 180) of historical delivery data. Skips weekends; 0.5s polite delay
    between requests.
  New Config field: delivery_backfill_days (int, default 180).
  build_analysis() now calls backfill_delivery_history() before fetch_delivery_data().
- 1.9.4
  - Fixed compute_pivots_support_resistance (4 confirmed bugs — only this
    function changed, all other code identical to v1.9.2):
    * Bug 1 (role reversal): former resistance below price now assigned as
      support; former support above price assigned as resistance. Was causing
      ~40 stocks to show completely empty S/R columns.
    * Bug 2 (WICK_MIN_ATR 0.50 → 0.25): low-volatility stocks (ASIANPAINT,
      HUL, TCS etc.) no longer excluded from swing candidate detection.
    * Bug 3 (MIN_REACTIONS 3 → 2): 252-bar window rarely yields 3
      well-separated touches; 2 is the correct threshold.
    * Bug 4 (BAR_GAP_MIN 10 → 5): packed consolidation zones where multiple
      touches fall within 10 bars were collapsed to 1 accepted touch,
      incorrectly failing the MIN_REACTIONS check.
    * Also: swing detection now uses actual High/Low prices (not Close),
      lookback extended from 120 → 252 bars, two-pass fallback prevents
      blank cells, dead unreachable code in v1.9.2 removed.
- 1.9.2
  - Previous production release.
- 1.7.0
- 2.0.0
  - Added Conviction Scorer: compute_conviction_score() + 6 helper sub-scorers.
    Adds "Conviction Score" (0-100) and "Conviction Explanation" columns to all
    four report tabs. Factors: PriceAction(30) + Volume(20) + Trend(20) +
    RSI(15) + MACD(10) + SuperTrend(5).
  - Fixed pre-existing v1.9.4 corruption: orphaned code fragment between
    compute_pivots_support_resistance and dow_theory_label removed.
- 1.9.4
  - See v1.9.4 changelog (S/R bugs fixed).
- 1.7.0
  - REPLACED NSE Bhavcopy data source with yfinance (yahoo finance).
    * yfinance returns split- and dividend-adjusted OHLCV prices, fixing the
      corporate-action distortion that affected all EMA calculations.
    * Removed: bhavcopy_candidate_urls(), download_bhavcopy_for_date(),
      normalize_bhavcopy(), and the day-by-day download loop in update_history().
    * Added: update_history_yfinance() – downloads adjusted daily OHLCV for the
      entire Nifty-100 universe in one yf.download() batch call, then merges
      incrementally into the existing local CSV cache (same schema: Date, Symbol,
      Open, High, Low, Close, Volume).  The local cache and all downstream
      functions (build_analysis, render_html, render_excel, email) are unchanged.
    * NSE ticker convention: appends ".NS" suffix for yfinance (e.g. RELIANCE→RELIANCE.NS).
    * Removed import: zipfile (no longer needed).
    * Added import: yfinance.
- 1.6.0
  - Previous production release.
- 1.5.0
  - Report 2: Added EMA 9/21 Bullish Crossover scan.
  - Report 3: Added EMA Ribbon scan – EMA 10 > EMA 20 > EMA 50 > EMA 200.
- 1.4.0 and earlier – see prior changelogs.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import math
import os
import sys
import html
import xml.etree.ElementTree as ET
from urllib.parse import quote_plus
from dataclasses import dataclass
from datetime import datetime, time, timedelta, timezone
from pathlib import Path
from typing import List, Optional

import numpy as np
import pandas as pd
import requests
import yfinance as yf
from openpyxl import Workbook
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
from openpyxl.styles import Font, PatternFill

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from google.auth.transport.requests import Request

APP_NAME = "Nifty 100 Daily EMA Cross Reporter"
APP_VERSION = "3.3.9"
APP_RELEASE_DATE = "2026-04-19"
IST = timezone(timedelta(hours=5, minutes=30))
NOW_IST  = datetime.now(IST)

# ── RESOLVED_TRADE_DATE ───────────────────────────────────────────────────────
# If the script is run between 08:30 and 15:45 IST on a weekday (market is open
# / data for today is incomplete), we step back one calendar day so all
# calculations use the previous fully-settled trading day.
# Weekends and holidays are handled downstream by the walk-back in
# fetch_delivery_data() and the yfinance OHLCV filter.
_candidate = NOW_IST.date()
if _candidate.weekday() < 5 and time(8, 30) <= NOW_IST.time() <= time(15, 45):
    _candidate = _candidate - timedelta(days=1)
RESOLVED_TRADE_DATE = _candidate   # date object used by all loaders
TODAY_IST = NOW_IST                # kept for backward-compat log messages

SENTIMENT_ANALYZER = SentimentIntensityAnalyzer()

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
REPORTS_DIR = BASE_DIR / "reports"
LOGS_DIR = BASE_DIR / "logs"
CONFIG_PATH = BASE_DIR / "config.json"
UNIVERSE_PATH = DATA_DIR / "nifty100_constituents.csv"
HISTORY_PATH          = DATA_DIR / "nifty100_ohlcv_history.csv"
DELIVERY_HISTORY_PATH = DATA_DIR / "nifty100_delivery_history.csv"
SIGNAL_LOG_PATH       = DATA_DIR / "nifty100_signal_log.csv"
SIGNAL_ARCHIVE_DIR    = DATA_DIR / "archives"
NIFTY50_CACHE_PATH    = DATA_DIR / "nifty50_members.json"
NIFTYNEXT50_CACHE_PATH= DATA_DIR / "niftynext50_members.json"

GMAIL_SCOPES = ["https://www.googleapis.com/auth/gmail.send"]

for d in [DATA_DIR, REPORTS_DIR, LOGS_DIR]:
    d.mkdir(parents=True, exist_ok=True)

LOG_FILE = LOGS_DIR / f"run_{TODAY_IST.strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[logging.FileHandler(LOG_FILE, encoding="utf-8"), logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger("nifty100_reporter")


@dataclass
class Config:
    email_enabled: bool = False
    smtp_host: str = ""
    smtp_port: int = 587
    smtp_username: str = ""
    smtp_password: str = ""
    mail_from: str = ""
    mail_to: List[str] = None
    news_api_key: str = ""
    history_backfill_days: int = 180
    delivery_backfill_days: int = 180
    signal_log_max_mb: int = 30
    corporate_action_lookahead_days: int = 14
    signal_log_archive_days: int = 90
    recent_news_days: int = 10
    recent_news_max_items: int = 5
    sentiment_pos_cutoff: float = 0.05
    sentiment_neg_cutoff: float = -0.05


def load_config() -> Config:
    if not CONFIG_PATH.exists():
        logger.warning("config.json not found. Email disabled. Using defaults.")
        return Config(mail_to=[])
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        raw = json.load(f)
    cfg = Config(
        email_enabled=bool(raw.get("email_enabled", False)),
        smtp_host=raw.get("smtp_host", ""),
        smtp_port=int(raw.get("smtp_port", 587)),
        smtp_username=raw.get("smtp_username", ""),
        smtp_password=raw.get("smtp_password", ""),
        mail_from=raw.get("mail_from", ""),
        mail_to=raw.get("mail_to", []) or [],
        news_api_key=raw.get("news_api_key", ""),
        history_backfill_days=int(raw.get("history_backfill_days", 180)),
    delivery_backfill_days=int(raw.get("delivery_backfill_days", 180)),
    signal_log_max_mb=int(raw.get("signal_log_max_mb", 30)),
        corporate_action_lookahead_days=int(raw.get("corporate_action_lookahead_days", 14)),
    signal_log_archive_days=int(raw.get("signal_log_archive_days", 90)),
        recent_news_days=int(raw.get("recent_news_days", 10)),
        recent_news_max_items=int(raw.get("recent_news_max_items", 5)),
        sentiment_pos_cutoff=float(raw.get("sentiment_pos_cutoff", 0.05)),
        sentiment_neg_cutoff=float(raw.get("sentiment_neg_cutoff", -0.05)),
    )
    logger.info("Loaded config. Email enabled: %s | Recipients: %s", cfg.email_enabled, len(cfg.mail_to))
    return cfg


def session_with_headers() -> requests.Session:
    s = requests.Session()
    s.headers.update(
        {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
            "Referer": "https://www.nseindia.com/",
            "Connection": "keep-alive",
        }
    )
    return s


import time as _time

def _http_get_with_retry(session, url, timeout=30, max_retries=5, retry_delay=10,
                         label="URL", params=None):
    """
    GET with up to max_retries attempts, retry_delay seconds apart.
    On HTTP 401/403/429 re-hits NSE homepage to refresh cookies before next retry.
    """
    last_exc = None
    for attempt in range(1, max_retries + 1):
        try:
            r = session.get(url, timeout=timeout, params=params)
            if r.status_code in (401, 403, 429) and attempt < max_retries:
                logger.warning(
                    "%s HTTP %s (attempt %d/%d) — refreshing NSE session cookie …",
                    label, r.status_code, attempt, max_retries,
                )
                try:
                    session.get("https://www.nseindia.com/", timeout=20)
                except Exception:
                    pass
                _time.sleep(retry_delay)
                continue
            return r
        except Exception as e:
            last_exc = e
            if attempt < max_retries:
                logger.warning(
                    "%s fetch failed (attempt %d/%d): %s — retrying in %ds …",
                    label, attempt, max_retries, e, retry_delay,
                )
                _time.sleep(retry_delay)
            else:
                logger.error(
                    "%s fetch failed after %d attempts: %s",
                    label, max_retries, e,
                )
    if last_exc:
        raise last_exc
    raise RuntimeError(f"{label}: all {max_retries} attempts exhausted")


def bootstrap_nse_session(s: requests.Session) -> None:
    try:
        r = _http_get_with_retry(s, "https://www.nseindia.com/", timeout=20,
                                 max_retries=5, retry_delay=10, label="NSE session bootstrap")
        logger.info("NSE session bootstrap status: %s", r.status_code)
    except Exception as e:
        logger.warning("NSE session bootstrap failed after retries: %s", e)


def fetch_universe(s: requests.Session) -> pd.DataFrame:
    url = "https://www.niftyindices.com/IndexConstituent/ind_nifty100list.csv"
    logger.info("Downloading Nifty 100 universe from %s", url)
    r = _http_get_with_retry(s, url, timeout=30, max_retries=5, retry_delay=10, label="Nifty100 universe")
    logger.info("Universe download status: %s", r.status_code)
    r.raise_for_status()
    text = r.text
    UNIVERSE_PATH.write_text(text, encoding="utf-8")
    # NiftyIndices CSV sometimes has preamble/disclaimer lines before the
    # actual header row.  Skip lines until we find one containing "Symbol"
    # (case-insensitive), then parse from that point onwards.
    lines = text.splitlines()
    header_idx = 0
    for _i, _line in enumerate(lines):
        if "symbol" in _line.lower() and "," in _line:
            header_idx = _i
            break
    clean_text = "\n".join(lines[header_idx:])
    df = pd.read_csv(io.StringIO(clean_text))
    cols = {c.lower().strip(): c for c in df.columns}
    sym_col = cols.get("symbol") or cols.get("ticker") or list(df.columns)[0]
    name_col = cols.get("company name") or cols.get("company") or sym_col
    out = df[[sym_col, name_col]].copy()
    out.columns = ["Symbol", "CompanyName"]
    out["Symbol"] = out["Symbol"].astype(str).str.strip().str.upper()
    out = out.drop_duplicates(subset=["Symbol"]).reset_index(drop=True)
    logger.info("Universe rows loaded: %s", len(out))
    return out





# ── Local history cache (schema unchanged from v1.6) ─────────────────────────

def fetch_index_membership(s) -> dict:
    """
    Returns {SYMBOL_UPPER: "NIFTY 50" or "Nifty Next 50"} for all members.
    Cache-first: loads from DATA_DIR JSON file; fetches NSE API only when absent.
    Delete cache files manually after NSE quarterly rebalancing.
    Falls back to {} on total failure; Index column shows "Nifty 100" for unmapped.
    """
    import json as _json
    result: dict = {}
    indices = [
        ("NIFTY 50",      "NIFTY 50",       NIFTY50_CACHE_PATH),
        ("Nifty Next 50", "NIFTY NEXT 50",  NIFTYNEXT50_CACHE_PATH),
    ]
    for index_label, index_param, cache_path in indices:
        if cache_path.exists():
            try:
                members = _json.loads(cache_path.read_text(encoding="utf-8"))
                for sym in members:
                    result[sym.upper()] = index_label
                logger.info("Index membership from cache: %s (%d symbols)",
                            index_label, len(members))
                continue
            except Exception as e:
                logger.warning("Index cache corrupt for %s: %s — re-fetching",
                               index_label, e)
        _im_url    = "https://www.nseindia.com/api/equity-stockIndices"
        _im_params = {"index": index_param}
        try:
            r = _http_get_with_retry(s, _im_url, timeout=20, max_retries=5,
                                     retry_delay=10,
                                     params=_im_params,
                                     label=f"NSE index {index_label}")
            r.raise_for_status()
            payload = r.json()
            rows = payload.get("data", [])
            members = [
                row["symbol"].upper()
                for row in rows
                if isinstance(row, dict) and row.get("symbol")
                and not row.get("symbol", "").startswith("NIFTY")
            ]
            if not members:
                logger.warning("Index membership: empty response for %s", index_label)
                continue
            cache_path.write_text(_json.dumps(members, indent=2), encoding="utf-8")
            for sym in members:
                result[sym.upper()] = index_label
            logger.info("Index membership fetched from NSE: %s (%d symbols) cached to %s",
                        index_label, len(members), cache_path.name)
        except Exception as e:
            logger.warning("Index membership fetch failed for %s: %s", index_label, e)
    logger.info("Index membership ready: %d symbols mapped", len(result))
    return result


def load_history() -> pd.DataFrame:
    if not HISTORY_PATH.exists():
        logger.info("History file does not exist yet: %s", HISTORY_PATH)
        return pd.DataFrame(columns=["Date", "Symbol", "Open", "High", "Low", "Close", "Volume"])
    try:
        df = pd.read_csv(HISTORY_PATH, parse_dates=["Date"])
        logger.info("Loaded history rows: %s", len(df))
        return df
    except Exception as e:
        logger.error("Failed to read history file: %s", e)
        return pd.DataFrame(columns=["Date", "Symbol", "Open", "High", "Low", "Close", "Volume"])


def save_history(df: pd.DataFrame) -> None:
    df = df.sort_values(["Symbol", "Date"]).drop_duplicates(["Date", "Symbol"], keep="last")
    df.to_csv(HISTORY_PATH, index=False)
    logger.info("Saved history rows: %s | file: %s", len(df), HISTORY_PATH)


# ── yfinance OHLCV helpers ────────────────────────────────────────────────────

def _nse_to_yf(symbol: str) -> str:
    """Convert a bare NSE symbol (e.g. 'RELIANCE') to its yfinance ticker ('RELIANCE.NS')."""
    return f"{symbol}.NS"


def download_yfinance_batch(
    symbols: List[str],
    start: str,
    end: str,
) -> pd.DataFrame:
    """
    Download adjusted daily OHLCV for *symbols* from yfinance for [start, end).

    Returns a tidy DataFrame with columns:
        Date, Symbol, Open, High, Low, Close, Volume
    where Date is a tz-naive date (no time component) and Close is the
    split- and dividend-adjusted close price.

    yfinance is called once for the whole batch to minimise network round-trips.
    Symbols that return no data (e.g. newly listed, delisted) are silently skipped.
    """
    yf_tickers = [_nse_to_yf(s) for s in symbols]
    logger.info(
        "yfinance batch download: %d symbols | %s → %s", len(yf_tickers), start, end
    )

    try:
        raw = yf.download(
            tickers=yf_tickers,
            start=start,
            end=end,
            interval="1d",
            auto_adjust=True,       # adjusts OHLC for splits & dividends
            progress=False,
            threads=True,
        )
    except Exception as e:
        logger.error("yfinance batch download failed: %s", e)
        return pd.DataFrame(columns=["Date", "Symbol", "Open", "High", "Low", "Close", "Volume"])

    if raw is None or raw.empty:
        logger.warning("yfinance returned empty DataFrame for batch download.")
        return pd.DataFrame(columns=["Date", "Symbol", "Open", "High", "Low", "Close", "Volume"])

    logger.info("yfinance raw shape: %s", raw.shape)

    # yfinance returns a MultiIndex columns frame when >1 ticker is requested:
    #   level-0 = field (Open, High, Low, Close, Volume)
    #   level-1 = ticker (RELIANCE.NS, …)
    # When exactly 1 ticker is requested it returns a flat frame – handle both.

    records = []

    if isinstance(raw.columns, pd.MultiIndex):
        fields = ["Open", "High", "Low", "Close", "Volume"]
        for yf_ticker, nse_sym in zip(yf_tickers, symbols):
            try:
                subset = raw.xs(yf_ticker, axis=1, level=1)[fields].copy()
            except KeyError:
                logger.debug("No data returned for %s", yf_ticker)
                continue
            subset = subset.dropna(subset=["Close"])
            subset = subset.reset_index()          # brings Date back as a column
            subset["Symbol"] = nse_sym
            subset["Date"] = pd.to_datetime(subset["Date"]).dt.tz_localize(None).dt.normalize()
            for col in ["Open", "High", "Low", "Close", "Volume"]:
                subset[col] = pd.to_numeric(subset[col], errors="coerce")
            subset = subset.dropna(subset=["Open", "High", "Low", "Close", "Volume"])
            records.append(subset[["Date", "Symbol", "Open", "High", "Low", "Close", "Volume"]])
    else:
        # Single-ticker flat frame
        sym = symbols[0]
        subset = raw[["Open", "High", "Low", "Close", "Volume"]].copy().dropna(subset=["Close"])
        subset = subset.reset_index()
        subset["Symbol"] = sym
        subset["Date"] = pd.to_datetime(subset["Date"]).dt.tz_localize(None).dt.normalize()
        for col in ["Open", "High", "Low", "Close", "Volume"]:
            subset[col] = pd.to_numeric(subset[col], errors="coerce")
        subset = subset.dropna(subset=["Open", "High", "Low", "Close", "Volume"])
        records.append(subset[["Date", "Symbol", "Open", "High", "Low", "Close", "Volume"]])

    if not records:
        logger.warning("yfinance: no usable rows after parsing.")
        return pd.DataFrame(columns=["Date", "Symbol", "Open", "High", "Low", "Close", "Volume"])

    out = pd.concat(records, ignore_index=True)
    out["Volume"] = out["Volume"].astype("int64", errors="ignore")

    # ── RESOLVED_TRADE_DATE filter ────────────────────────────────────────────
    # Strip any partial intraday candle for today when script runs pre-close
    # (08:30–15:45 IST). RESOLVED_TRADE_DATE is already set to the correct
    # last fully-settled trading day at startup.
    if "Date" in out.columns:
        try:
            out["_d"] = pd.to_datetime(out["Date"]).dt.date
            out = out[out["_d"] <= RESOLVED_TRADE_DATE].drop(columns=["_d"])
            out = out.reset_index(drop=True)
        except Exception as _exc:
            logger.debug("RESOLVED_TRADE_DATE filter skipped: %s", _exc)

    logger.info("yfinance parsed rows after date-filter: %d", len(out))
    return out



# ── NSE Delivery Bhavcopy helpers ────────────────────────────────────────────

DELIVERY_COLS_MAP = {
    # NSE column name variations → normalised name
    "symbol":       "Symbol",
    "series":       "Series",
    "deliv_qty":    "DelivQty",
    "deliv_per":    "DelivPct",
    "deliveryqty":  "DelivQty",
    "deliveryper":  "DelivPct",
    "%dly_qt_to_traded_qty": "DelivPct",
}

def fetch_delivery_data(
    s: requests.Session,
    trade_date: datetime,
) -> tuple:
    """
    Download the NSE CM bhavcopy full-delivery CSV for the most recent
    available trading day on or before *trade_date* and return
    (delivery_dict, fallback_used) where:

      delivery_dict : {SYMBOL: delivery_pct_float}  — EQ series only.
      fallback_used : True  → file unavailable; values are the last-5-day
                              average from the local cache.
                    : False → file successfully downloaded.

    Date walk-back logic
    --------------------
    NSE publishes the delivery file post-market (~17:30 IST).  Running the
    script pre-market (or on a holiday) means today's file does not exist
    yet.  The function tries dates: today, today-1, today-2 … today-6 (skipping
    Saturdays/Sundays) until it gets HTTP 200 with content > 500 bytes.
    This transparently handles weekends, public holidays, and pre-market runs.

    NSE URL pattern:
      https://archives.nseindia.com/products/content/sec_bhavdata_full_DDMMYYYY.csv
    """
    raw_text     = None
    used_date    = None

    for days_back in range(0, 8):          # 0 = today; walk back up to 7 days
        candidate = trade_date - timedelta(days=days_back)
        if candidate.weekday() >= 5:       # skip Saturday (5) and Sunday (6)
            continue
        date_str = candidate.strftime("%d%m%Y")
        url = (
            f"https://archives.nseindia.com/products/content/"
            f"sec_bhavdata_full_{date_str}.csv"
        )
        logger.info("Delivery fetch attempt [%d/7] — URL: %s", days_back, url)
        try:
            r = _http_get_with_retry(s, url, timeout=30, max_retries=5, retry_delay=10, label=f"Delivery data {url[-20:]}")
            if r.status_code == 200 and len(r.content) > 500:
                raw_text   = r.text
                used_date  = candidate
                logger.info(
                    "Delivery fetch SUCCESS — date: %s | size: %d bytes | URL: %s",
                    candidate.strftime("%Y-%m-%d"), len(r.content), url,
                )
                break
            else:
                logger.info(
                    "Delivery fetch FAILED for %s — HTTP %s — URL: %s",
                    candidate.strftime("%Y-%m-%d"), r.status_code, url,
                )
        except Exception as exc:
            logger.warning("Delivery fetch ERROR — date: %s | URL: %s | error: %s", date_str, url, exc)

    if raw_text is None:
        logger.warning(
            "No delivery file found in last 7 calendar days. Using fallback."
        )

    # ── Parse the delivery file ───────────────────────────────────────────────
    today_dict: dict = {}
    if raw_text and used_date:
        try:
            df_raw = pd.read_csv(io.StringIO(raw_text))
            df_raw.columns = [c.strip().lower().replace(" ", "_") for c in df_raw.columns]
            rename = {}
            for col in df_raw.columns:
                for pattern, norm in DELIVERY_COLS_MAP.items():
                    if pattern in col:
                        rename[col] = norm
                        break
            df_raw.rename(columns=rename, inplace=True)

            if "Symbol" not in df_raw.columns or "DelivPct" not in df_raw.columns:
                logger.warning(
                    "Delivery CSV missing expected columns. Cols: %s", list(df_raw.columns)
                )
                raw_text = None
            else:
                if "Series" in df_raw.columns:
                    df_raw = df_raw[
                        df_raw["Series"].astype(str).str.strip().str.upper() == "EQ"
                    ]
                df_raw["Symbol"]   = df_raw["Symbol"].astype(str).str.strip().str.upper()
                df_raw["DelivPct"] = pd.to_numeric(df_raw["DelivPct"], errors="coerce")
                df_raw = df_raw.dropna(subset=["DelivPct"])
                today_dict = dict(zip(df_raw["Symbol"], df_raw["DelivPct"]))

                # ── Append / update local delivery cache ─────────────────────
                cache_rows          = df_raw[["Symbol", "DelivPct"]].copy()
                cache_rows["Date"]  = used_date.strftime("%Y-%m-%d")
                cache_rows          = cache_rows[["Date", "Symbol", "DelivPct"]]

                if DELIVERY_HISTORY_PATH.exists():
                    existing = pd.read_csv(DELIVERY_HISTORY_PATH, parse_dates=["Date"])
                    combined = pd.concat([existing, cache_rows], ignore_index=True)
                else:
                    combined = cache_rows.copy()

                # Primary guard: ensure single row per (Date, Symbol)
                combined["Date"] = pd.to_datetime(combined["Date"])
                combined = combined.drop_duplicates(subset=["Date", "Symbol"], keep="last")

                # Optional prune window (keep last ~45 days of rows)
                try:
                    cutoff = combined["Date"].max() - pd.Timedelta(days=45)
                    combined = combined[combined["Date"] >= cutoff]
                except Exception:
                    pass

                combined = combined.sort_values(["Date", "Symbol"]).reset_index(drop=True)
                combined.to_csv(DELIVERY_HISTORY_PATH, index=False)

                # Extra safety repair: remove any residual duplicates on disk
                try:
                    _rep = pd.read_csv(DELIVERY_HISTORY_PATH, parse_dates=["Date"])
                    _before = len(_rep)
                    _rep = _rep.sort_values(["Date", "Symbol", "DelivPct"], ascending=[True, True, False])
                    _rep = _rep.drop_duplicates(subset=["Date", "Symbol"], keep="first")
                    if len(_rep) < _before:
                        _rep.to_csv(DELIVERY_HISTORY_PATH, index=False)
                        logger.info(
                            "Delivery cache dedup repair: removed %d duplicate rows.",
                            _before - len(_rep),
                        )
                except Exception as _e:
                    logger.warning("Delivery cache dedup repair failed: %s", _e)

                logger.info(
                    "Delivery cache updated: %d symbols for %s (rows in cache: %d)",
                    len(today_dict), used_date.strftime("%Y-%m-%d"),
                    len(combined),
                )
        except Exception as exc:
            logger.warning("Delivery CSV parse error: %s", exc)
            today_dict = {}

    # ── Fallback: 5-day average from cache ───────────────────────────────────
    if not today_dict:
        logger.warning("Using delivery fallback: 5-day avg from cache.")
        if DELIVERY_HISTORY_PATH.exists():
            try:
                hist_d = pd.read_csv(DELIVERY_HISTORY_PATH, parse_dates=["Date"])
                hist_d["DelivPct"] = pd.to_numeric(hist_d["DelivPct"], errors="coerce")
                hist_d = hist_d.dropna(subset=["DelivPct"])
                last5_dates = sorted(hist_d["Date"].unique())[-5:]
                hist_d      = hist_d[hist_d["Date"].isin(last5_dates)]
                avg_df      = hist_d.groupby("Symbol")["DelivPct"].mean()
                today_dict  = avg_df.to_dict()
                logger.info(
                    "Fallback delivery: %d symbols from %d cached dates.",
                    len(today_dict), len(last5_dates),
                )
            except Exception as exc:
                logger.warning("Delivery cache read error: %s", exc)
        return today_dict, True   # fallback_used = True

    return today_dict, False      # fallback_used = False


def get_delivery_series(
    symbol: str,
    trade_date: datetime,
    delivery_dict: dict,
    fallback_used: bool,
) -> tuple:
    """
    Build a 5-element delivery % series for *symbol* for use in _conv_delivery().

    Returns (series: pd.Series or None, today_pct: float or None, fallback_used: bool).

    The series covers [day-4 .. today] — 5 trading days — from the delivery cache.
    Today's value is spliced in from *delivery_dict* (which may itself be a fallback).
    """
    today_pct = delivery_dict.get(symbol.upper())
    if today_pct is None:
        return None, None, fallback_used

    if not DELIVERY_HISTORY_PATH.exists():
        # Only today's data available — return a single-element series
        return pd.Series([today_pct]), today_pct, fallback_used

    try:
        hist_d = pd.read_csv(DELIVERY_HISTORY_PATH, parse_dates=["Date"])
        sym_hist = hist_d[hist_d["Symbol"].str.upper() == symbol.upper()].copy()
        sym_hist["DelivPct"] = pd.to_numeric(sym_hist["DelivPct"], errors="coerce")
        sym_hist = sym_hist.dropna(subset=["DelivPct"])
        # If duplicates exist for a given Date, keep the highest DelivPct
        sym_hist = sym_hist.sort_values(["Date", "DelivPct"], ascending=[True, False])
        sym_hist = sym_hist.drop_duplicates(subset=["Date"], keep="first")

        # Normalise trade_date to a plain date object so the strict-less-than
        # comparison works regardless of whether a datetime or date was passed in.
        # Without this, when trade_date is a datetime with time=00:00:00 the
        # comparison sym_hist["Date"].dt.date < trade_date.date() can silently
        # include today's row (already in the cache), causing today_pct to
        # appear twice in the sparkline.
        if hasattr(trade_date, "date") and callable(trade_date.date):
            _td = trade_date.date()   # datetime → date
        else:
            _td = trade_date          # already a date object

        if fallback_used:
            # Fallback today_pct is a synthetic 5-day average — NOT a real delivery value.
            # Return the last 5 actual cache rows directly; do NOT append the fake average.
            real_vals = list(sym_hist.tail(5)["DelivPct"].values)
            real_today = real_vals[-1] if real_vals else today_pct
            return pd.Series(real_vals, dtype=float), real_today, fallback_used

        # today_pct is a real download for used_date (may be earlier than _td on weekends).
        # Exclude used_date from prior to avoid double-counting.
        latest_cache_date = sym_hist["Date"].dt.date.max() if not sym_hist.empty else None
        if latest_cache_date is not None and sym_hist[sym_hist["Date"].dt.date == latest_cache_date]["DelivPct"].iloc[0] == today_pct:
            prior = sym_hist[sym_hist["Date"].dt.date < latest_cache_date].tail(4)
        else:
            prior = sym_hist[sym_hist["Date"].dt.date < _td].tail(4)
        vals  = list(prior["DelivPct"].values) + [today_pct]
        return pd.Series(vals, dtype=float), today_pct, fallback_used
    except Exception:
        return pd.Series([today_pct]), today_pct, fallback_used

# ═══════════════════════════════════════════════════════════════════════════════
#  MARKET HEALTH DASHBOARD — v3.3.0
#  Components: Trend(35%) + Breadth(25%) + Momentum(20%) + Volatility(10%) + RS(10%)
#  Score: -100 (bearish) to +100 (bullish). Zero changes to existing logic.
# ═══════════════════════════════════════════════════════════════════════════════

MH_VIX_CACHE         = DATA_DIR / "mh_vix_cache.csv"
MH_BREADTH_EMA_CACHE = DATA_DIR / "mh_breadth_ema_cache.csv"
MH_HISTORY_DAYS      = 520


def _mh_clamp(v, lo=-100.0, hi=100.0):
    return max(lo, min(hi, float(v)))


def _mh_emoji(score):
    if score >= 60:   return "&#x1F7E2;"   # green circle
    if score >= 20:   return "&#x1F7E1;"   # yellow circle
    if score >= -20:  return "&#x1F7E0;"   # orange circle
    if score >= -60:  return "&#x1F534;"   # red circle
    return "&#x26D4;"                       # no entry


def _mh_date_chunks(start, end, chunk_days=90):
    cur = start
    while cur <= end:
        nxt = min(cur + pd.Timedelta(days=chunk_days - 1), end)
        yield cur.strftime("%d-%m-%Y"), nxt.strftime("%d-%m-%Y")
        cur = nxt + pd.Timedelta(days=1)


def fetch_index_ohlcv(s, index_name, days=None):
    """Fetch index OHLCV from 2020-10-20.
    Primary  : niftyindices.com POST API (confirmed working, no bot protection).
    Fallback : yfinance interval=1d.
    Cache-first; incremental on subsequent runs."""
    import json as _json
    HISTORY_START = pd.Timestamp("2020-10-20")
    end_dt        = pd.Timestamp(TODAY_IST.date())
    cache_key     = index_name.upper().replace(" ", "_")
    cache_path    = DATA_DIR / ("mh_idx_" + cache_key + ".csv")

    # ── Cache check ──────────────────────────────────────────────────────────
    cached_df = pd.DataFrame()
    if cache_path.exists():
        try:
            cached_df = pd.read_csv(cache_path, parse_dates=["Date"])
            cached_df = cached_df.drop_duplicates("Date").sort_values("Date")
            if cached_df["Date"].max() >= end_dt - pd.Timedelta(days=3):
                logger.info("Index OHLCV cache fresh for %s (%d rows)", index_name, len(cached_df))
                return cached_df
            fetch_from = cached_df["Date"].max() + pd.Timedelta(days=1)
        except Exception as e:
            logger.warning("Index OHLCV cache corrupt for %s: %s — full re-fetch", index_name, e)
            fetch_from = HISTORY_START
    else:
        fetch_from = HISTORY_START

    # ── Primary: niftyindices.com POST API ───────────────────────────────────
    # Confirmed working (test v4): cinfo-wrapped JSON, yyyy-MM-dd dates,
    # returns fields: INDEX_NAME, HistoricalDate, OPEN, HIGH, LOW, CLOSE
    NI_URL = "https://www.niftyindices.com/Backpage.aspx/getHistoricaldatatabletoString"
    NI_HDR = {
        "User-Agent":     ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                           "AppleWebKit/537.36 (KHTML, like Gecko) "
                           "Chrome/124.0.0.0 Safari/537.36"),
        "Accept":         "application/json, text/plain, */*",
        "Content-Type":   "application/json;charset=UTF-8",
        "Referer":        "https://www.niftyindices.com/",
        "X-Requested-With": "XMLHttpRequest",
    }
    # niftyindices index name mapping
    _NI_NAME = {
        "NIFTY 50":      "NIFTY 50",
        "NIFTY NEXT 50": "Nifty Next 50",
    }
    ni_name = _NI_NAME.get(index_name.upper(), index_name)
    ni_sess = requests.Session()
    try:
        ni_sess.get("https://www.niftyindices.com/", timeout=15, headers={
            "User-Agent": NI_HDR["User-Agent"],
            "Accept-Language": "en-US,en;q=0.9",
        })
    except Exception as e:
        logger.warning("niftyindices session prime failed: %s", e)

    ni_rows = []
    cur = fetch_from
    while cur <= end_dt:
        nxt     = min(cur + pd.Timedelta(days=59), end_dt)
        s_str   = cur.strftime("%Y-%m-%d")
        e_str   = nxt.strftime("%Y-%m-%d")
        inner   = _json.dumps({"name": ni_name, "startDate": s_str,
                                "endDate": e_str, "indexName": ni_name})
        payload = {"cinfo": inner}
        try:
            r = ni_sess.post(NI_URL, json=payload, headers=NI_HDR, timeout=20)
            r.raise_for_status()
            outer = r.json().get("d", "")
            if outer and outer != "[]":
                for row in _json.loads(outer):
                    try:
                        ni_rows.append({
                            "Date":   pd.to_datetime(row["HistoricalDate"], format="%d %b %Y"),
                            "Open":   float(row.get("OPEN",  0) or 0),
                            "High":   float(row.get("HIGH",  0) or 0),
                            "Low":    float(row.get("LOW",   0) or 0),
                            "Close":  float(row.get("CLOSE", 0) or 0),
                            "Volume": 0.0,
                        })
                    except Exception:
                        pass
        except Exception as e:
            logger.warning("niftyindices failed %s %s-%s: %s", index_name, s_str, e_str, e)
        cur = nxt + pd.Timedelta(days=1)

    if ni_rows:
        new_df   = pd.DataFrame(ni_rows).drop_duplicates("Date").sort_values("Date")
        combined = pd.concat([cached_df, new_df], ignore_index=True)
        combined = combined.drop_duplicates("Date").sort_values("Date")
        combined.to_csv(cache_path, index=False)
        logger.info("Index OHLCV via niftyindices for %s: %d rows (+%d)",
                    index_name, len(combined), len(ni_rows))
        return combined

    # ── Fallback: yfinance ───────────────────────────────────────────────────
    _YF_TICKERS = {
        "NIFTY 50":      ["^NSEI"],
        "NIFTY NEXT 50": ["JUNIORBEES.NS", "NIFTYNXT50.NS"],
    }
    yf_tickers = _YF_TICKERS.get(index_name.upper(), [])
    if yf_tickers:
        import yfinance as _yf
        yf_start = (fetch_from if not cached_df.empty else HISTORY_START).strftime("%Y-%m-%d")
        yf_end   = (end_dt + pd.Timedelta(days=1)).strftime("%Y-%m-%d")
        for ticker in yf_tickers:
            try:
                _raw = _yf.download(ticker, start=yf_start, end=yf_end,
                                    interval="1d", auto_adjust=True,
                                    progress=False, ignore_tz=True)
                if _raw.empty:
                    logger.warning("yfinance empty for %s (%s)", index_name, ticker)
                    continue
                _df = _raw.reset_index()
                _df.columns = [c[0] if isinstance(c, tuple) else c for c in _df.columns]
                _df = _df.rename(columns={"index": "Date", "Datetime": "Date"})
                _df["Date"] = pd.to_datetime(_df["Date"]).dt.normalize()
                for mc in ["Open", "High", "Low", "Volume"]:
                    if mc not in _df.columns:
                        _df[mc] = _df["Close"]
                _df = _df[["Date","Open","High","Low","Close","Volume"]].dropna(subset=["Close"])
                combined = pd.concat([cached_df, _df], ignore_index=True)
                combined = combined.drop_duplicates("Date").sort_values("Date")
                combined.to_csv(cache_path, index=False)
                logger.info("Index OHLCV via yfinance (%s) for %s: %d rows",
                            ticker, index_name, len(combined))
                return combined
            except Exception as e:
                logger.warning("yfinance failed %s for %s: %s", ticker, index_name, e)

    logger.warning("All sources failed for index OHLCV: %s", index_name)
    return cached_df if not cached_df.empty else pd.DataFrame(
        columns=["Date","Open","High","Low","Close","Volume"])


def fetch_vix_history(s, days=90):
    """Fetch India VIX.
    Primary  : niftyindices.com POST API (name='India VIX').
    Fallback : yfinance ^INDIAVIX.
    Cache-first, incremental, 60-day chunks."""
    import json as _json
    end_dt    = pd.Timestamp(TODAY_IST.date())
    start_dt  = end_dt - pd.Timedelta(days=max(days, 90))
    cached_df = pd.DataFrame()
    if MH_VIX_CACHE.exists():
        try:
            cached_df = pd.read_csv(MH_VIX_CACHE, parse_dates=["Date"])
            cached_df = cached_df.drop_duplicates("Date").sort_values("Date")
            if cached_df["Date"].max() >= end_dt - pd.Timedelta(days=3):
                logger.info("VIX cache fresh (%d rows)", len(cached_df))
                return cached_df
            fetch_from = cached_df["Date"].max() + pd.Timedelta(days=1)
        except Exception as e:
            logger.warning("VIX cache corrupt: %s — re-fetch", e)
            fetch_from = start_dt
    else:
        fetch_from = start_dt

    NI_URL = "https://www.niftyindices.com/Backpage.aspx/getHistoricaldatatabletoString"
    NI_HDR = {
        "User-Agent":     ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                           "AppleWebKit/537.36 (KHTML, like Gecko) "
                           "Chrome/124.0.0.0 Safari/537.36"),
        "Accept":         "application/json, text/plain, */*",
        "Content-Type":   "application/json;charset=UTF-8",
        "Referer":        "https://www.niftyindices.com/",
        "X-Requested-With": "XMLHttpRequest",
    }
    ni_sess = requests.Session()
    try:
        ni_sess.get("https://www.niftyindices.com/", timeout=15, headers={
            "User-Agent": NI_HDR["User-Agent"],
            "Accept-Language": "en-US,en;q=0.9",
        })
    except Exception as e:
        logger.warning("niftyindices VIX session prime failed: %s", e)

    ni_rows = []
    cur = fetch_from
    while cur <= end_dt:
        nxt   = min(cur + pd.Timedelta(days=59), end_dt)
        s_str = cur.strftime("%Y-%m-%d")
        e_str = nxt.strftime("%Y-%m-%d")
        inner   = _json.dumps({"name": "India VIX", "startDate": s_str,
                                "endDate": e_str, "indexName": "India VIX"})
        payload = {"cinfo": inner}
        try:
            r = ni_sess.post(NI_URL, json=payload, headers=NI_HDR, timeout=20)
            r.raise_for_status()
            outer = r.json().get("d", "")
            if outer and outer != "[]":
                for row in _json.loads(outer):
                    try:
                        ni_rows.append({
                            "Date": pd.to_datetime(row["HistoricalDate"], format="%d %b %Y"),
                            "VIX":  float(row.get("CLOSE", row.get("EOD_INDEX_VALUES", 0)) or 0),
                        })
                    except Exception:
                        pass
        except Exception as e:
            logger.warning("niftyindices VIX failed %s-%s: %s", s_str, e_str, e)
        cur = nxt + pd.Timedelta(days=1)

    if ni_rows:
        new_df   = pd.DataFrame(ni_rows).drop_duplicates("Date").sort_values("Date")
        combined = pd.concat([cached_df, new_df], ignore_index=True)
        combined = combined.drop_duplicates("Date").sort_values("Date")
        combined.to_csv(MH_VIX_CACHE, index=False)
        logger.info("VIX via niftyindices: %d rows (+%d)", len(combined), len(ni_rows))
        return combined

    # ── yfinance fallback ────────────────────────────────────────────────────
    try:
        import yfinance as _yf
        yf_start = (fetch_from if not cached_df.empty else start_dt).strftime("%Y-%m-%d")
        yf_end   = (end_dt + pd.Timedelta(days=1)).strftime("%Y-%m-%d")
        _vdf = _yf.download("^INDIAVIX", start=yf_start, end=yf_end,
                             interval="1d", auto_adjust=True, progress=False, ignore_tz=True)
        if not _vdf.empty:
            _vdf = _vdf.reset_index()
            _vdf.columns = [c[0] if isinstance(c, tuple) else c for c in _vdf.columns]
            _vdf = _vdf.rename(columns={"index": "Date", "Datetime": "Date", "Close": "VIX"})
            _vdf["Date"] = pd.to_datetime(_vdf["Date"]).dt.normalize()
            _vdf = _vdf[["Date", "VIX"]].dropna()
            combined = pd.concat([cached_df, _vdf], ignore_index=True)
            combined = combined.drop_duplicates("Date").sort_values("Date")
            combined.to_csv(MH_VIX_CACHE, index=False)
            logger.info("VIX via yfinance (^INDIAVIX): %d rows", len(combined))
            return combined
    except Exception as e:
        logger.warning("yfinance VIX failed: %s", e)

    return cached_df if not cached_df.empty else pd.DataFrame(columns=["Date", "VIX"])


def _mh_ema(series, span):
    return series.ewm(span=span, adjust=False).mean()


def _compute_hh_hl(close, window=20):
    """Returns (int_signal, plain_english_label) tuple."""
    if len(close) < window:
        return 0, "Not enough data to assess price structure"
    sl    = close.iloc[-window:]
    highs = sl.rolling(5, center=True).max().dropna().values
    lows  = sl.rolling(5, center=True).min().dropna().values
    if len(highs) < 2 or len(lows) < 2:
        return 0, "Not enough swing points to assess structure"
    hh = all(highs[i] <= highs[i+1] for i in range(len(highs)-1))
    hl = all(lows[i]  <= lows[i+1]  for i in range(len(lows)-1))
    ll = all(highs[i] >= highs[i+1] for i in range(len(highs)-1))
    lh = all(lows[i]  >= lows[i+1]  for i in range(len(lows)-1))
    if hh and hl:
        return  1, "Rising highs and rising lows — textbook uptrend, buyers in control"
    if ll and lh:
        return -1, "Falling highs and falling lows — textbook downtrend, sellers in control"
    # Mixed — differentiate the sub-cases
    if hh and not hl:
        return 0, "Mixed — highs pushing up but lows not holding (support weakening, bullish lean)"
    if hl and not hh:
        return 0, "Mixed — lows holding firm but no breakout in highs yet (base building, wait for confirmation)"
    if ll and not lh:
        return 0, "Mixed — lows sliding down with occasional bounces (bearish lean, rallies may not sustain)"
    if lh and not ll:
        return 0, "Mixed — highs being sold off but lows still holding (distribution phase, monitor closely)"
    return 0, "Mixed — no consistent pattern in highs or lows (choppy, stand aside and wait for clarity)"


def compute_trend_score(df):
    """EMA 10/20/50/200 alignment + slopes + HH+HL structure -> [-100, +100]."""
    _empty = {"score": 0, "label": "Insufficient data", "structure": "N/A",
              "price": None, "ema10": None, "ema20": None, "ema50": None, "ema200": None}
    if df.empty or len(df) < 30:
        return _empty
    df = df.sort_values("Date").copy()
    df["E10"]  = _mh_ema(df["Close"], 10)
    df["E20"]  = _mh_ema(df["Close"], 20)
    df["E50"]  = _mh_ema(df["Close"], 50)
    df["E200"] = _mh_ema(df["Close"], 200)
    last = df.iloc[-1]
    p, e10, e20, e50, e200 = last["Close"], last["E10"], last["E20"], last["E50"], last["E200"]
    pts = 0
    if   p > e20 > e50 > e200:            pts, label = 40, "All averages aligned upward — strong uptrend, full participation"
    elif p > e20 > e50 and e10 > e20:     pts, label = 30, "Uptrend confirmed — short-term momentum also pointing up"
    elif p > e20 > e50 and e10 < e20:     pts, label = 15, "Uptrend intact but short-term momentum fading — watch for slowdown"
    elif p < e10 and e10 > e20 > e50:     pts, label = 10, "Healthy pullback within uptrend — trend still up, price taking a breather"
    elif e20 > e50 and p < e20:           pts, label =  0, "Price dipped below 20-day avg but uptrend structure intact — consolidating"
    elif p < e20 < e50 < e200:            pts, label = -40, "Price below all averages — deep downtrend, avoid fresh longs"
    elif p < e20 < e50:                   pts, label = -30, "Price below 20 and 50-day avg — downtrend in place"
    elif e20 < e50:                        pts, label = -20, "20-day avg below 50-day avg — medium-term trend turned bearish"
    else:                                  pts, label =   0, "Averages tangled with no clear direction — market is undecided"
    if len(df) >= 10:
        s50  = (df["E50"].iloc[-1]  - df["E50"].iloc[-6])  / (df["E50"].iloc[-6]  + 1e-9) * 100
        s200 = (df["E200"].iloc[-1] - df["E200"].iloc[-6]) / (df["E200"].iloc[-6] + 1e-9) * 100
        if s50  > 0: pts += 10
        if s200 > 0: pts += 10
    struct, struct_label = _compute_hh_hl(df["Close"])
    if struct ==  1: pts += 15
    if struct == -1: pts -= 15
    return {"score": _mh_clamp(pts), "label": label, "structure": struct_label,
            "price": round(p, 2), "ema10": round(e10, 2),
            "ema20": round(e20, 2), "ema50": round(e50, 2), "ema200": round(e200, 2)}


def update_breadth_ema_cache(stock_history, index_map):
    """Incremental per-stock EMA cache for breadth computation."""
    cached = pd.DataFrame()
    if MH_BREADTH_EMA_CACHE.exists():
        try:
            cached = pd.read_csv(MH_BREADTH_EMA_CACHE, parse_dates=["Date"])
        except Exception as e:
            logger.warning("Breadth EMA cache corrupt: %s - rebuilding", e)
    all_syms   = [s for s in stock_history["Symbol"].unique() if s.upper() in index_map]
    today_rows = []
    for sym in all_syms:
        sh = stock_history[stock_history["Symbol"] == sym].sort_values("Date")
        if sh.empty or len(sh) < 10:
            continue
        sym_cached = cached[cached["Symbol"] == sym] if not cached.empty else pd.DataFrame()
        latest_c   = sym_cached["Date"].max() if not sym_cached.empty else pd.NaT
        last_date  = sh["Date"].max()
        if pd.notna(latest_c) and latest_c >= last_date:
            continue
        e10 = _mh_ema(sh["Close"], 10)
        e20 = _mh_ema(sh["Close"], 20)
        e50 = _mh_ema(sh["Close"], 50)
        mask = sh["Date"] > latest_c if pd.notna(latest_c) else pd.Series([True]*len(sh), index=sh.index)
        for i in sh[mask].index:
            pos = sh.index.get_loc(i)
            today_rows.append({
                "Symbol": sym,
                "Date":   sh.loc[i, "Date"],
                "EMA10":  round(float(e10.iloc[pos]), 4),
                "EMA20":  round(float(e20.iloc[pos]), 4),
                "EMA50":  round(float(e50.iloc[pos]), 4),
                "Close":  round(float(sh.loc[i, "Close"]), 4),
            })
    if today_rows:
        new_df   = pd.DataFrame(today_rows)
        combined = pd.concat([cached, new_df], ignore_index=True)
        combined = combined.drop_duplicates(["Symbol", "Date"]).sort_values(["Symbol", "Date"])
        combined.to_csv(MH_BREADTH_EMA_CACHE, index=False)
        logger.info("Breadth EMA cache: %d total rows (+%d)", len(combined), len(today_rows))
        return combined
    return cached


def compute_breadth_score(breadth_cache, index_map, index_label):
    """% stocks above EMA20 and EMA50 -> score [-100, +100]."""
    members = [s for s, lbl in index_map.items() if lbl == index_label]
    _zero   = {"score": 0, "pct_ema20": 0.0, "pct_ema50": 0.0,
               "above_ema20": 0, "above_ema50": 0, "total": 0}
    if not members or breadth_cache.empty:
        return _zero
    latest = (breadth_cache[breadth_cache["Symbol"].isin(members)]
              .sort_values("Date").groupby("Symbol").last().reset_index())
    total = len(latest)
    if total == 0:
        return _zero
    above20 = int((latest["Close"] > latest["EMA20"]).sum())
    above50 = int((latest["Close"] > latest["EMA50"]).sum())
    pct20   = round(above20 / total * 100, 1)
    pct50   = round(above50 / total * 100, 1)
    s20 = (50 if pct20 > 75 else 30 if pct20 > 60 else 10 if pct20 > 50
           else -10 if pct20 > 40 else -30)
    s50 = (50 if pct50 > 70 else 30 if pct50 > 55 else 10 if pct50 > 45
           else -10 if pct50 > 35 else -30)
    return {"score": _mh_clamp(s20 + s50), "pct_ema20": pct20, "pct_ema50": pct50,
            "above_ema20": above20, "above_ema50": above50, "total": total}


def compute_momentum_score(df):
    """RSI-14 + MACD(12/26/9) on index OHLCV -> score [-100, +100]."""
    _empty = {"score": 0, "rsi": None, "macd_bullish": None,
              "macd_val": None, "signal_val": None, "hist_expanding": None}
    if df.empty or len(df) < 35:
        return _empty
    df    = df.sort_values("Date").copy()
    close = df["Close"]
    delta = close.diff()
    gain  = delta.clip(lower=0).ewm(com=13, adjust=False).mean()
    loss  = (-delta.clip(upper=0)).ewm(com=13, adjust=False).mean()
    rsi   = float(100 - 100 / (1 + gain.iloc[-1] / (loss.iloc[-1] + 1e-9)))
    ema12    = _mh_ema(close, 12)
    ema26    = _mh_ema(close, 26)
    macd_l   = ema12 - ema26
    signal_l = _mh_ema(macd_l, 9)
    hist     = macd_l - signal_l
    macd_val   = float(macd_l.iloc[-1])
    signal_val = float(signal_l.iloc[-1])
    macd_bull  = macd_val > signal_val
    hist_exp   = bool(len(hist) >= 4 and abs(hist.iloc[-1]) > abs(hist.iloc[-3]))
    rsi_pts  = (40 if rsi > 70 else 30 if rsi > 60 else 10 if rsi > 50
                else -10 if rsi > 40 else -30 if rsi > 30 else -40)
    macd_pts = 25 if macd_bull else -25
    if hist_exp:
        macd_pts += 10 if macd_bull else -10
    return {"score": _mh_clamp(rsi_pts + macd_pts), "rsi": round(rsi, 1),
            "macd_bullish": macd_bull, "macd_val": round(macd_val, 2),
            "signal_val": round(signal_val, 2), "hist_expanding": hist_exp}


def compute_volatility_score(vix_df):
    """India VIX -> score [-100, +100]."""
    if vix_df.empty:
        return {"score": 0, "vix": None, "label": "No VIX data"}
    vix_val = float(vix_df.sort_values("Date")["VIX"].iloc[-1])
    if vix_val < 12:    score, label = 100,  "Ultra-low fear (bull-friendly)"
    elif vix_val < 15:  score, label =  50,  "Calm / Normal"
    elif vix_val < 18:  score, label =   0,  "Neutral"
    elif vix_val < 22:  score, label = -50,  "Elevated anxiety"
    else:               score, label = -100, "Risk-off / High uncertainty"
    return {"score": score, "vix": round(vix_val, 2), "label": label}


def compute_rs_score(n50_df, nn50_df, window=20):
    """Nifty Next 50 vs Nifty 50 20-day RS ratio -> score [-100, +100]."""
    _na = {"score": 0, "label": "Insufficient data", "rs_ratio": None,
           "rs_avg": None, "pct_dev": None}
    if n50_df.empty or nn50_df.empty:
        return _na
    n50    = n50_df.sort_values("Date").set_index("Date")["Close"]
    nn50   = nn50_df.sort_values("Date").set_index("Date")["Close"]
    common = n50.index.intersection(nn50.index)
    if len(common) < window + 1:
        return _na
    rs      = nn50[common] / n50[common]
    rs_avg  = float(rs.rolling(window).mean().iloc[-1])
    rs_last = float(rs.iloc[-1])
    pct_dev = (rs_last - rs_avg) / (rs_avg + 1e-9) * 100
    if pct_dev > 1:    score, label = 100,  "Risk-on: NN50 outperforming (broad rally)"
    elif pct_dev < -1: score, label = -100, "Defensive: N50 leading (selective market)"
    else:              score, label =    0, "Neutral RS (parity)"
    return {"score": score, "label": label,
            "rs_ratio": round(rs_last, 4), "rs_avg": round(rs_avg, 4),
            "pct_dev": round(pct_dev, 2)}


def _mh_market_type(score):
    if score >= 60:   return "Strong Bull &#x1F7E2;", "Full size (100%)",      "#1b5e20"
    if score >= 20:   return "Weak Bull &#x1F7E1;",   "Normal (75%)",          "#f57f17"
    if score >= -20:  return "Sideways &#x1F7E0;",    "Reduced (50%)",         "#e65100"
    if score >= -60:  return "Weak Bear &#x1F534;",   "Defensive (25%)",       "#b71c1c"
    return              "Strong Bear &#x26D4;",        "Avoid / Hedge (≤10%)", "#4a0000"


def compute_market_health(s, stock_history, index_map):
    """Orchestrator: returns market health dict for NIFTY 50 and Nifty Next 50."""
    logger.info("Market Health: fetching index OHLCV...")
    n50_df  = fetch_index_ohlcv(s, "NIFTY 50")
    nn50_df = fetch_index_ohlcv(s, "NIFTY NEXT 50")
    logger.info("Market Health: fetching VIX...")
    vix_df  = fetch_vix_history(s)
    logger.info("Market Health: updating breadth EMA cache...")
    breadth_cache = update_breadth_ema_cache(stock_history, index_map)
    rs  = compute_rs_score(n50_df, nn50_df)
    vol = compute_volatility_score(vix_df)
    result = {}
    for idx_label, ohlcv_df in [("NIFTY 50", n50_df), ("Nifty Next 50", nn50_df)]:
        trend    = compute_trend_score(ohlcv_df)
        breadth  = compute_breadth_score(breadth_cache, index_map, idx_label)
        momentum = compute_momentum_score(ohlcv_df)
        raw   = (0.35 * trend["score"]    + 0.25 * breadth["score"] +
                 0.20 * momentum["score"] + 0.10 * vol["score"]     +
                 0.10 * rs["score"])
        score = round(_mh_clamp(raw), 1)
        mtype, possize, colour = _mh_market_type(score)
        result[idx_label] = {
            "score": score, "trend": trend, "breadth": breadth,
            "momentum": momentum, "volatility": vol, "rs": rs,
            "market_type": mtype, "pos_size": possize, "colour": colour,
        }
        logger.info("Market Health [%s]: Score=%.1f | %s", idx_label, score, mtype)
    return result



def render_high_conviction_cards(reports: dict, threshold: float = 75.0) -> str:
    """Render High Conviction stock cards (Score > threshold) across all reports."""
    import pandas as _pd

    def _to_rows(val):
        if val is None:
            return []
        if isinstance(val, _pd.DataFrame):
            rows = val.to_dict(orient="records")
            return [r for r in rows if isinstance(r, dict) and r.get("Symbol")]
        if isinstance(val, list):
            return [r for r in val if isinstance(r, dict) and r.get("Symbol")]
        return []

    seen: dict = {}
    for rpt_val in reports.values():
        for row in _to_rows(rpt_val):
            try:
                score = float(row.get("Conviction Score", 0) or 0)
            except (ValueError, TypeError):
                score = 0.0
            if score <= threshold:
                continue
            sym = str(row.get("Symbol", "")).strip()
            if not sym:
                continue
            prev = float(seen[sym].get("Conviction Score", 0) or 0) if sym in seen else -1
            if score > prev:
                seen[sym] = row

    if not seen:
        return ""

    sorted_rows = sorted(seen.values(),
                         key=lambda r: float(r.get("Conviction Score", 0) or 0),
                         reverse=True)

    def _score_cls(s):
        return "" if s >= 85 else (" mid" if s >= 80 else " low")

    def _fire(s):
        return " \U0001f525\U0001f525" if s >= 90 else (" \U0001f525" if s >= 85 else "")

    def _tags(row):
        parts = []
        idx = str(row.get("Index", "") or "").strip()
        if idx:
            parts.append(f"<span class='hc-card-tag tag-idx'>{idx}</span>")
        mom = str(row.get("Momentum", "") or row.get("Dow Theory Trend", "") or "").strip()
        if mom:
            parts.append(f"<span class='hc-card-tag tag-trend'>{mom.split('|')[0].strip()[:14]}</span>")
        try:
            rsi = float(row.get("RSI14", row.get("RSI", 0)) or 0)
            parts.append(f"<span class='hc-card-tag tag-rsi'>RSI {rsi:.0f}</span>")
        except Exception:
            pass
        if str(row.get("VolumeAbove14Avg", "") or "").strip().lower() in ("yes", "true", "1"):
            parts.append("<span class='hc-card-tag tag-vol'>Vol &#8593;</span>")
        return "".join(parts)

    q = '"'   # helper to embed double-quotes inside f-strings
    cards = []
    for row in sorted_rows:
        sym   = str(row.get("Symbol", "")).strip()
        try:
            score = float(row.get("Conviction Score", 0) or 0)
        except Exception:
            score = 0.0
        sc    = f"{score:.1f}"
        anc   = sym.replace(" ", "_")
        cards.append(
            f"<a class={q}hc-card{q} href={q}#row-{anc}{q} "
            f"onclick={q}var el=document.getElementById('row-{anc}');"
            f"if(el){{el.scrollIntoView({{behavior:'smooth',block:'center'}});}}"
            f"return false;{q}>"
            f"<div class={q}hc-card-symbol{q}>[ {sym} ]</div>"
            f"<div class={q}hc-card-score{_score_cls(score)}{q}>Score: {sc}{_fire(score)}</div>"
            f"<div style={q}margin-bottom:4px{q}>{_tags(row)}</div>"
            f"<span class={q}hc-card-arrow{q}>&#8594; View in report</span>"
            f"</a>"
        )

    count = len(sorted_rows)
    body  = "\n".join(cards)
    pl    = "s" if count != 1 else ""
    return (
        "<div class='hc-section'>"
        "<div class='section-head'>"
        f"<h2>&#x2B50;&nbsp; High Conviction Stocks"
        f"&nbsp;<span style='font-size:.75rem;color:#c8dcf4;font-weight:400'>"
        f"score &gt; {threshold:.0f} &mdash; {count} stock{pl} across all reports"
        f"</span></h2>"
        f"<span class='section-badge'>{count} stocks</span>"
        "</div>"
        f"<div class='hc-grid'>{body}</div>"
        "</div>"
    )


def render_market_health_html(mh):
    """Market Health Dashboard — fully dark-theme, self-contained inline styles."""
    if not mh:
        return ""

    def _bar(score):
        pct = max(0, min(100, int((score + 100) / 2)))
        col = ("#2e7d32" if score >= 60 else "#f9a825" if score >= 20
               else "#e65100" if score >= -20 else "#c62828")
        return (
            '<div style="background:#1a2f50;border-radius:4px;height:7px;width:100%;margin:4px 0">'
            '<div style="background:' + col + ';width:' + str(pct) + '%;height:7px;border-radius:4px"></div></div>'
        )

    def _row(label, score, extra=""):
        em = _mh_emoji(score)
        score_col = ("#4caf75" if score >= 60 else "#ffd966" if score >= 20
                     else "#f5a623" if score >= -20 else "#e05252")
        return (
            "<tr style='border-bottom:1px solid #1a3058'>"
            "<td style='padding:5px 10px;color:#c8dcf4;font-size:12px;width:130px'>" + label + "</td>"
            "<td style='padding:5px 8px;font-weight:700;font-size:13px;text-align:right;color:"
            + score_col + "'>" + ("{:+.0f}".format(score)) + "</td>"
            "<td style='padding:5px 6px;font-size:14px;text-align:center'>" + em + "</td>"
            "<td style='padding:5px 10px;color:#dce8f8;font-size:12px'>" + extra + "</td>"
            "</tr>"
        )

    def _inf(label, value):
        return (
            "<tr style='border-bottom:1px solid #12284a'>"
            "<td style='padding:4px 10px;color:#c8dcf4;font-size:12px;white-space:nowrap;width:130px'>"
            + label + "</td>"
            "<td colspan='3' style='padding:4px 10px;font-size:11px;color:#dce8f8'>"
            + value + "</td></tr>"
        )

    def _pane(idx_label, d):
        badge_bg  = "#1565c0" if idx_label == "NIFTY 50" else "#b84800"
        badge_txt = "N50"     if idx_label == "NIFTY 50" else "NN50"
        tr  = d["trend"];    br  = d["breadth"]
        mo  = d["momentum"]; vol = d["volatility"];  rs = d["rs"]

        ema_str = "N/A"
        if tr["price"] is not None:
            p, e10, e20, e50, e200 = tr["price"], tr["ema10"], tr["ema20"], tr["ema50"], tr["ema200"]
            if   p > e10 > e20 > e50 > e200: ema_str = "P&gt;10&gt;20&gt;50&gt;200 &#x1F7E2;"
            elif p > e20 > e50 and e10 > e20: ema_str = "P&gt;20&gt;50, EMA10 bullish &#x1F7E1;"
            elif p > e20 > e50:               ema_str = "P&gt;20&gt;50 &#x1F7E1;"
            elif p < e20 < e50:               ema_str = "P&lt;20&lt;50 &#x1F534;"
            else:
                ema_str = ("P=" + "{:,.0f}".format(p)
                           + " | EMA20=" + "{:,.0f}".format(e20)
                           + " | EMA50=" + "{:,.0f}".format(e50))

        macd_str = ("&#x2705; Bullish" if mo["macd_bullish"] is True
                    else "&#x274C; Bearish" if mo["macd_bullish"] is False else "N/A")
        if mo["hist_expanding"] and mo["macd_bullish"] is not None:
            macd_str += " (expanding)"
        rsi_str = str(mo["rsi"]) if mo["rsi"] is not None else "N/A"
        vix_str = (str(vol["vix"]) + " — " + vol["label"]) if vol["vix"] is not None else "N/A"

        def _be(pct, hi, lo):
            return "&#x1F7E2;" if pct > hi else ("&#x1F7E1;" if pct > lo else "&#x1F534;")

        b20e = _be(br["pct_ema20"], 65, 50)
        b50e = _be(br["pct_ema50"], 60, 45)
        b_summary  = ("{:.1f}".format(br["pct_ema20"]) + "% &gt;EMA20 &nbsp;|&nbsp; "
                      + "{:.1f}".format(br["pct_ema50"]) + "% &gt;EMA50")
        b20_detail = ("{:.1f}".format(br["pct_ema20"]) + "% ("
                      + str(br["above_ema20"]) + "/" + str(br["total"]) + ") " + b20e)
        b50_detail = ("{:.1f}".format(br["pct_ema50"]) + "% ("
                      + str(br["above_ema50"]) + "/" + str(br["total"]) + ") " + b50e)

        # composite score colour band for header strip
        hdr_col = (d["colour"] if "colour" in d and d["colour"]
                   else ("#1b5e20" if d["score"] >= 60 else "#e65100" if d["score"] >= 0 else "#7f0000"))

        score_rows = (
            _row("Trend",      tr["score"],  tr["label"])
            + _row("Breadth",  br["score"],  b_summary)
            + _row("Momentum", mo["score"],  "RSI " + rsi_str)
            + _row("Volatility", vol["score"], "VIX " + vix_str)
            + _row("Rel Strength", rs["score"], rs["label"])
        )
        detail_rows = (
            _inf("EMA Stack",           ema_str)
            + _inf("Structure",         tr["structure"])
            + _inf("Breadth &gt;EMA20", b20_detail)
            + _inf("Breadth &gt;EMA50", b50_detail)
            + _inf("MACD",              macd_str)
        )

        return (
            # Card container
            '<div style="background:#0c1b33;border:1px solid #1e3a6b;border-radius:12px;'
            'min-width:320px;flex:1;box-shadow:0 4px 16px rgba(0,0,0,.5);overflow:hidden">'
            # Header strip (coloured by market regime)
            '<div style="background:' + hdr_col + ';padding:12px 16px;'
            'display:flex;align-items:center;gap:10px">'
            '<span style="background:' + badge_bg + ';color:#fff;padding:3px 10px;'
            'border-radius:4px;font-size:11px;font-weight:700">' + badge_txt + '</span>'
            '<span style="color:#fff;font-size:14px;font-weight:700">' + idx_label + '</span>'
            '<span style="margin-left:auto;color:#fff;font-size:24px;font-weight:800">'
            + ("{:+.0f}".format(d["score"])) + '</span></div>'
            # Subheader: market type + position size + score bar
            '<div style="background:#0f2040;padding:8px 16px;border-bottom:1px solid #1e3a6b">'
            '<div style="font-size:14px;font-weight:600;color:#dce8f8;margin-bottom:2px">'
            + d["market_type"] + '</div>'
            '<div style="font-size:11px;color:#c8dcf4">Position Size: '
            '<b style="color:#ffd966">' + d["pos_size"] + '</b></div>'
            + _bar(d["score"]) + '</div>'
            # Score rows
            '<div style="padding:4px 0;background:#0c1b33">'
            '<table style="width:100%;border-collapse:collapse">' + score_rows + '</table></div>'
            # Detail rows
            '<div style="border-top:1px solid #1a3058;padding:4px 0;background:#0a1628">'
            '<table style="width:100%;border-collapse:collapse">' + detail_rows + '</table></div>'
            '</div>'
        )

    pane_n50  = _pane("NIFTY 50",      mh["NIFTY 50"])
    pane_nn50 = _pane("Nifty Next 50", mh["Nifty Next 50"])
    return (
        '<div style="margin:24px 0">'
        '<div style="font-size:.95rem;font-weight:700;color:#ffd966;'
        'margin-bottom:14px;padding:10px 16px;letter-spacing:.3px;'
        'background:#0d1e3a;border-left:4px solid #ffd966;border-radius:0 8px 8px 0">'
        '&#x1F4CA;&nbsp; Market Health Dashboard</div>'
        '<div style="display:flex;gap:16px;flex-wrap:wrap;align-items:flex-start">'
        + pane_n50 + pane_nn50
        + '</div></div>'
    )


def render_market_health_excel(wb, mh):
    """Insert Market Health as Sheet 1 in the workbook."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    ws = wb.create_sheet("Market Health", 0)

    def _fill(hex6):
        return PatternFill("solid", fgColor=hex6.lstrip("#"))

    def _score_fill(score):
        if score >= 60:   return _fill("C8E6C9")
        if score >= 20:   return _fill("FFF9C4")
        if score >= -20:  return _fill("FFE0B2")
        return _fill("FFCDD2")

    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"),  bottom=Side(style="thin"))

    def _wc(row, col, value, bold=False, fill=None, align="left", fc="000000", sz=10, wrap=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(bold=bold, color=fc, size=sz)
        c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
        c.border    = thin
        if fill:
            c.fill = fill
        return c

    ws.merge_cells("A1:F1")
    tc = ws.cell(row=1, column=1, value="Market Health Dashboard")
    tc.font      = Font(bold=True, size=14, color="FFFFFF")
    tc.fill      = _fill("1A237E")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for ci, h in enumerate(["Metric", "NIFTY 50 Score", "N50 Detail",
                             "", "Nifty Next 50 Score", "NN50 Detail"], 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font      = Font(bold=True, size=10, color="FFFFFF")
        c.fill      = _fill("3949AB")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin
    ws.row_dimensions[2].height = 20

    metrics = [
        ("Market Score",
         lambda d, tr, br, mo, v, rs: d["score"],
         lambda d, tr, br, mo, v, rs: d["market_type"]),
        ("Position Size",
         lambda d, tr, br, mo, v, rs: d["pos_size"],
         lambda d, tr, br, mo, v, rs: ""),
        ("Trend Score",
         lambda d, tr, br, mo, v, rs: tr["score"],
         lambda d, tr, br, mo, v, rs: tr["label"]),
        ("  EMA Stack",
         lambda d, tr, br, mo, v, rs: (
             "P={:,.0f} | EMA10={:,.0f} | EMA20={:,.0f} | EMA50={:,.0f} | EMA200={:,.0f}".format(
                 tr["price"], tr["ema10"], tr["ema20"], tr["ema50"], tr["ema200"])
             if tr["price"] else "N/A"),
         lambda d, tr, br, mo, v, rs: ""),
        ("  Structure",
         lambda d, tr, br, mo, v, rs: tr["structure"],
         lambda d, tr, br, mo, v, rs: ""),
        ("Breadth Score",
         lambda d, tr, br, mo, v, rs: br["score"],
         lambda d, tr, br, mo, v, rs: "{}% > EMA20  |  {}% > EMA50".format(
             br["pct_ema20"], br["pct_ema50"])),
        ("  % Stocks > EMA20",
         lambda d, tr, br, mo, v, rs: "{}%".format(br["pct_ema20"]),
         lambda d, tr, br, mo, v, rs: "{} of {} stocks".format(br["above_ema20"], br["total"])),
        ("  % Stocks > EMA50",
         lambda d, tr, br, mo, v, rs: "{}%".format(br["pct_ema50"]),
         lambda d, tr, br, mo, v, rs: "{} of {} stocks".format(br["above_ema50"], br["total"])),
        ("Momentum Score",
         lambda d, tr, br, mo, v, rs: mo["score"],
         lambda d, tr, br, mo, v, rs: "RSI={} | MACD={}".format(
             mo["rsi"], "Bullish" if mo["macd_bullish"] else "Bearish")),
        ("Volatility Score",
         lambda d, tr, br, mo, v, rs: v["score"],
         lambda d, tr, br, mo, v, rs: "VIX={} - {}".format(v["vix"], v["label"])),
        ("Rel Strength Score",
         lambda d, tr, br, mo, v, rs: rs["score"],
         lambda d, tr, br, mo, v, rs: rs["label"]),
    ]

    bold_set   = {"Market Score", "Trend Score", "Breadth Score",
                  "Momentum Score", "Volatility Score", "Rel Strength Score"}
    rows_data  = [("NIFTY 50", mh["NIFTY 50"]), ("Nifty Next 50", mh["Nifty Next 50"])]

    for ri, (metric, val_fn, det_fn) in enumerate(metrics, start=3):
        lc = ws.cell(row=ri, column=1, value=metric)
        lc.font      = Font(bold=(metric in bold_set), size=10)
        lc.alignment = Alignment(horizontal="left", vertical="top")
        lc.border    = thin; lc.fill = _fill("FFFFFF")
        sc = ws.cell(row=ri, column=4, value="")
        sc.fill = _fill("FFFFFF"); sc.border = thin
        for col_off, (idx_label, d) in zip([2, 5], rows_data):
            tr, br, mo, vol, rs = (d["trend"], d["breadth"], d["momentum"],
                                    d["volatility"], d["rs"])
            val  = val_fn(d, tr, br, mo, vol, rs)
            det  = det_fn(d, tr, br, mo, vol, rs)
            fill = _score_fill(val) if isinstance(val, (int, float)) else _fill("FFFFFF")
            vc = ws.cell(row=ri, column=col_off, value=val)
            vc.font      = Font(bold=isinstance(val, (int, float)), size=10)
            vc.alignment = Alignment(horizontal="center", vertical="top")
            vc.border    = thin; vc.fill = fill
            dc = ws.cell(row=ri, column=col_off + 1, value=det)
            dc.font      = Font(size=9)
            dc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            dc.border    = thin; dc.fill = _fill("FFFFFF")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 2
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 55
    ws.freeze_panes = "A3"
    logger.info("Market Health Excel sheet written.")


def fetch_corporate_actions(s, symbols: list, lookahead_days: int = 14) -> dict:
    """
    Fetch upcoming corporate actions from NSE India for all given symbols.

    Confirmed endpoint (17-Apr-2026):
      GET /api/corporates-corporateActions
          ?index=equities&from_date=DD-MM-YYYY&to_date=DD-MM-YYYY
    Confirmed response fields: symbol, subject, exDate, recDate

    Returns dict keyed by UPPER symbol:
      "RELIANCE": [{event, kind, ex_date, rec_date, days_away}, ...]
      "PFC":      []   # explicit empty = checked, no events in window
    Returns {} on total failure (non-critical, run continues).
    """
    sym_set  = {sym.upper() for sym in symbols}
    today    = RESOLVED_TRADE_DATE
    cutoff   = today + timedelta(days=lookahead_days)
    from_str = today.strftime("%d-%m-%Y")
    to_str   = cutoff.strftime("%d-%m-%Y")
    result: dict = {}

    def _classify(subject: str) -> str:
        sl = subject.lower()
        if "split" in sl or "sub-division" in sl: return "SPLIT"
        if "bonus" in sl:                          return "BONUS"
        if "buy back" in sl or "buyback" in sl:   return "BUYBACK"
        if "rights" in sl:                         return "RIGHTS"
        if "agm" in sl or "annual general" in sl:  return "AGM"
        if "dividend" in sl:                       return "DIVIDEND"
        if "earnings" in sl or "results" in sl:    return "EARNINGS"
        return "CORPORATE ACTION"

    def _parse_rows(rows: list) -> None:
        for row in rows:
            sym     = str(row.get("symbol", "")).strip().upper()
            if sym not in sym_set:
                continue
            subject = str(row.get("subject", "")).strip()
            ex_raw  = str(row.get("exDate",  "")).strip()
            rec_raw = str(row.get("recDate", "")).strip()
            if not ex_raw or not subject:
                continue
            ex_date = None
            for fmt in ("%d-%b-%Y", "%d-%m-%Y", "%Y-%m-%d"):
                try:
                    ex_date = datetime.strptime(ex_raw, fmt).date()
                    break
                except (ValueError, AttributeError):
                    pass
            if ex_date is None or not (today <= ex_date <= cutoff):
                continue
            rec_date = None
            for fmt in ("%d-%b-%Y", "%d-%m-%Y", "%Y-%m-%d"):
                try:
                    rec_date = datetime.strptime(rec_raw, fmt).date()
                    break
                except (ValueError, AttributeError):
                    pass
            days_away = (ex_date - today).days
            result.setdefault(sym, []).append({
                "event":     subject,
                "kind":      _classify(subject),
                "ex_date":   ex_date.strftime("%d-%b-%Y"),
                "rec_date":  rec_date.strftime("%d-%b-%Y") if rec_date else "-",
                "days_away": days_away,
            })

    bulk_url    = "https://www.nseindia.com/api/corporates-corporateActions"
    bulk_params = {"index": "equities", "from_date": from_str, "to_date": to_str}
    try:
        r = _http_get_with_retry(s, bulk_url, params=bulk_params,
                                  timeout=20, max_retries=5, retry_delay=5,
                                  label="NSE corp-actions bulk")
        r.raise_for_status()
        payload = r.json()
        rows = payload if isinstance(payload, list) else payload.get("data", [])
        _parse_rows(rows)
        logger.info("Corporate actions bulk OK: %d raw rows | %d Nifty100 symbols with events",
                    len(rows), sum(1 for v in result.values() if v))
    except Exception as e:
        logger.warning("Corporate actions bulk fetch failed: %s", e)

    missing = sym_set - set(result.keys())
    if missing:
        per_url = "https://www.nseindia.com/api/corporates-corporateActions"
        fetched_n = 0
        for sym in sorted(missing):
            try:
                r = _http_get_with_retry(s, per_url,
                                          params={"symbol": sym, "series": "EQ"},
                                          timeout=15, max_retries=3, retry_delay=5,
                                          label=f"NSE corp-actions {sym}")
                r.raise_for_status()
                payload = r.json()
                rows = payload if isinstance(payload, list) else payload.get("data", [])
                _parse_rows(rows)
                fetched_n += 1
                _time.sleep(0.3)
            except Exception as e:
                logger.debug("Corp-actions per-symbol %s failed: %s", sym, e)
        if fetched_n:
            logger.info("Corporate actions per-symbol fallback: %d symbols fetched", fetched_n)

    for sym in sym_set:
        result.setdefault(sym, [])

    found_n = sum(1 for v in result.values() if v)
    _syms_with_events = ", ".join(sorted(k for k, v in result.items() if v))
    logger.info(
        "Corporate actions: %d/%d symbols have events (window %s to %s) | Symbols: %s",
        found_n, len(sym_set), from_str, to_str,
        _syms_with_events if _syms_with_events else "none",
    )
    return result


def backfill_delivery_history(
    s: requests.Session,
    cfg: Config,
) -> None:
    """
    One-time (or repair) backfill of the local delivery history cache.

    Called at the start of build_analysis().  It is a no-op when the cache
    already has >= 5 distinct trading dates (meaning a prior run populated it).

    When the cache is missing or sparse it downloads historical delivery
    bhavcopy files day-by-day from
      https://archives.nseindia.com/products/content/sec_bhavdata_full_DDMMYYYY.csv

    going back cfg.delivery_backfill_days calendar days (default 180).

    Design choices
    --------------
    * Skips weekends — NSE never publishes on Sat/Sun.
    * 0.5-second sleep between successful downloads — polite rate limiting.
    * Skips dates already in the cache (incremental / idempotent).
    * Stops on 3 consecutive HTTP errors to avoid hammering NSE on a bad run.
    * Logs a progress line every 10 successful downloads.
    * On failure for a given date (404 / timeout) the date is silently skipped
      (holiday or exchange closure).
    """
    import time

    # ── Check if backfill is needed ───────────────────────────────────────────
    if DELIVERY_HISTORY_PATH.exists():
        try:
            existing = pd.read_csv(DELIVERY_HISTORY_PATH, parse_dates=["Date"])
            n_dates  = existing["Date"].nunique()
            if n_dates >= 5:
                logger.info(
                    "Delivery cache has %d dates — skipping backfill.", n_dates
                )
                return
            logger.info(
                "Delivery cache has only %d dates — running backfill.", n_dates
            )
        except Exception:
            logger.warning("Could not read delivery cache — will recreate via backfill.")
    else:
        logger.info("Delivery cache does not exist — running initial backfill.")

    # ── Build list of candidate dates to try ─────────────────────────────────
    end_date   = RESOLVED_TRADE_DATE - timedelta(days=1)       # day before analysis date
    start_date = TODAY_IST.date() - timedelta(days=cfg.delivery_backfill_days)
    logger.info(
        "Delivery backfill: %s → %s (%d calendar days)",
        start_date, end_date, cfg.delivery_backfill_days,
    )

    # Build existing date set to skip already-cached dates
    existing_dates: set = set()
    if DELIVERY_HISTORY_PATH.exists():
        try:
            ex = pd.read_csv(DELIVERY_HISTORY_PATH, parse_dates=["Date"])
            existing_dates = set(ex["Date"].dt.date.unique())
        except Exception:
            pass

    # Iterate newest → oldest so the most recent data lands in cache first
    current      = end_date
    success_count = 0
    consec_errors = 0
    all_rows: list = []

    while current >= start_date:
        # Skip weekends
        if current.weekday() >= 5:
            current -= timedelta(days=1)
            continue

        # Skip already-cached dates
        if current in existing_dates:
            current -= timedelta(days=1)
            continue

        date_str = current.strftime("%d%m%Y")
        url = (
            f"https://archives.nseindia.com/products/content/"
            f"sec_bhavdata_full_{date_str}.csv"
        )
        try:
            r = _http_get_with_retry(s, url, timeout=30, max_retries=5, retry_delay=10, label=f"Delivery data {url[-20:]}")
            if r.status_code == 200 and len(r.content) > 500:
                consec_errors = 0
                try:
                    df_raw = pd.read_csv(io.StringIO(r.text))
                    df_raw.columns = [
                        c.strip().lower().replace(" ", "_") for c in df_raw.columns
                    ]
                    rename = {}
                    for col in df_raw.columns:
                        for pattern, norm in DELIVERY_COLS_MAP.items():
                            if pattern in col:
                                rename[col] = norm
                                break
                    df_raw.rename(columns=rename, inplace=True)

                    if "Symbol" in df_raw.columns and "DelivPct" in df_raw.columns:
                        if "Series" in df_raw.columns:
                            df_raw = df_raw[
                                df_raw["Series"].astype(str).str.strip().str.upper() == "EQ"
                            ]
                        df_raw["Symbol"]   = (
                            df_raw["Symbol"].astype(str).str.strip().str.upper()
                        )
                        df_raw["DelivPct"] = pd.to_numeric(
                            df_raw["DelivPct"], errors="coerce"
                        )
                        df_raw = df_raw.dropna(subset=["DelivPct"])
                        rows        = df_raw[["Symbol", "DelivPct"]].copy()
                        rows["Date"] = current.strftime("%Y-%m-%d")
                        all_rows.append(rows[["Date", "Symbol", "DelivPct"]])
                        success_count += 1
                        if success_count % 10 == 0:
                            logger.info(
                                "Delivery backfill: %d dates downloaded so far "
                                "(latest: %s)…",
                                success_count, current,
                            )
                        time.sleep(0.5)   # polite delay
                except Exception as parse_exc:
                    logger.debug("Backfill parse error for %s: %s", date_str, parse_exc)
            else:
                # 404 or holiday — silent skip
                consec_errors += 1
                if consec_errors >= 3:
                    logger.debug(
                        "3 consecutive HTTP errors at %s — continuing.", current
                    )
                    consec_errors = 0
        except Exception as req_exc:
            logger.debug("Backfill request error for %s: %s", date_str, req_exc)

        current -= timedelta(days=1)

    # ── Write everything to cache in one shot ────────────────────────────────
    if all_rows:
        new_df = pd.concat(all_rows, ignore_index=True)
        new_df["Date"] = pd.to_datetime(new_df["Date"])

        if DELIVERY_HISTORY_PATH.exists():
            try:
                existing_df = pd.read_csv(DELIVERY_HISTORY_PATH, parse_dates=["Date"])
                combined = pd.concat([existing_df, new_df], ignore_index=True)
            except Exception:
                combined = new_df
        else:
            combined = new_df

        combined = combined.drop_duplicates(subset=["Date", "Symbol"], keep="last")
        combined = combined.sort_values(["Date", "Symbol"]).reset_index(drop=True)
        combined.to_csv(DELIVERY_HISTORY_PATH, index=False)
        logger.info(
            "Delivery backfill complete: %d dates, %d rows written to %s",
            success_count, len(combined), DELIVERY_HISTORY_PATH,
        )
    else:
        logger.warning("Delivery backfill: no data rows collected.")


def update_history(
    _s,                          # requests.Session – kept for API compatibility; not used
    universe: pd.DataFrame,
    cfg: Config,
) -> pd.DataFrame:
    """
    Incrementally refresh the local OHLCV history CSV using yfinance
    (split- and dividend-adjusted prices).

    Strategy
    --------
    • If the local history file is empty / missing, back-fill
      `cfg.history_backfill_days` calendar days.
    • Otherwise, download only the days since the last recorded date
      (inclusive of that date so any late-arriving data is refreshed).
    • New rows are merged into the existing history, deduped by (Date, Symbol),
      and saved back to disk.
    • The function raises RuntimeError if the cache is still empty after the
      download attempt, which preserves the original fail-fast behaviour.
    """
    hist = load_history()
    symbols = list(universe["Symbol"].unique())

    if hist.empty:
        start_dt = TODAY_IST.date() - timedelta(days=cfg.history_backfill_days)
    else:
        last_recorded = pd.to_datetime(hist["Date"]).max().date()
        # Re-download from one day before the last date so any corporate-action
        # re-adjustments that yfinance applies retroactively are captured.
        start_dt = last_recorded - timedelta(days=1)

    # yfinance `end` is exclusive – add one day to include today
    end_dt = TODAY_IST.date() + timedelta(days=1)

    logger.info(
        "yfinance history update: %d symbols | %s → %s",
        len(symbols), start_dt, end_dt,
    )

    new_data = download_yfinance_batch(
        symbols,
        start=start_dt.strftime("%Y-%m-%d"),
        end=end_dt.strftime("%Y-%m-%d"),
    )

    if new_data.empty:
        logger.warning("yfinance returned no new data rows.")
    else:
        logger.info("New rows from yfinance: %d", len(new_data))
        hist = pd.concat([hist, new_data], ignore_index=True)
        hist["Date"] = pd.to_datetime(hist["Date"])
        save_history(hist)

    hist = load_history()
    if hist.empty:
        raise RuntimeError(
            "History file is empty after yfinance update. "
            "Check network connectivity and that yfinance can reach finance.yahoo.com."
        )
    return hist




def ema(s, span):
    # If not enough data to compute SMA seed, fall back to standard ewm
    if len(s) < span:
        return s.ewm(span=span, adjust=False).mean()
    
    # Seed with SMA of first `span` candles — matches TradingView behaviour
    sma_seed = s.iloc[:span].mean()
    result = s.ewm(span=span, adjust=False).mean().copy()
    result.iloc[span - 1] = sma_seed
    return result


def rsi(close, period=14):
    # Warm-up guard: RSI values on the first 2*period bars are statistically
    # unreliable and should not be used for signal detection (Fix 4).
    if len(close) < period * 2:
        return pd.Series([np.nan] * len(close), index=close.index)
    delta = close.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    avg_gain = gain.ewm(alpha=1/period, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1/period, adjust=False).mean()
    rs = avg_gain / avg_loss.replace(0, np.nan)
    return 100 - (100 / (1 + rs))


def macd(close):
    ema12  = ema(close, 12)
    ema26  = ema(close, 26)
    line   = ema12 - ema26
    # Seed signal line with ema() for consistency with ema12/ema26 (Fix 3)
    signal = ema(line, 9)
    hist   = line - signal
    return line, signal, hist


def atr(df, period=10):
    hl = df["High"] - df["Low"]
    hc = (df["High"] - df["Close"].shift()).abs()
    lc = (df["Low"] - df["Close"].shift()).abs()
    tr = pd.concat([hl, hc, lc], axis=1).max(axis=1)
    return tr.ewm(alpha=1/period, adjust=False).mean()


def supertrend(df, period=10, multiplier=3.0):
    a = atr(df, period)
    hl2 = (df["High"] + df["Low"]) / 2
    upperband = hl2 + multiplier * a
    lowerband = hl2 - multiplier * a
    final_upper = upperband.copy()
    final_lower = lowerband.copy()
    trend = pd.Series(index=df.index, dtype="int64")
    st = pd.Series(index=df.index, dtype="float64")
    for i in range(len(df)):
        if i == 0:
            trend.iloc[i] = 1
            st.iloc[i] = lowerband.iloc[i]
            continue
        final_upper.iloc[i] = upperband.iloc[i] if (upperband.iloc[i] < final_upper.iloc[i-1] or df["Close"].iloc[i-1] > final_upper.iloc[i-1]) else final_upper.iloc[i-1]
        final_lower.iloc[i] = lowerband.iloc[i] if (lowerband.iloc[i] > final_lower.iloc[i-1] or df["Close"].iloc[i-1] < final_lower.iloc[i-1]) else final_lower.iloc[i-1]
        if trend.iloc[i-1] == 1:
            trend.iloc[i] = -1 if df["Close"].iloc[i] < final_lower.iloc[i] else 1
        else:
            trend.iloc[i] = 1 if df["Close"].iloc[i] > final_upper.iloc[i] else -1
        st.iloc[i] = final_lower.iloc[i] if trend.iloc[i] == 1 else final_upper.iloc[i]
    return trend, st


def detect_candlestick(df: pd.DataFrame) -> str:
    if len(df) < 3:
        return "None"

    x = df.iloc[-1]
    prev = df.iloc[-2]
    prev2 = df.iloc[-3]

    o, h, l, c = x["Open"], x["High"], x["Low"], x["Close"]
    body = abs(c - o)
    rng = max(h - l, 1e-9)
    upper = h - max(o, c)
    lower = min(o, c) - l

    bullish = c > o
    bearish = c < o

    if body / rng <= 0.1:
        return "Doji"

    if body / rng >= 0.8:
        return "Bullish Marubozu" if bullish else "Bearish Marubozu"

    if body / rng <= 0.3 and upper / rng > 0.2 and lower / rng > 0.2:
        return "Spinning Top"

    if lower > body * 2 and upper <= body:
        # Correct classification: Hammer = same shape at bottom of downtrend
        #                         Hanging Man = same shape at top of uptrend
        # Use _slope_direction on the prior bars (exclude the current candle)
        prior_trend = _slope_direction(df["Close"].iloc[:-1], min_points=5, max_points=10)
        if prior_trend == "down":
            return "Hammer"
        else:
            return "Hanging Man"

    # Shooting Star: long upper wick regardless of candle color (body direction
    # is irrelevant per Nison 1991 — the long upper wick IS the signal)
    # Only valid at the top of an uptrend.
    if upper > body * 2 and lower <= body:
        prior_trend = _slope_direction(df["Close"].iloc[:-1], min_points=5, max_points=10)
        if prior_trend == "up":
            return "Shooting Star"

    prev_o, prev_c = prev["Open"], prev["Close"]
    prev_body = abs(prev_c - prev_o)
    prev_rng = max(prev["High"] - prev["Low"], 1e-9)

    if bullish and prev_c < prev_o and c >= prev_o and o <= prev_c:
        return "Bullish Engulfing"
    if bearish and prev_c > prev_o and o >= prev_c and c <= prev_o:
        return "Bearish Engulfing"

    small_body = body / max(prev_body, 1e-9) < 0.6
    prev_large = prev_body / prev_rng >= 0.5
    if prev_large and small_body:
        if prev_c < prev_o and bullish and (min(o, c) > min(prev_o, prev_c)) and (max(o, c) < max(prev_o, prev_c)):
            return "Bullish Harami"
        if prev_c > prev_o and bearish and (min(o, c) > min(prev_o, prev_c)) and (max(o, c) < max(prev_o, prev_c)):
            return "Bearish Harami"

    if (
        prev2["Close"] < prev2["Open"]
        and abs(prev["Close"] - prev["Open"]) / max(prev["High"] - prev["Low"], 1e-9) < 0.3
        and bullish
        and c > ((prev2["Open"] + prev2["Close"]) / 2)
    ):
        return "Morning Star"

    if (
        prev2["Close"] > prev2["Open"]
        and abs(prev["Close"] - prev["Open"]) / max(prev["High"] - prev["Low"], 1e-9) < 0.3
        and bearish
        and c < ((prev2["Open"] + prev2["Close"]) / 2)
    ):
        return "Evening Star"

    return "None"


def _slope_direction(series: pd.Series, min_points: int = 10, max_points: int = 21, flat_threshold: float = 0.005) -> str:
    if len(series) < min_points:
        return "insufficient"
    window = series.dropna().tail(max_points)
    if len(window) < min_points:
        return "insufficient"
    first = float(window.iloc[0])
    last = float(window.iloc[-1])
    base = max(abs(first), 1.0)
    change = (last - first) / base
    if change > flat_threshold:
        return "up"
    if change < -flat_threshold:
        return "down"
    return "flat"


def compute_rsi_divergence(price: pd.Series, rsi_series: pd.Series) -> str:
    price_dir = _slope_direction(price)
    rsi_dir = _slope_direction(rsi_series)
    if price_dir == "insufficient" or rsi_dir == "insufficient":
        return "Insufficient data"

    if price_dir == "flat" and rsi_dir == "up":
        return "Bullish divergence"
    if price_dir == "down" and rsi_dir in ("flat", "up"):
        return "Bullish divergence"

    if price_dir == "flat" and rsi_dir == "down":
        return "Bearish divergence"
    if price_dir == "up" and rsi_dir in ("flat", "down"):
        return "Bearish divergence"

    return "No divergence"


def _find_swings(prices: pd.Series, lookback: int = 60, pct_tol: float = 0.02):
    """Legacy swing finder on Close prices — used by compute_pivots_support_resistance."""
    window = prices.tail(lookback)
    highs, lows = [], []
    vals = window.values
    idxs = window.index.to_list()
    for i in range(1, len(vals) - 1):
        if vals[i] > vals[i - 1] and vals[i] > vals[i + 1]:
            highs.append((idxs[i], vals[i]))
        if vals[i] < vals[i - 1] and vals[i] < vals[i + 1]:
            lows.append((idxs[i], vals[i]))
    return highs, lows


def _significant_swings_hl(df: pd.DataFrame, lookback: int = 90, atr_mult: float = 1.0):
    """
    ATR-filtered swing detection on actual High/Low prices.
    Only swings exceeding atr_mult * ATR(14) qualify as significant.
    Returns (highs, lows) as lists of (integer_position, price_value).
    """
    window = df.tail(lookback).reset_index(drop=True)
    if len(window) < 14:
        return [], []
    atr_val = float(atr(window, period=14).iloc[-1])
    threshold = atr_mult * atr_val
    highs, lows = [], []
    h_vals = window["High"].values
    l_vals = window["Low"].values
    for i in range(1, len(window) - 1):
        if (h_vals[i] > h_vals[i - 1] + threshold and
                h_vals[i] > h_vals[i + 1] + threshold):
            highs.append((i, h_vals[i]))
        if (l_vals[i] < l_vals[i - 1] - threshold and
                l_vals[i] < l_vals[i + 1] - threshold):
            lows.append((i, l_vals[i]))
    return highs, lows


# ── Phase 2: HMM Regime Detection ────────────────────────────────────────────

def _hmm_regime(df: pd.DataFrame, n_states: int = 3) -> str:
    """
    Hidden Markov Model regime classifier using log-returns and normalised volume.

    States are mapped to regime labels by their mean log-return:
      highest mean  → uptrend
      lowest mean   → downtrend
      middle        → ranging

    Falls back gracefully to the EMA/slope classifier if hmmlearn is not
    installed or if there is insufficient data (< 60 bars).

    Returns one of: "uptrend", "downtrend", "ranging"
    """
    if len(df) < 60:
        return _ema_regime(df)
    try:
        from hmmlearn.hmm import GaussianHMM  # optional dependency
    except ImportError:
        return _ema_regime(df)

    try:
        close  = df["Close"].values.astype(float)
        volume = df["Volume"].values.astype(float)

        log_ret = np.diff(np.log(np.where(close > 0, close, np.nan)))
        log_ret = np.nan_to_num(log_ret, nan=0.0)

        # Normalise volume to [0, 1] to keep features on similar scale
        vol_norm = volume[1:]
        vol_range = vol_norm.max() - vol_norm.min()
        if vol_range > 0:
            vol_norm = (vol_norm - vol_norm.min()) / vol_range
        else:
            vol_norm = np.zeros_like(vol_norm)

        X = np.column_stack([log_ret, vol_norm])

        model = GaussianHMM(
            n_components=n_states,
            covariance_type="diag",
            n_iter=200,
            random_state=42,
            tol=1e-4,
        )
        model.fit(X)
        hidden_states = model.predict(X)

        # Map state index → regime by mean log-return of each state
        state_means = {
            s: float(log_ret[hidden_states == s].mean())
            for s in range(n_states)
            if (hidden_states == s).any()
        }
        if not state_means:
            return _ema_regime(df)

        sorted_states = sorted(state_means, key=lambda s: state_means[s])
        down_state   = sorted_states[0]
        up_state     = sorted_states[-1]
        range_states = sorted_states[1:-1]

        current_state = int(hidden_states[-1])
        if current_state == up_state:
            return "uptrend"
        if current_state == down_state:
            return "downtrend"
        return "ranging"

    except Exception:
        return _ema_regime(df)


def _ema_regime(df: pd.DataFrame) -> str:
    """EMA/slope fallback regime classifier (used when hmmlearn unavailable)."""
    if len(df) < 50:
        return "ranging"
    close      = df["Close"]
    ema50_val  = float(ema(close, 50).iloc[-1])
    ema200_val = float(ema(close, 200).iloc[-1]) if len(df) >= 200 else ema50_val
    last_close = float(close.iloc[-1])
    slope      = _slope_direction(close, min_points=20, max_points=40)
    if last_close > ema50_val > ema200_val and slope == "up":
        return "uptrend"
    if last_close < ema50_val < ema200_val and slope == "down":
        return "downtrend"
    return "ranging"


def _detect_regime(df: pd.DataFrame) -> str:
    """
    Primary regime entry point.
    Uses HMM when hmmlearn is available, falls back to EMA/slope otherwise.
    """
    return _hmm_regime(df)


# ── Pattern scorers ───────────────────────────────────────────────────────────

def _score_double_top(highs, df: pd.DataFrame, regime: str, vol_trend: str) -> float:
    if regime != "uptrend" or len(highs) < 2:
        return 0.0
    h1, h2 = highs[-2][1], highs[-1][1]
    similarity = 1.0 - abs(h1 - h2) / max(abs(h1), 1.0)
    if similarity < 0.93:
        return 0.0
    score = 40.0 if similarity >= 0.97 else (25.0 if similarity >= 0.95 else 10.0)
    score += 30.0 if vol_trend == "contracting" else (10.0 if vol_trend == "flat" else 0.0)
    rsi_now = float(df["RSI14"].iloc[-1]) if "RSI14" in df.columns else 50.0
    score += 20.0 if rsi_now < 50 else (10.0 if rsi_now < 60 else 0.0)
    time_gap = abs(highs[-1][0] - highs[-2][0])
    if 10 <= time_gap <= 40:
        score += 10.0
    return score


def _score_double_bottom(lows, df: pd.DataFrame, regime: str, vol_trend: str) -> float:
    if regime != "downtrend" or len(lows) < 2:
        return 0.0
    l1, l2 = lows[-2][1], lows[-1][1]
    similarity = 1.0 - abs(l1 - l2) / max(abs(l1), 1.0)
    if similarity < 0.93:
        return 0.0
    score = 40.0 if similarity >= 0.97 else (25.0 if similarity >= 0.95 else 10.0)
    score += 30.0 if vol_trend == "contracting" else (10.0 if vol_trend == "flat" else 0.0)
    rsi_now = float(df["RSI14"].iloc[-1]) if "RSI14" in df.columns else 50.0
    score += 20.0 if rsi_now > 50 else (10.0 if rsi_now > 40 else 0.0)
    time_gap = abs(lows[-1][0] - lows[-2][0])
    if 10 <= time_gap <= 40:
        score += 10.0
    return score


def _score_head_shoulders(highs, df: pd.DataFrame, regime: str) -> float:
    if regime != "uptrend" or len(highs) < 3:
        return 0.0
    h1, hm, h2 = highs[-3][1], highs[-2][1], highs[-1][1]
    if not (hm > h1 and hm > h2):
        return 0.0
    score = 0.0
    shoulder_sym = 1.0 - abs(h1 - h2) / max(abs(h1), 1.0)
    score += 40.0 if shoulder_sym >= 0.95 else (20.0 if shoulder_sym >= 0.90 else 0.0)
    head_prom = (hm - max(h1, h2)) / max(abs(hm), 1.0)
    score += 30.0 if head_prom >= 0.03 else (15.0 if head_prom >= 0.02 else 0.0)
    rsi_now = float(df["RSI14"].iloc[-1]) if "RSI14" in df.columns else 50.0
    score += 20.0 if rsi_now < 50 else (10.0 if rsi_now < 60 else 0.0)
    return score


def _score_inv_head_shoulders(lows, df: pd.DataFrame, regime: str) -> float:
    if regime != "downtrend" or len(lows) < 3:
        return 0.0
    l1, lm, l2 = lows[-3][1], lows[-2][1], lows[-1][1]
    if not (lm < l1 and lm < l2):
        return 0.0
    score = 0.0
    shoulder_sym = 1.0 - abs(l1 - l2) / max(abs(l1), 1.0)
    score += 40.0 if shoulder_sym >= 0.95 else (20.0 if shoulder_sym >= 0.90 else 0.0)
    head_depth = (min(l1, l2) - lm) / max(abs(lm), 1.0)
    score += 30.0 if head_depth >= 0.03 else (15.0 if head_depth >= 0.02 else 0.0)
    rsi_now = float(df["RSI14"].iloc[-1]) if "RSI14" in df.columns else 50.0
    score += 20.0 if rsi_now > 50 else (10.0 if rsi_now > 40 else 0.0)
    return score


def _score_flag(df: pd.DataFrame, regime: str, vol_trend: str) -> tuple:
    """Returns (bull_flag_score, bear_flag_score)."""
    if len(df) < 40:
        return 0.0, 0.0
    full     = df["Close"].tail(40)
    pole     = full.head(25)
    rest     = full.tail(15)
    pole_chg = (float(pole.iloc[-1]) - float(pole.iloc[0])) / max(abs(float(pole.iloc[0])), 1.0)
    rest_chg = (float(rest.iloc[-1]) - float(rest.iloc[0])) / max(abs(float(rest.iloc[0])), 1.0)
    bull, bear = 0.0, 0.0
    if abs(pole_chg) >= 0.06:
        if pole_chg > 0:
            bull += 40.0
            bull += 35.0 if -0.04 <= rest_chg <= 0.01 else (15.0 if rest_chg <= 0.0 else 0.0)
            bull += 15.0 if vol_trend == "contracting" else 0.0
            bull += 10.0 if regime == "uptrend" else 0.0
        else:
            bear += 40.0
            bear += 35.0 if -0.01 <= rest_chg <= 0.04 else (15.0 if rest_chg >= 0.0 else 0.0)
            bear += 15.0 if vol_trend == "contracting" else 0.0
            bear += 10.0 if regime == "downtrend" else 0.0
    return bull, bear


def _score_rectangle(df: pd.DataFrame, regime: str) -> tuple:
    """Returns (bull_rect_score, bear_rect_score)."""
    if len(df) < 60:
        return 0.0, 0.0
    recent  = df["Close"].tail(30)
    high_r  = float(recent.max())
    low_r   = float(recent.min())
    mid     = (high_r + low_r) / 2.0
    spread  = (high_r - low_r) / max(abs(mid), 1.0)
    bull, bear = 0.0, 0.0
    if spread <= 0.06:
        base  = 50.0 if spread <= 0.03 else 30.0
        prior = df["Close"].tail(90).head(60)
        p_chg = (float(prior.iloc[-1]) - float(prior.iloc[0])) / max(abs(float(prior.iloc[0])), 1.0)
        if p_chg > 0.05:
            bull = base + (20.0 if regime == "uptrend" else 0.0)
        elif p_chg < -0.05:
            bear = base + (20.0 if regime == "downtrend" else 0.0)
    return bull, bear


def _score_range_formation(df: pd.DataFrame) -> float:
    """
    Range Formation: price oscillating inside a well-defined horizontal band
    for an extended period (30+ candles), regardless of prior trend direction.
    Scored on tightness and duration of the range.
    """
    if len(df) < 60:
        return 0.0
    recent  = df["Close"].tail(60)
    h_vals  = df["High"].tail(60).values
    l_vals  = df["Low"].tail(60).values
    band_hi = float(np.percentile(h_vals, 85))
    band_lo = float(np.percentile(l_vals, 15))
    spread  = (band_hi - band_lo) / max(abs((band_hi + band_lo) / 2.0), 1.0)
    if spread > 0.08:
        return 0.0
    # Count candles that stayed inside the band
    inside = int(np.sum((h_vals <= band_hi * 1.01) & (l_vals >= band_lo * 0.99)))
    score  = 0.0
    score += 40.0 if spread <= 0.04 else (25.0 if spread <= 0.06 else 10.0)
    score += 40.0 if inside >= 50 else (25.0 if inside >= 38 else 10.0)
    # Penalise if EMA trend is strong (range less likely if trending hard)
    if len(df) >= 50:
        ema50_slope = _slope_direction(df["Close"], min_points=20, max_points=50)
        if ema50_slope != "flat":
            score *= 0.7
    return score


def _score_range_breakout(df: pd.DataFrame, regime: str) -> tuple:
    """
    Range Breakout: recent candles broke above/below a prior consolidation band.
    Returns (bull_breakout_score, bear_breakout_score).
    Requires at least 60 bars: 40 for the range + 20 for the breakout leg.
    """
    if len(df) < 60:
        return 0.0, 0.0
    # Define the range from bars -60 to -21 (40 bars)
    range_window  = df.iloc[-60:-20]
    breakout_leg  = df.iloc[-20:]
    if len(range_window) < 20:
        return 0.0, 0.0
    h_vals   = range_window["High"].values
    l_vals   = range_window["Low"].values
    band_hi  = float(np.percentile(h_vals, 90))
    band_lo  = float(np.percentile(l_vals, 10))
    spread   = (band_hi - band_lo) / max(abs((band_hi + band_lo) / 2.0), 1.0)
    if spread > 0.12:
        return 0.0, 0.0

    last_close = float(breakout_leg["Close"].iloc[-1])
    bull, bear = 0.0, 0.0

    if last_close > band_hi:
        excess = (last_close - band_hi) / max(abs(band_hi), 1.0)
        bull  += 50.0 if excess >= 0.03 else (30.0 if excess >= 0.01 else 15.0)
        bull  += 20.0 if regime == "uptrend" else 0.0
        # Volume expansion on breakout leg is bullish confirmation
        vol_break  = float(breakout_leg["Volume"].mean())
        vol_range  = float(range_window["Volume"].mean())
        if vol_range > 0 and vol_break / vol_range >= 1.5:
            bull += 20.0
        elif vol_range > 0 and vol_break / vol_range >= 1.2:
            bull += 10.0
        bull = min(bull, 100.0)

    elif last_close < band_lo:
        excess = (band_lo - last_close) / max(abs(band_lo), 1.0)
        bear  += 50.0 if excess >= 0.03 else (30.0 if excess >= 0.01 else 15.0)
        bear  += 20.0 if regime == "downtrend" else 0.0
        vol_break = float(breakout_leg["Volume"].mean())
        vol_range = float(range_window["Volume"].mean())
        if vol_range > 0 and vol_break / vol_range >= 1.5:
            bear += 20.0
        elif vol_range > 0 and vol_break / vol_range >= 1.2:
            bear += 10.0
        bear = min(bear, 100.0)

    return bull, bear


def _score_cup_and_handle(df: pd.DataFrame, regime: str, vol_trend: str) -> float:
    """
    Cup and Handle (bullish continuation):
    - Cup: U-shaped price decline then recovery over ~40-60 bars
      (prior high → rounded low → recovery to prior high level)
    - Handle: shallow pullback of 5-15% over the final 10-20 bars
    - Breakout: latest close near or above the cup's right rim
    Scored on cup roundness, handle depth, volume contraction in handle,
    and volume expansion on breakout.
    """
    if len(df) < 90 or regime not in ("uptrend", "ranging"):
        return 0.0
    score = 0.0

    cup    = df.iloc[-90:-15]
    handle = df.iloc[-15:]
    if len(cup) < 45 or len(handle) < 5:
        return 0.0

    cup_left_high  = float(cup["High"].iloc[:15].max())
    cup_bottom     = float(cup["Low"].min())
    cup_right_high = float(cup["High"].iloc[-15:].max())

    # Cup depth: should be meaningful (8-33% decline from rim) — O'Neil standard
    cup_depth = (cup_left_high - cup_bottom) / max(abs(cup_left_high), 1.0)
    if not (0.08 <= cup_depth <= 0.33):
        return 0.0

    # Both rims should be at similar levels (within 5%)
    rim_sym = 1.0 - abs(cup_left_high - cup_right_high) / max(abs(cup_left_high), 1.0)
    score += 30.0 if rim_sym >= 0.97 else (20.0 if rim_sym >= 0.95 else (10.0 if rim_sym >= 0.92 else 0.0))
    if score == 0.0:
        return 0.0

    # Handle: should be a shallow pullback (5-15% from right rim)
    handle_high = float(handle["High"].max())
    handle_low  = float(handle["Low"].min())
    handle_ret  = (handle_high - handle_low) / max(abs(handle_high), 1.0)
    score += 25.0 if 0.05 <= handle_ret <= 0.15 else (10.0 if handle_ret <= 0.20 else 0.0)

    # Cup roundness: check that the bottom 30% of the cup is wider than 40% of cup length
    cup_close   = cup["Close"].values
    bottom_thr  = cup_bottom + 0.30 * (cup_left_high - cup_bottom)
    in_bottom   = int(np.sum(cup_close <= bottom_thr))
    score += 20.0 if in_bottom >= int(len(cup) * 0.40) else (10.0 if in_bottom >= int(len(cup) * 0.25) else 0.0)

    # Volume: should contract in handle, expand on latest bar (breakout attempt)
    if vol_trend == "contracting":
        score += 15.0
    latest_close = float(df["Close"].iloc[-1])
    if latest_close >= cup_right_high * 0.98:
        score += 10.0  # approaching breakout point

    return min(score, 100.0)


def _score_inv_cup_and_handle(df: pd.DataFrame, regime: str, vol_trend: str) -> float:
    """
    Inverse Cup and Handle (bearish continuation):
    Mirror image of Cup & Handle — an inverted U-shape followed by a
    shallow bounce (handle), then breakdown below the right rim.
    """
    if len(df) < 90 or regime not in ("downtrend", "ranging"):
        return 0.0
    score = 0.0

    cup    = df.iloc[-90:-15]
    handle = df.iloc[-15:]
    if len(cup) < 45 or len(handle) < 5:
        return 0.0

    cup_left_low   = float(cup["Low"].iloc[:15].min())
    cup_top        = float(cup["High"].max())
    cup_right_low  = float(cup["Low"].iloc[-15:].min())

    cup_rise = (cup_top - cup_left_low) / max(abs(cup_left_low), 1.0)
    if not (0.08 <= cup_rise <= 0.33):
        return 0.0

    rim_sym = 1.0 - abs(cup_left_low - cup_right_low) / max(abs(cup_left_low), 1.0)
    score += 30.0 if rim_sym >= 0.97 else (20.0 if rim_sym >= 0.95 else (10.0 if rim_sym >= 0.92 else 0.0))
    if score == 0.0:
        return 0.0

    handle_high = float(handle["High"].max())
    handle_low  = float(handle["Low"].min())
    handle_ret  = (handle_high - handle_low) / max(abs(handle_high), 1.0)
    score += 25.0 if 0.05 <= handle_ret <= 0.15 else (10.0 if handle_ret <= 0.20 else 0.0)

    cup_close  = cup["Close"].values
    top_thr    = cup_top - 0.30 * (cup_top - cup_left_low)
    in_top     = int(np.sum(cup_close >= top_thr))
    score += 20.0 if in_top >= int(len(cup) * 0.40) else (10.0 if in_top >= int(len(cup) * 0.25) else 0.0)

    if vol_trend == "contracting":
        score += 15.0
    latest_close = float(df["Close"].iloc[-1])
    if latest_close <= cup_right_low * 1.02:
        score += 10.0

    return min(score, 100.0)


# ── Main pattern detection entry point ───────────────────────────────────────

def detect_chart_patterns(df_or_series) -> str:
    """
    Phase 1 + Phase 2 enhanced chart pattern detection.

    Accepts either a full OHLCV DataFrame (preferred) or a Close price Series
    (legacy fallback) so existing call-sites need no change.

    Phase 1: ATR-filtered High/Low swings, volume confirmation, graduated scoring.
    Phase 2: HMM-based regime detection (hmmlearn); falls back to EMA/slope if
             hmmlearn is not installed — script works without it.

    Patterns detected:
      Double Top / Double Bottom
      Head & Shoulders / Inverse Head & Shoulders
      Bullish Flag / Bearish Flag
      Bullish Rectangle / Bearish Rectangle
      Range Formation
      Bullish Range Breakout / Bearish Range Breakout
      Cup and Handle / Inverse Cup and Handle

    Each detected pattern is labelled with a confidence percentage:
      >= 60  → confirmed   e.g. "Double Top (74%)"
      40–59  → probable    e.g. "Possible Double Top (48%)"
    """
    if isinstance(df_or_series, pd.Series):
        price = df_or_series
        stub  = pd.DataFrame({
            "Close":  price.values,
            "High":   price.values,
            "Low":    price.values,
            "Open":   price.values,
            "Volume": np.ones(len(price)),
        }, index=price.index)
        df = stub
    else:
        df = df_or_series

    if len(df) < 30:
        return ""

    # ── Feature extraction ────────────────────────────────────────────────
    highs, lows = _significant_swings_hl(df, lookback=90, atr_mult=0.75)
    regime      = _detect_regime(df)          # HMM or EMA fallback

    vol10 = float(df["Volume"].tail(10).mean()) if "Volume" in df.columns else 1.0
    vol30 = float(df["Volume"].tail(30).mean()) if "Volume" in df.columns else 1.0
    ratio = vol10 / max(vol30, 1.0)
    vol_trend = "expanding" if ratio > 1.10 else ("contracting" if ratio < 0.90 else "flat")

    patterns = []

    def _add(name: str, score: float) -> None:
        s = int(round(score))
        if score >= 60:
            patterns.append(f"{name} ({s}%)")
        elif score >= 40:
            patterns.append(f"Possible {name} ({s}%)")

    # ── Score all patterns ────────────────────────────────────────────────
    _add("Double Top",                   _score_double_top(highs, df, regime, vol_trend))
    _add("Double Bottom",                _score_double_bottom(lows, df, regime, vol_trend))
    _add("Head & Shoulders",             _score_head_shoulders(highs, df, regime))
    _add("Inverse Head & Shoulders",     _score_inv_head_shoulders(lows, df, regime))

    bull_flag, bear_flag = _score_flag(df, regime, vol_trend)
    _add("Bullish Flag",  bull_flag)
    _add("Bearish Flag",  bear_flag)

    bull_rect, bear_rect = _score_rectangle(df, regime)
    _add("Bullish Rectangle", bull_rect)
    _add("Bearish Rectangle", bear_rect)

    _add("Range Formation", _score_range_formation(df))

    bull_brk, bear_brk = _score_range_breakout(df, regime)
    _add("Bullish Range Breakout", bull_brk)
    _add("Bearish Range Breakout", bear_brk)

    _add("Cup and Handle",         _score_cup_and_handle(df, regime, vol_trend))
    _add("Inverse Cup and Handle", _score_inv_cup_and_handle(df, regime, vol_trend))

    # Fix 7 — Conflicting pattern warning.
    # When both bullish and bearish patterns are detected simultaneously the
    # signals are contradictory. Flag this explicitly so the analyst knows to
    # wait for resolution rather than acting on either signal alone.
    _bullish_keys = {"Double Bottom", "Inverse Head", "Cup and Handle",
                     "Bullish", "Range Breakout"}
    _bearish_keys = {"Double Top", "Head & Shoulders", "Bearish",
                     "Inverse Cup"}
    _has_bull = any(any(k in p for k in _bullish_keys) for p in patterns)
    _has_bear = any(any(k in p for k in _bearish_keys) for p in patterns)
    if _has_bull and _has_bear:
        patterns.append("⚠ Conflicting Signals — wait for resolution")

    return "; ".join(patterns) if patterns else ""


def compute_pivots_support_resistance(df: pd.DataFrame, lookback: int = 252):
    """
    v1.9.4 — ATR-filtered High/Low swing S&R with role-reversal support.

    Fixes vs v1.9.2 (four confirmed bugs):
      Bug 1 — Role reversal: a former resistance level now below price is
               repurposed as support (and vice-versa). Previously those zones
               were silently discarded, causing ~40 stocks to show empty S/R.
      Bug 2 — WICK_MIN_ATR lowered 0.50 → 0.25 so smoothly-trending stocks
               (ASIANPAINT, HUL, TCS etc.) with small daily wicks are not
               excluded from candidate detection.
      Bug 3 — MIN_REACTIONS lowered 3 → 2; a 252-bar window rarely produces
               3 well-separated touches at the same price level.
      Bug 4 — BAR_GAP_MIN lowered 10 → 5; packed consolidation zones where
               multiple touches occur within 10 bars were being collapsed to a
               single accepted touch, incorrectly failing MIN_REACTIONS=2.
      Also:  Swing detection now uses actual High/Low prices (not Close),
               lookback extended from 120 → 252 bars, single-touch fallback
               prevents blank cells even for strongly trending stocks, and
               dead code present in v1.9.2 has been removed.
    """
    if df.empty or len(df) < 20:
        return "", ""

    recent     = df.tail(lookback).reset_index(drop=True)
    n          = len(recent)
    last_close = float(recent["Close"].iloc[-1])

    atr_val = float(atr(recent, period=14).iloc[-1])
    if atr_val <= 0:
        atr_val = last_close * 0.01

    PIVOT_WING    = 3      # bars either side required for swing confirmation
    WICK_MIN_ATR  = 0.25   # minimum wick as ATR multiple         (was 0.50 — Bug 2)
    CLUSTER_ATR   = 1.0    # zone-clustering width in ATR units
    MIN_REACTIONS = 2      # minimum well-separated touches        (was 3  — Bug 3)
    BAR_GAP_MIN   = 5      # minimum bars between counted touches  (was 10 — Bug 4)

    high_arr  = recent["High"].values
    low_arr   = recent["Low"].values
    open_arr  = recent["Open"].values
    close_arr = recent["Close"].values

    # Each entry: (price, bar_index, wick_ratio, direction)
    # direction: +1 = resistance, -1 = support
    candidates = []

    for i in range(PIVOT_WING, n - PIVOT_WING):
        is_sh = (all(high_arr[i] >= high_arr[i - k] for k in range(1, PIVOT_WING + 1)) and
                 all(high_arr[i] >= high_arr[i + k] for k in range(1, PIVOT_WING + 1)))
        is_sl = (all(low_arr[i]  <= low_arr[i - k]  for k in range(1, PIVOT_WING + 1)) and
                 all(low_arr[i]  <= low_arr[i + k]  for k in range(1, PIVOT_WING + 1)))

        if is_sh:
            upper_wick = high_arr[i] - max(close_arr[i], open_arr[i])
            if upper_wick >= WICK_MIN_ATR * atr_val:
                # Bug 1 fix — role reversal: swing-high below current price → support
                direction = -1 if high_arr[i] < last_close else +1
                candidates.append((high_arr[i], i, upper_wick / atr_val, direction))

        if is_sl:
            lower_wick = min(close_arr[i], open_arr[i]) - low_arr[i]
            if lower_wick >= WICK_MIN_ATR * atr_val:
                # Bug 1 fix — role reversal: swing-low above current price → resistance
                direction = +1 if low_arr[i] > last_close else -1
                candidates.append((low_arr[i], i, lower_wick / atr_val, direction))

    res_cands = [(p, i, w) for p, i, w, d in candidates if d == +1]
    sup_cands = [(p, i, w) for p, i, w, d in candidates if d == -1]

    def _cluster(cands):
        if not cands:
            return []
        cs = sorted(cands, key=lambda x: x[0])
        zones, cur = [], [cs[0]]
        for item in cs[1:]:
            if abs(item[0] - cur[-1][0]) <= CLUSTER_ATR * atr_val:
                cur.append(item)
            else:
                zones.append(cur)
                cur = [item]
        zones.append(cur)
        return zones

    def _score_zone(zone_items, min_r=MIN_REACTIONS):
        by_time  = sorted(zone_items, key=lambda x: x[1])
        accepted = [by_time[0]]
        for item in by_time[1:]:
            if item[1] - accepted[-1][1] >= BAR_GAP_MIN:
                accepted.append(item)
        if len(accepted) < min_r:
            return None
        prices = [x[0] for x in accepted]
        lo  = min(prices) - 0.15 * atr_val
        hi  = max(prices) + 0.15 * atr_val
        mid = (lo + hi) / 2.0
        sc  = (len(accepted) * 3.0
               + (max(x[1] for x in accepted) - min(x[1] for x in accepted)) / 20.0
               + float(np.mean([x[2] for x in accepted])))
        return round(lo, 2), round(hi, 2), mid, sc

    def _build_scored(cands, min_r=MIN_REACTIONS):
        scored = []
        for z in _cluster(cands):
            r = _score_zone(z, min_r=min_r)
            if r:
                scored.append(r)
        return scored

    scored_res = _build_scored(res_cands)
    scored_sup = _build_scored(sup_cands)

    # Fallback pass 1: relax to single-touch zones
    if not scored_res and res_cands:
        scored_res = _build_scored(res_cands, min_r=1)
    if not scored_sup and sup_cands:
        scored_sup = _build_scored(sup_cands, min_r=1)

    # Fallback pass 2: if still empty (no wick-qualified swings at all), use
    # raw swing-high / swing-low prices without wick filter as last resort
    if not scored_res or not scored_sup:
        raw_highs = [(high_arr[i], i, 0.1)
                     for i in range(PIVOT_WING, n - PIVOT_WING)
                     if (all(high_arr[i] >= high_arr[i-k] for k in range(1, PIVOT_WING+1)) and
                         all(high_arr[i] >= high_arr[i+k] for k in range(1, PIVOT_WING+1)))]
        raw_lows  = [(low_arr[i],  i, 0.1)
                     for i in range(PIVOT_WING, n - PIVOT_WING)
                     if (all(low_arr[i] <= low_arr[i-k] for k in range(1, PIVOT_WING+1)) and
                         all(low_arr[i] <= low_arr[i+k] for k in range(1, PIVOT_WING+1)))]
        raw_res = [(p, i, w) for p, i, w in raw_highs if p >= last_close] +                   [(p, i, w) for p, i, w in raw_lows  if p >  last_close]
        raw_sup = [(p, i, w) for p, i, w in raw_lows  if p <= last_close] +                   [(p, i, w) for p, i, w in raw_highs if p <  last_close]
        if not scored_res and raw_res:
            scored_res = _build_scored(raw_res, min_r=1)
        if not scored_sup and raw_sup:
            scored_sup = _build_scored(raw_sup, min_r=1)

    # Pick top 2 closest to last_close on each side
    top_res = sorted(scored_res, key=lambda z: z[2])[:2]   # lowest resistance first
    top_sup = sorted(scored_sup, key=lambda z: -z[2])[:2]  # highest support first

    def _fmt(zones, prefix):
        parts = []
        for idx, (lo, hi, _, _) in enumerate(zones, start=1):
            if hi - lo <= 0.005 * max(lo, 1.0):
                parts.append(f"{prefix}{idx}: {lo:.2f}")
            else:
                parts.append(f"{prefix}{idx}: {lo:.2f}–{hi:.2f}")
        return "; ".join(parts)

    return _fmt(top_sup, "S"), _fmt(top_res, "R")

def dow_theory_label(df: pd.DataFrame) -> str:
    """
    Enhanced Dow Theory analysis returning a rich multi-line label:

    Line 1 – Primary Trend   (200-bar structure): Uptrend / Downtrend / Sideways
    Line 2 – Secondary Trend (60-bar structure):  Uptrend / Downtrend / Sideways
    Line 3 – Minor Trend     (20-bar structure):  Uptrend / Downtrend / Sideways
    Line 4 – Phase           (position in cycle): see below

    Bull phases:  Accumulation | Markup | Distribution
    Bear phases:  Distribution | Markdown | Accumulation
    Transition:   Reversal Watch

    Phase logic (all three timeframes must agree on primary direction):
    ┌─────────────────────────────────────────────────────────────────┐
    │ Bull cycle                                                       │
    │  Accumulation : Primary=Up  + Secondary=Sideways/Down           │
    │                 + EMA200 slope flat/just turning up              │
    │  Markup       : Primary=Up  + Secondary=Up + Minor=Up           │
    │                 + price above EMA50 & EMA200                    │
    │  Distribution : Primary=Up  + Minor=Sideways/Down               │
    │                 + RSI divergence bearish / price near 52w high   │
    │                                                                  │
    │ Bear cycle                                                       │
    │  Distribution : Primary=Down + Secondary=Sideways/Up            │
    │                 + EMA200 slope flat/just turning down            │
    │  Markdown     : Primary=Down + Secondary=Down + Minor=Down      │
    │                 + price below EMA50 & EMA200                    │
    │  Accumulation : Primary=Down + Minor=Sideways/Up                │
    │                 + RSI divergence bullish / near 52w low          │
    └─────────────────────────────────────────────────────────────────┘

    Volume confirmation is used to boost phase confidence.
    Falls back gracefully when insufficient data is available.
    """

    # ── helpers ──────────────────────────────────────────────────────
    def _trend(window: pd.DataFrame) -> str:
        """Determine trend from ATR-filtered swing structure on a sub-window."""
        n = len(window)
        if n < 10:
            return "Sideways"
        # Use simple percentile-based swing detection (no external call needed)
        h = window["High"].values
        l = window["Low"].values
        c = window["Close"].values
        # Compute ATR-like noise floor
        tr = np.array([max(h[i]-l[i],
                           abs(h[i]-c[i-1]) if i>0 else 0,
                           abs(l[i]-c[i-1]) if i>0 else 0)
                       for i in range(n)])
        noise = float(np.mean(tr)) * 0.75

        swing_highs, swing_lows = [], []
        for i in range(1, n - 1):
            if h[i] > h[i-1] + noise and h[i] > h[i+1] + noise:
                swing_highs.append(h[i])
            if l[i] < l[i-1] - noise and l[i] < l[i+1] - noise:
                swing_lows.append(l[i])

        # Need at least 2 swings of each type for HH/HL or LH/LL confirmation
        bull = (len(swing_highs) >= 2 and swing_highs[-1] > swing_highs[-2] and
                len(swing_lows)  >= 2 and swing_lows[-1]  > swing_lows[-2])
        bear = (len(swing_highs) >= 2 and swing_highs[-1] < swing_highs[-2] and
                len(swing_lows)  >= 2 and swing_lows[-1]  < swing_lows[-2])

        # Secondary confirmation via EMA slope
        ema_s = float(c[-1]) - float(c[max(0, n//2)])
        if bull and ema_s >= 0:
            return "Uptrend"
        if bear and ema_s <= 0:
            return "Downtrend"
        # Partial signals — one swing type confirms
        if (len(swing_highs) >= 2 and swing_highs[-1] > swing_highs[-2] and ema_s > 0):
            return "Uptrend"
        if (len(swing_highs) >= 2 and swing_highs[-1] < swing_highs[-2] and ema_s < 0):
            return "Downtrend"
        return "Sideways"

    def _vol_trend(window: pd.DataFrame) -> str:
        v = window["Volume"].values.astype(float)
        if len(v) < 6:
            return "flat"
        half = len(v) // 2
        return "expanding" if v[half:].mean() > v[:half].mean() * 1.1 else (
               "contracting" if v[half:].mean() < v[:half].mean() * 0.9 else "flat")

    # ── determine data availability ───────────────────────────────────
    n = len(df)
    MIN_PRIMARY   = 120   # ~6 months for primary
    MIN_SECONDARY = 60    # ~3 months for secondary
    MIN_MINOR     = 20    # ~1 month for minor

    # ── compute trends at each timeframe ─────────────────────────────
    primary_trend = (
        _trend(df.tail(MIN_PRIMARY))   if n >= MIN_PRIMARY   else "Insufficient data"
    )
    secondary_trend = (
        _trend(df.tail(MIN_SECONDARY)) if n >= MIN_SECONDARY else "Insufficient data"
    )
    minor_trend = (
        _trend(df.tail(MIN_MINOR))     if n >= MIN_MINOR     else "Insufficient data"
    )

    if primary_trend == "Insufficient data":
        return "Insufficient data"

    # ── EMA context for phase detection ──────────────────────────────
    close       = df["Close"]
    ema50_val   = float(ema(close, 50).iloc[-1])  if n >= 50  else None
    ema200_val  = float(ema(close, 200).iloc[-1]) if n >= 200 else None
    last_close  = float(close.iloc[-1])

    # EMA200 slope: compare last value vs 20 bars ago
    if n >= 220:
        ema200_20ago = float(ema(close, 200).iloc[-21])
        ema200_slope = "up"   if ema200_val > ema200_20ago * 1.001 else (
                       "down" if ema200_val < ema200_20ago * 0.999 else "flat")
    else:
        ema200_slope = "flat"

    # RSI on latest bar (may already be computed upstream but safe to recompute)
    rsi_val = float(rsi(close, 14).iloc[-1]) if n >= 15 else 50.0

    # Volume trend on primary window
    vol = _vol_trend(df.tail(MIN_PRIMARY) if n >= MIN_PRIMARY else df)

    # 52-week high/low proximity
    w52 = df.tail(252) if n >= 252 else df
    hi52 = float(w52["High"].max())
    lo52 = float(w52["Low"].min())
    rng52 = max(hi52 - lo52, 1.0)
    pct_from_hi = (hi52 - last_close) / rng52   # 0 = at high, 1 = at low
    pct_from_lo = (last_close - lo52) / rng52   # 0 = at low,  1 = at high

    # ── phase detection ───────────────────────────────────────────────
    phase = "Indeterminate"

    if primary_trend == "Uptrend":
        if secondary_trend in ("Sideways", "Downtrend") and ema200_slope in ("flat", "up"):
            phase = "Bull: Accumulation"
        elif (secondary_trend == "Uptrend" and minor_trend == "Uptrend"
              and ema50_val and ema200_val
              and last_close > ema50_val and last_close > ema200_val):
            phase = "Bull: Markup"
        elif (minor_trend in ("Sideways", "Downtrend")
              and (rsi_val > 65 or pct_from_hi < 0.10)):
            phase = "Bull: Distribution"
        else:
            phase = "Bull: Markup"          # default for confirmed uptrend

    elif primary_trend == "Downtrend":
        if secondary_trend in ("Sideways", "Uptrend") and ema200_slope in ("flat", "down"):
            phase = "Bear: Distribution"
        elif (secondary_trend == "Downtrend" and minor_trend == "Downtrend"
              and ema50_val and ema200_val
              and last_close < ema50_val and last_close < ema200_val):
            phase = "Bear: Markdown"
        elif (minor_trend in ("Sideways", "Uptrend")
              and (rsi_val < 35 or pct_from_lo < 0.10)):
            phase = "Bear: Accumulation"
        else:
            phase = "Bear: Markdown"        # default for confirmed downtrend

    else:
        # Primary = Sideways — check for potential reversal setup
        if secondary_trend == "Uptrend" and minor_trend == "Uptrend":
            phase = "Reversal Watch (Bullish)"
        elif secondary_trend == "Downtrend" and minor_trend == "Downtrend":
            phase = "Reversal Watch (Bearish)"
        else:
            phase = "Consolidation"

    # ── volume annotation + Fix 10: vol changes phase label ─────────────
    # Vol Contraction in a bullish phase warns of weakening momentum.
    # Vol Surge in a bearish phase warns of accelerating selling pressure.
    # Vol Contraction in Bear: Accumulation is a positive sign (supply drying up).
    # Vol Surge in Bull: Accumulation confirms smart-money absorption.
    if vol == "contracting":
        if phase == "Bull: Markup":
            phase = "Bull: Markup ⚠ Vol Contraction"
        elif phase == "Bear: Accumulation":
            phase = "Bear: Accumulation ✓ Vol Contraction"
        elif phase == "Bull: Distribution":
            phase = "Bull: Distribution ⚠ Vol Contraction"
        vol_note = " ↓Vol"
    elif vol == "expanding":
        if phase == "Bear: Markdown":
            phase = "Bear: Markdown ⚠ Vol Surge"
        elif phase == "Bull: Accumulation":
            phase = "Bull: Accumulation ✓ Vol Surge"
        elif phase == "Bear: Distribution":
            phase = "Bear: Distribution ⚠ Vol Surge"
        vol_note = " ✓Vol"
    else:
        vol_note = ""

    # ── compose output string ─────────────────────────────────────────
    return (
        f"Primary: {primary_trend} | "
        f"Secondary: {secondary_trend} | "
        f"Minor: {minor_trend} | "
        f"Phase: {phase}"
    )




# Fix 16 — India-market domain keyword sets.
# VADER is trained on English social media text and has no awareness of
# Indian market terminology. These sets provide a high-precision pre-pass
# that fires before VADER and returns immediately when a known phrase is found.
INDIA_BULLISH_KW = {
    "bulk deal", "block deal", "promoter buying", "upper circuit",
    "beat estimate", "buyback", "dividend", "upgrades", "strong q",
    "order win", "order book", "capex", "expansion", "fundraise",
    "qip", "rights issue", "bonus issue", "stock split", "nse listing",
    "strong results", "margin expansion", "debt free", "zero debt",
}
INDIA_BEARISH_KW = {
    "lower circuit", "operator", "sebi probe", "miss estimate",
    "downgrade", "promoter selling", "pledge", "npa", "fraud",
    "insider trading", "margin call", "default", "insolvency",
    "forensic audit", "income tax raid", "ed raid", "cbi probe",
    "profit warning", "revenue miss", "weak results", "margin pressure",
}

def classify_sentiment(text: str, cfg: Config) -> str:
    """Return 'positive', 'negative', or 'neutral' for a headline.

    Fix 16: Indian market domain keyword pre-pass fires before VADER so that
    high-signal NSE-specific phrases (circuit limits, SEBI actions, bulk deals,
    promoter activity, etc.) are classified correctly without VADER's
    US-centric bias.
    """
    if not text:
        return "neutral"
    t_lower = text.lower()
    # Fix 16 — India keyword pre-pass (high precision, zero latency)
    if any(k in t_lower for k in INDIA_BULLISH_KW):
        return "positive"
    if any(k in t_lower for k in INDIA_BEARISH_KW):
        return "negative"
    # Fall through to VADER for generic English text
    scores = SENTIMENT_ANALYZER.polarity_scores(text)
    comp = scores.get("compound", 0.0)
    if comp >= cfg.sentiment_pos_cutoff:
        return "positive"
    if comp <= cfg.sentiment_neg_cutoff:
        return "negative"
    return "neutral"

def momentum_label(row) -> str:
    bullish_count = 0
    bearish_count = 0
    if row["Close"] > row["EMA21"] > row["EMA50"]:
        bullish_count += 1
    else:
        bearish_count += 1
    if row["RSI14"] >= 55:
        bullish_count += 1
    elif row["RSI14"] <= 45:
        bearish_count += 1
    if row["MACD"] > row["MACDSignal"]:
        bullish_count += 1
    else:
        bearish_count += 1
    if row.get("SupertrendGreen_10_3", False):
        bullish_count += 1
    else:
        bearish_count += 1
    if bullish_count >= 3:
        return "Bullish"
    if bearish_count >= 3:
        return "Bearish"
    return "Sideways"


# ── News source label map ─────────────────────────────────────────────────────
_NEWS_SOURCE_LABELS = {
    "economictimes.indiatimes.com": "ET",
    "moneycontrol.com":             "MC",
    "businessstandard.com":         "BS",
    "livemint.com":                 "Mint",
    "financialexpress.com":         "FE",
    "thehindu.com":                 "Hindu",
    "ndtvprofit.com":               "NDTV",
    "bloombergquint.com":           "BQ",
    "reuters.com":                  "Reuters",
    "bloomberg.com":                "Bloomberg",
}

def _news_source_label(url: str) -> str:
    """Return a short source badge like [ET] from an article URL."""
    try:
        from urllib.parse import urlparse
        host = urlparse(url).netloc.lower().replace("www.", "")
        for domain, label in _NEWS_SOURCE_LABELS.items():
            if domain in host:
                return f"[{label}]"
    except Exception:
        pass
    return ""

def fetch_recent_news(symbol: str, company: str, cfg: Config, s: requests.Session):
    """Fetch recent Google News RSS headlines for this stock with sentiment."""
    # Generic queries first, then source-targeted to surface ET / MC articles
    queries = [
        f'"{company}" NSE',
        f'"{company}" stock',
        f'{symbol} NSE',
        f'site:economictimes.indiatimes.com "{company}"',
        f'site:moneycontrol.com "{company}"',
        f'site:businessstandard.com "{company}"',
    ]
    cutoff = TODAY_IST - timedelta(days=cfg.recent_news_days)
    items = []
    seen = set()

    for q in queries:
        try:
            url = (
                "https://news.google.com/rss/search?"
                f"q={quote_plus(q)}+when:{cfg.recent_news_days}d"
                "&hl=en-IN&gl=IN&ceid=IN:en"
            )
            logger.info("News RSS query for %s: %s", symbol, url)
            r = _http_get_with_retry(s, url, timeout=25, max_retries=3, retry_delay=5, label=f"News RSS {symbol}")
            logger.info("News RSS status for %s: %s", symbol, r.status_code)
            if r.status_code != 200:
                continue

            root = ET.fromstring(r.content)
            for item in root.findall("./channel/item"):
                title = (item.findtext("title") or "").strip()
                link = (item.findtext("link") or "").strip()
                pub = (item.findtext("pubDate") or "").strip()
                if not title or not link:
                    continue

                key = (title.lower(), link)
                if key in seen:
                    continue
                seen.add(key)

                pub_dt = None
                try:
                    pub_dt = pd.to_datetime(pub, utc=True).tz_convert(IST)
                except Exception:
                    pass

                if pub_dt is not None and pub_dt.to_pydatetime().replace(tzinfo=IST) < cutoff:
                    continue

                sentiment = classify_sentiment(title, cfg)
                items.append({
                    "title":     title,
                    "link":      link,
                    "pub":       pub_dt,
                    "sentiment": sentiment,
                    "source":    _news_source_label(link),
                })

                if len(items) >= cfg.recent_news_max_items:
                    break

            if len(items) >= cfg.recent_news_max_items:
                break

        except Exception as e:
            logger.warning("News fetch failed for %s | query=%s | error=%s", symbol, q, e)

    if not items:
        return [], "No recent news found"

    html_items = []
    text_items = []

    # Sort articles newest-first; items with no pub date fall to the bottom.
    # CA block is prepended separately by _build_stock_row so this sort
    # does not affect corporate action ordering.
    items.sort(key=lambda x: x["pub"] if x["pub"] is not None else pd.Timestamp.min.tz_localize("UTC"), reverse=True)

    for n in items[: cfg.recent_news_max_items]:
        safe_title  = html.escape(n["title"])
        safe_link   = html.escape(n["link"], quote=True)
        sent        = n["sentiment"]
        sent_tag    = f"({sent})"
        source_badge = n.get("source", "")
        date_str    = ""
        if n["pub"] is not None:
            date_str = " – " + n["pub"].strftime("%d-%b-%Y %H:%M")
        source_html = f" <span style='color:#c8dcf4;font-size:0.85em'>{source_badge}</span>" if source_badge else ""
        html_items.append(
            f"<a href='{safe_link}' target='_blank' rel='noopener noreferrer'>"
            f"{safe_title}</a>{source_html} {sent_tag}{date_str}"
        )
        source_txt = f" {source_badge}" if source_badge else ""
        text_items.append(f"{n['title']}{source_txt} {sent_tag} – {n['link']}")

    return html_items, " | ".join(text_items)



# ── Conviction Scorer (v2.0.0) ────────────────────────────────────────────────
#
# Scores a shortlisted stock 0-100 across 6 weighted factor groups.
# Returns (score: float, breakdown: str) where breakdown is a pipe-separated
# string with one segment per factor group — ideal for a single Excel/HTML cell.
#
# Weights: PriceAction(30) + Volume(20) + Trend(20) + RSI(15) + MACD(10) + ST(5)
#
# Industry references:
#   EMA alignment (9>21>50): Appel (2005), "Technical Analysis – Power Tools"
#   Dow Theory phases:       Rhea (1932), "The Dow Theory"
#   RSI zones (55-70 bullish): Wilder (1978), "New Concepts in Technical Trading"
#   Volume / breakout:       Pring (2002), "Technical Analysis Explained", Ch.6
#   SuperTrend:              Anand (2009), Futures & Options Analyst
#
# ⚠ Non-standard / proprietary notes:
#   - Pattern Quality uses ATR-based range contraction as a "tight base" proxy
#     (inspired by O'Neil VCP but simplified — not formally published).
#   - Candle Strength body-ratio thresholds (0.6 / 0.45) empirically calibrated
#     for Indian large-cap daily data; not from a formal academic source.
#   - Market Structure uses 5-bar & 20-bar HH/HL check — shorter than Dow's
#     20-bar standard but appropriate for daily conviction screening.
#   - Aggregate score is a simple weighted sum; no covariance adjustment applied.

def _conv_price_action(g: pd.DataFrame):
    """Price Action — max 25 pts: KeyLvl(8)+PatQual(7)+CandleStr(3)+MktStr(7)."""
    reasons = []
    if len(g) < 21:
        return 0.0, ["PriceAction: insufficient data"]
    close = g["Close"].values
    high  = g["High"].values
    low   = g["Low"].values
    open_ = g["Open"].values
    last_c = close[-1]; last_h = high[-1]; last_l = low[-1]; last_o = open_[-1]

    atr_v = float(atr(g.tail(20).reset_index(drop=True), period=14).iloc[-1])
    if atr_v <= 0:
        atr_v = last_c * 0.01

    # ── (A) Key Level with Breakout Quality — max 8 ──────────────────────────
    lk        = min(20, len(g) - 1)
    res_level = float(np.max(high[-lk-1:-1]))
    sup_level = float(np.min(low[-lk-1:-1]))
    spread    = max(res_level - sup_level, last_c * 0.001)
    dist      = (res_level - last_c) / spread

    vol_arr   = g["Volume"].values.astype(float)
    avg14_vol = float(np.mean(vol_arr[-15:-1])) if len(vol_arr) >= 15 else float(np.mean(vol_arr[:-1]))
    vol_spike = (float(vol_arr[-1]) / max(avg14_vol, 1.0)) >= 1.5
    upper_wick_pct = (last_h - last_c) / max(last_h - last_l, atr_v * 0.1)
    fake_breakout  = upper_wick_pct > 0.5   # wick > 50% of range = rejection

    if last_c > res_level:
        if vol_spike and not fake_breakout:
            kl = 8.0; reasons.append(f"KeyLvl=8/8 breakout+vol+close>{res_level:.2f} 🔥")
        elif fake_breakout:
            kl = 3.5; reasons.append(f"KeyLvl=4/8 wick-rejection>{res_level:.2f} ⚠")
        else:
            kl = 6.0; reasons.append(f"KeyLvl=6/8 breakout>{res_level:.2f} no vol confirm")
    elif dist < 0.05:
        kl = 5.5; reasons.append(f"KeyLvl=6/8 nearing resistance {res_level:.2f}")
    elif last_c < sup_level * 1.02:
        kl = 5.0; reasons.append(f"KeyLvl=5/8 near support {sup_level:.2f}")
    elif dist < 0.25:
        kl = 3.5; reasons.append(f"KeyLvl=4/8 upper-mid range")
    else:
        kl = max(0.0, min(3.0, 3.0 * (1.0 - dist)))
        reasons.append(f"KeyLvl={kl:.0f}/8 mid-range")

    # ── (B) Pattern Quality — structure-aware, max 7 ─────────────────────────
    rng_pct  = float(np.max(high[-10:]) - np.min(low[-10:])) / max(last_c, 1.0)
    lows10   = low[-10:]
    hl_count = sum(1 for i in range(1, len(lows10)) if lows10[i] > lows10[i-1])
    hl_ratio = hl_count / max(len(lows10) - 1, 1)

    if rng_pct < 0.04 and hl_ratio >= 0.6:
        pq = 7.0; reasons.append("PatQual=7/7 tight+Higher-Lows ✓")
    elif rng_pct < 0.05 and hl_ratio >= 0.4:
        pq = 6.0; reasons.append("PatQual=6/7 tight flat base")
    elif rng_pct < 0.08 and hl_ratio >= 0.5:
        pq = 5.0; reasons.append("PatQual=5/7 wide but trending up")
    elif rng_pct < 0.06:
        pq = 5.0; reasons.append("PatQual=5/7 tight base (no HL)")
    elif rng_pct < 0.10 and hl_ratio >= 0.4:
        pq = 4.0; reasons.append("PatQual=4/7 moderate-wide trending")
    elif rng_pct < 0.12:
        pq = 2.5; reasons.append("PatQual=3/7 wide range")
    else:
        pq = 1.0; reasons.append("PatQual=1/7 very noisy/wide")

    # ── (C) Candle Strength — max 3 ──────────────────────────────────────────
    body = abs(last_c - last_o)
    crng = max(last_h - last_l, atr_v * 0.1)
    br   = body / crng
    ctop = (last_h - last_c) / crng
    if last_c > last_o:
        if br >= 0.6 and ctop < 0.15:
            cs = 3.0; reasons.append("CandleStr=3/3 strong bull@high")
        elif br >= 0.45:
            cs = 2.0; reasons.append("CandleStr=2/3 moderate bull body")
        elif br >= 0.25:
            cs = 1.0; reasons.append("CandleStr=1/3 small bull body")
        else:
            cs = 0.5; reasons.append("CandleStr=1/3 doji/indecision")
    else:
        if br >= 0.6:
            cs = 0.0; reasons.append("CandleStr=0/3 strong bear candle ⚠")
        elif br >= 0.35:
            cs = 0.5; reasons.append("CandleStr=1/3 bearish candle")
        else:
            cs = 1.0; reasons.append("CandleStr=1/3 spinning top")

    # ── (D) Market Structure — max 7 ─────────────────────────────────────────
    hh5  = bool(last_h > high[-6])  if len(high) > 5  else False
    hl5  = bool(last_l > low[-6])   if len(low)  > 5  else False
    hh20 = bool(last_h > high[-21]) if len(high) > 20 else False
    hl20 = bool(last_l > low[-21])  if len(low)  > 20 else False
    e50v = (float(g["EMA50"].iloc[-1])
            if "EMA50" in g.columns and pd.notna(g["EMA50"].iloc[-1]) else None)
    if hh5 and hl5 and hh20 and hl20:
        ms = 7.0; reasons.append("MktStr=7/7 HH+HL both TFs ✓")
    elif hh5 and hl5:
        ms = 5.0; reasons.append("MktStr=5/7 HH+HL short-term")
    elif hh20 and hl20:
        ms = 4.0; reasons.append("MktStr=4/7 HH+HL medium-term")
    elif hh5 or hh20:
        ms = 2.0; reasons.append("MktStr=2/7 partial HH only")
    else:
        ms = 0.0; reasons.append("MktStr=0/7 no HH+HL ⚠")
    if e50v is not None and last_c > e50v and ms < 7.0:
        ms = min(ms + 0.5, 7.0)

    return round(min(kl + pq + cs + ms, 25.0), 2), reasons

def _conv_volume(g: pd.DataFrame):
    """Volume — max 15 pts: BreakVol(7)+VolTrend(5)+PullVol(3)."""
    reasons = []
    if "Volume" not in g.columns or len(g) < 15:
        return 0.0, ["Volume: insufficient data"]
    vol = g["Volume"].values.astype(float)

    # ── (A) Breakout Volume — max 7 ──────────────────────────────────────────
    # Spike ratio vs 14-day average (excludes today); bonus if breakout context
    avg14 = float(np.mean(vol[-15:-1]))  # 14-bar avg before today
    today_vol = float(vol[-1])
    ratio = today_vol / max(avg14, 1.0)

    if ratio >= 2.5:
        bv = 7.0; reasons.append(f"BreakVol=7/7 climactic spike {ratio:.1f}x 🔥")
    elif ratio >= 2.0:
        bv = 6.0; reasons.append(f"BreakVol=6/7 strong breakout {ratio:.1f}x")
    elif ratio >= 1.5:
        bv = 5.0; reasons.append(f"BreakVol=5/7 good confirmation {ratio:.1f}x")
    elif ratio >= 1.2:
        bv = 4.0; reasons.append(f"BreakVol=4/7 decent vol {ratio:.1f}x")
    elif ratio >= 0.9:
        bv = 3.0; reasons.append(f"BreakVol=3/7 normal vol {ratio:.1f}x")
    else:
        bv = 1.0; reasons.append(f"BreakVol=1/7 weak vol {ratio:.1f}x ❌")

    # Bonus +1 if breakout (close > 20-day high) AND vol >= 2x (cap at 7)
    if len(g) >= 21:
        close = g["Close"].values
        high  = g["High"].values
        prior_high20 = float(np.max(high[-21:-1]))
        if close[-1] > prior_high20 and ratio >= 2.0:
            bv = min(bv + 1.0, 7.0)
            reasons[-1] += " +breakout bonus"

    # ── (B) Volume Trend — max 5 ─────────────────────────────────────────────
    # 5-bar avg vs 20-bar avg
    if len(vol) >= 20:
        v5  = float(np.mean(vol[-5:]))
        v20 = float(np.mean(vol[-20:]))
        vr  = v5 / max(v20, 1.0)
        if vr >= 1.30:
            vt = 5.0; reasons.append("VolTrend=5/5 expanding strongly ✓")
        elif vr >= 1.10:
            vt = 4.0; reasons.append("VolTrend=4/5 expanding")
        elif vr >= 0.90:
            vt = 3.0; reasons.append("VolTrend=3/5 flat/neutral")
        elif vr >= 0.70:
            vt = 2.0; reasons.append("VolTrend=2/5 contracting slightly")
        else:
            vt = 0.0; reasons.append("VolTrend=0/5 contracting sharply ⚠")
    else:
        vt = 2.0; reasons.append("VolTrend=2/5 insuf history")

    # ── (C) Pullback Volume — max 3 ──────────────────────────────────────────
    # Low-volume pullback after a prior advance = healthy; high vol = distribution
    if len(g) >= 10:
        close = g["Close"].values
        # Identify last 3 bars where price declined
        pb_vols = [vol[-i] for i in range(1, 6) if close[-i] < close[-i-1]]
        if pb_vols:
            pb_avg = float(np.mean(pb_vols))
            pb_r   = pb_avg / max(avg14, 1.0)
            if pb_r < 0.70:
                pv = 3.0; reasons.append("PullVol=3/3 low-vol pullback ✓")
            elif pb_r < 0.90:
                pv = 2.0; reasons.append("PullVol=2/3 moderate pullback vol")
            elif pb_r < 1.10:
                pv = 1.0; reasons.append("PullVol=1/3 avg pullback vol")
            else:
                pv = 0.0; reasons.append("PullVol=0/3 high-vol pullback ⚠")
        else:
            pv = 2.0; reasons.append("PullVol=2/3 no recent pullback")
    else:
        pv = 1.0; reasons.append("PullVol=1/3 insuf history")

    return round(min(bv + vt + pv, 15.0), 2), reasons


def _conv_delivery(
    deliv_series,
    today_pct,
    last_close: float,
    prev_close: float,
    fallback_used: bool,
):
    """
    Delivery sub-factor scorer.  Max = 15 pts (DelivPct=6, DelivTrend=5, PriceConfirm=4).

    v2.3.0 redesign:
    - DelivTrend is level-aware (L0<50 / L1 50-60 / L2>60).
    - "Sharply declining" only fires when drop is large AND current level is low.
    - Dominant-recent-strength override: if today near 5-day high and >=50, floor=3.
    - Choppy/indecisive: fires when 5-day range >=12pp and no clear direction.
    - DeliveryStrength composite (0-100): combines level, trend, recent-change.
    - PriceConfirm uses a scoring model (no fixed grid) to avoid contradictions.
    """
    reasons = []

    # ── Guard: missing data ───────────────────────────────────────────────────
    if today_pct is None or (deliv_series is None or len(deliv_series) == 0):
        reasons.append("DelivPct=0/6 no delivery data ⚠")
        reasons.append("DelivTrend=3/5 insufficient history (neutral)")
        reasons.append("PriceConfirm=2/4 mixed signals")
        if fallback_used:
            reasons.insert(0, "⚠ Delivery file for today unavailable. Last 5 day del% avg used.")
        return 0.0, reasons

    if fallback_used:
        reasons.append("⚠ Delivery file for today unavailable. Last 5 day del% avg used for calculation")

    # ══════════════════════════════════════════════════════════════════════════
    # A) DelivPct — max 6
    # ══════════════════════════════════════════════════════════════════════════
    if today_pct > 70:
        dp = 6.0; reasons.append(f"DelivPct=6/6 very high ({today_pct:.1f}%) 🔥")
    elif today_pct >= 60:
        dp = 5.0; reasons.append(f"DelivPct=5/6 high ({today_pct:.1f}%)")
    elif today_pct >= 50:
        dp = 3.0; reasons.append(f"DelivPct=3/6 moderate ({today_pct:.1f}%)")
    else:
        dp = 1.0; reasons.append(f"DelivPct=1/6 low ({today_pct:.1f}%) ⚠")

    # ══════════════════════════════════════════════════════════════════════════
    # B) DelivTrend — max 5  (level-aware, v2.3.0 engine)
    # ══════════════════════════════════════════════════════════════════════════
    vals = deliv_series.values.astype(float)
    n    = len(vals)

    # Sparkline: use the series exactly as returned by get_delivery_series(),
    # which already includes today_pct as the last element. Do NOT append
    # today_pct again here, otherwise the last value appears twice and the
    # oldest one is dropped in the 5-point window.
    _spark = " → ".join(f"{v:.1f}%" for v in vals[-5:])
    _sp    = f"  [{_spark}]"

    if n < 2:
        dt = 3.0
        reasons.append(f"DelivTrend=3/5 insufficient history (neutral){_sp}")
    else:
        moves      = [vals[i] - vals[i - 1] for i in range(1, n)]
        up_moves   = sum(1 for m in moves if m > 0)
        down_moves = sum(1 for m in moves if m < 0)
        net_change = vals[-1] - vals[0]
        highest    = float(max(vals))
        lowest     = float(min(vals))
        rng        = highest - lowest
        avg_val    = float(sum(vals) / n)
        today_v    = float(vals[-1])
        up_ratio   = up_moves   / max(len(moves), 1)
        down_ratio = down_moves / max(len(moves), 1)
        near_highest = today_v >= (highest - 1.0)

        # Early accumulation: flat start then spike in last 2 days
        early_accum = (
            n >= 4
            and abs(vals[-3] - vals[0]) / max(abs(vals[0]), 1.0) < 0.06
            and (vals[-1] - vals[-3]) >= 5.0
        )

        # Choppy: wide range but no dominant direction
        choppy = (rng >= 12.0) and not (up_ratio >= 0.6 or down_ratio >= 0.6)

        if early_accum:
            dt = 4.0
            reasons.append(f"DelivTrend=4/5 early accumulation (flat→spike){_sp}")
        elif choppy:
            dt = 3.0
            reasons.append(f"DelivTrend=3/5 choppy/indecisive{_sp}")
        elif up_ratio >= 0.75 and near_highest and today_v >= 50:
            dt = 5.0
            reasons.append(f"DelivTrend=5/5 strong accumulation ✓{_sp}")
        elif net_change >= 3.0 and today_v >= 50:
            dt = 4.0
            reasons.append(f"DelivTrend=4/5 moderate accumulation{_sp}")
        elif abs(vals[-1] - vals[0]) / max(abs(vals[0]), 1.0) < 0.04 or rng < 6.0:
            dt = 3.0
            reasons.append(f"DelivTrend=3/5 sideways/flat{_sp}")
        elif net_change < -3.0 and today_v >= 50:
            dt = 2.0
            reasons.append(f"DelivTrend=2/5 slightly declining ⚠{_sp}")
        elif net_change <= -8.0 and (today_v < 50 or avg_val < 50):
            dt = 1.0
            reasons.append(f"DelivTrend=1/5 sharply declining ❌{_sp}")
        else:
            dt = 2.0
            reasons.append(f"DelivTrend=2/5 slightly declining ⚠{_sp}")

        # Dominant Recent Strength Override:
        # today near 5-day high AND healthy level → never worse than neutral
        if near_highest and today_v >= 50 and dt < 3.0:
            dt = 3.0
            reasons[-1] = f"DelivTrend=3/5 choppy/indecisive (recovery override){_sp}"

        # Net change filter (±1)
        if net_change >= 8.0 and dt < 5.0:
            dt = min(dt + 1.0, 5.0)
            reasons[-1] += " (+net boost)"
        elif net_change <= -8.0 and dt > 1.0:
            dt = max(dt - 1.0, 1.0)
            reasons[-1] += " (-net penalty)"

    # ══════════════════════════════════════════════════════════════════════════
    # C) DeliveryStrength composite (0-100)
    # ══════════════════════════════════════════════════════════════════════════
    if today_pct > 70:     level_s = 85.0
    elif today_pct >= 60:  level_s = 70.0
    elif today_pct >= 50:  level_s = 55.0
    elif today_pct >= 40:  level_s = 35.0
    else:                  level_s = 20.0

    trend_map = {1: 20, 2: 40, 3: 55, 4: 75, 5: 90}
    trend_s   = float(trend_map.get(int(round(dt)), 55))

    delta = float(vals[-1] - vals[-2]) if n >= 2 else 0.0
    if delta >= 5.0:     recent_s = 90.0
    elif delta >= 2.0:   recent_s = 75.0
    elif delta >= 0.5:   recent_s = 60.0
    elif delta >= -0.5:  recent_s = 50.0
    elif delta >= -2.0:  recent_s = 35.0
    else:                recent_s = 20.0

    delivery_strength = max(0.0, min(100.0,
        0.5 * level_s + 0.3 * trend_s + 0.2 * recent_s
    ))
    if   delivery_strength >= 70: ds_band = "Strong"
    elif delivery_strength >= 55: ds_band = "Moderate"
    elif delivery_strength >= 45: ds_band = "Neutral"
    else:                         ds_band = "Weak"
    reasons.append(f"DelivStrength={delivery_strength:.0f}/100 ({ds_band})")

    # ══════════════════════════════════════════════════════════════════════════
    # D) PriceConfirm — max 4 (scoring model)
    # ══════════════════════════════════════════════════════════════════════════
    pct_chg = (last_close - prev_close) / max(abs(prev_close), 0.01) * 100
    if pct_chg > 0.5:    price_dir = "up"
    elif pct_chg < -0.5: price_dir = "down"
    else:                price_dir = "flat"

    p_score  = 2.0 if price_dir == "up" else (1.0 if price_dir == "flat" else 0.0)
    ds_score = {"Strong": 2.0, "Moderate": 1.5, "Neutral": 1.0, "Weak": 0.0}.get(ds_band, 1.0)

    if delta >= 5.0:    rc_score =  1.0
    elif delta >= 1.0:  rc_score =  0.5
    elif delta >= -1.0: rc_score =  0.2
    elif delta >= -4.0: rc_score = -0.5
    else:               rc_score = -1.0

    total_pc = p_score + ds_score + rc_score

    if total_pc >= 4.5:
        pc = 4.0; pc_txt = "PriceConfirm=4/4 strong accumulation 🔥"
    elif total_pc >= 3.5:
        pc = 3.0; pc_txt = "PriceConfirm=3/4 healthy / accumulation"
    elif total_pc >= 2.5:
        pc = 2.0; pc_txt = "PriceConfirm=2/4 mixed but acceptable"
    else:
        pc = 1.0; pc_txt = "PriceConfirm=1/4 weak/speculative rally ⚠"

    dir_sym   = {"up": "↑", "flat": "→", "down": "↓"}
    deliv_sym = "↑" if delta > 1.0 else ("↓" if delta < -1.0 else "→")
    reasons.append(
        f"{pc_txt}  [Price{dir_sym[price_dir]}+Deliv{deliv_sym} | score={total_pc:.1f}]"
    )

    total = min(dp + dt + pc, 15.0)
    return total, reasons


def _conv_trend(g: pd.DataFrame, dow_label: str):
    """Trend (EMA + Dow) — max 20 pts: EMAAlignment(12)+DowTheory(8)."""
    reasons = []
    e9  = (float(g["EMA9"].iloc[-1])
           if "EMA9"  in g.columns and pd.notna(g["EMA9"].iloc[-1])  else None)
    e21 = (float(g["EMA21"].iloc[-1])
           if "EMA21" in g.columns and pd.notna(g["EMA21"].iloc[-1]) else None)
    e50 = (float(g["EMA50"].iloc[-1])
           if "EMA50" in g.columns and pd.notna(g["EMA50"].iloc[-1]) else None)
    c   = float(g["Close"].iloc[-1])

    # (A) EMA Alignment — max 12
    if e9 is not None and e21 is not None and e50 is not None:
        if e9 > e21 > e50 and c > e9:
            ea = 12.0; reasons.append("EMA=12/12 9>21>50 price>EMA9 ✓")
        elif e9 > e21 > e50:
            ea = 10.0; reasons.append("EMA=10/12 9>21>50 price<=EMA9")
        elif e9 > e21 and c > e21:
            ea = 7.0;  reasons.append("EMA=7/12 9>21 partial align")
        elif c > e50:
            ea = 4.0;  reasons.append("EMA=4/12 price>EMA50 only")
        elif c < e50 and e50 < e21:
            ea = 1.0;  reasons.append("EMA=1/12 bearish alignment ⚠")
        else:
            ea = 2.0;  reasons.append("EMA=2/12 mixed EMA order")
    else:
        ea = 4.0; reasons.append("EMA=4/12 partial data")

    # EMA-respect bounce bonus: price touched EMA21 and bounced up (last 20 bars)
    # Counts bars where Low <= EMA21 <= High AND next bar closes higher
    if e21 is not None and "EMA21" in g.columns and len(g) >= 20 and ea < 12.0:
        ema21_arr = g["EMA21"].values.astype(float)
        low_arr   = g["Low"].values.astype(float)
        high_arr  = g["High"].values.astype(float)
        close_arr = g["Close"].values.astype(float)
        bounce_count = 0
        window_start = max(0, len(g) - 21)
        for i in range(window_start, len(g) - 1):
            if (pd.notna(ema21_arr[i])
                    and low_arr[i] <= ema21_arr[i] <= high_arr[i]
                    and close_arr[i + 1] > close_arr[i]):
                bounce_count += 1
        if bounce_count >= 2:
            ea = min(ea + 1.0, 12.0)
            reasons[-1] += f" +EMA21 respect ({bounce_count}x bounces)"

    # (B) Dow Theory phase — max 8
    phase   = _dow_phase(dow_label).lower()
    primary = ""
    for part in dow_label.split("|"):
        p = part.strip()
        if p.startswith("Primary:"):
            primary = p.replace("Primary:", "").strip().lower()
            break
    if "markup" in phase:
        dt = 8.0; reasons.append("Dow=8/8 Bull Markup phase")
    elif "accumulation" in phase and "bull" in phase:
        dt = 6.0; reasons.append("Dow=6/8 Bull Accumulation")
    elif "reversal watch (bullish)" in phase:
        dt = 5.5; reasons.append("Dow=6/8 Reversal Watch Bullish")
    elif "consolidation" in phase:
        dt = 4.0; reasons.append("Dow=4/8 Consolidation")
    elif "distribution" in phase and "bear" in phase:
        dt = 2.0; reasons.append("Dow=2/8 Bear Distribution ⚠")
    elif "markdown" in phase:
        dt = 1.0; reasons.append("Dow=1/8 Bear Markdown ⚠")
    elif "bear: accumulation" in phase:
        dt = 3.0; reasons.append("Dow=3/8 Bear Accum (bottoming?)")
    elif "uptrend" in primary:
        dt = 6.0; reasons.append("Dow=6/8 Primary Uptrend")
    else:
        dt = 3.0; reasons.append(f"Dow=3/8 phase={phase or 'unknown'}")

    return round(min(ea + dt, 20.0), 2), reasons

def _conv_rsi(g: pd.DataFrame, rsi_div: str):
    """RSI Momentum — max 10 pts: Zone(5)+Slope(3)+Divergence(2)."""
    reasons = []
    if "RSI14" not in g.columns or len(g) < 5:
        return 0.0, ["RSI: insufficient data"]
    rv = g["RSI14"].dropna()
    if len(rv) < 3:
        return 0.0, ["RSI: insufficient data"]
    rn = float(rv.iloc[-1])

    # (A) Zone — max 5
    if 55 <= rn <= 70:
        rz = 5.0; reasons.append(f"RSIZone=5/5 bullish ({rn:.1f})")
    elif 50 <= rn < 55:
        rz = 3.5; reasons.append(f"RSIZone=4/5 above50 ({rn:.1f})")
    elif 45 <= rn < 50:
        rz = 2.0; reasons.append(f"RSIZone=2/5 below50 ({rn:.1f})")
    elif rn > 70:
        rz = 3.0; reasons.append(f"RSIZone=3/5 overbought ({rn:.1f}) ⚠")
    else:
        rz = 0.5; reasons.append(f"RSIZone=1/5 weak ({rn:.1f}) ⚠")

    # (B) Slope (3-bar) — max 3
    r3 = float(rv.iloc[-4]) if len(rv) >= 4 else float(rv.iloc[0])
    rd = rn - r3
    if rd >= 3.0:
        rs = 3.0; reasons.append(f"RSISlope=3/3 rising (+{rd:.1f})")
    elif rd >= 1.0:
        rs = 2.0; reasons.append(f"RSISlope=2/3 slight rise (+{rd:.1f})")
    elif rd >= -1.0:
        rs = 1.0; reasons.append(f"RSISlope=1/3 flat ({rd:+.1f})")
    else:
        rs = 0.0; reasons.append(f"RSISlope=0/3 falling ({rd:.1f}) ⚠")

    # (C) Divergence — max 2
    dv = rsi_div.lower() if rsi_div else ""
    if "bullish" in dv:
        rdiv = 2.0; reasons.append("RSIDiv=2/2 bullish div ✓")
    elif "no divergence" in dv:
        rdiv = 2.0; reasons.append("RSIDiv=2/2 no div (clean)")
    elif "bearish" in dv:
        rdiv = 0.0; reasons.append("RSIDiv=0/2 bearish div ⚠")
    else:
        rdiv = 1.0; reasons.append("RSIDiv=1/2 insuf data")

    return round(min(rz + rs + rdiv, 10.0), 2), reasons

def _conv_macd(g: pd.DataFrame):
    """MACD — max 10 pts: Crossover(4)+Histogram(4)+ZeroLine(2)."""
    reasons = []
    req = ["MACD", "MACDSignal", "MACDHist"]
    if not all(c in g.columns for c in req) or len(g) < 3:
        return 0.0, ["MACD: insufficient data"]
    ms2 = g["MACD"].dropna()
    ss2 = g["MACDSignal"].dropna()
    hs2 = g["MACDHist"].dropna()
    if len(ms2) < 2 or len(ss2) < 2 or len(hs2) < 2:
        return 0.0, ["MACD: insufficient data"]
    mn = float(ms2.iloc[-1]); mp = float(ms2.iloc[-2])
    sn = float(ss2.iloc[-1]); sp = float(ss2.iloc[-2])
    hn = float(hs2.iloc[-1]); hp = float(hs2.iloc[-2])

    # (A) Crossover — max 4
    if mp <= sp and mn > sn:
        mc = 4.0; reasons.append("MACD=4/4 fresh bull cross ✓")
    elif mn > sn:
        mc = 2.5; reasons.append("MACD=3/4 above signal")
    else:
        mc = 0.0; reasons.append("MACD=0/4 below signal ⚠")

    # (B) Histogram — max 4 (increased from 3)
    if hn > 0 and hn > hp:
        mh = 4.0; reasons.append("Hist=4/4 positive & growing ✓")
    elif hn > hp:
        mh = 2.5; reasons.append("Hist=3/4 growing (still neg)")
    elif hn > 0:
        mh = 2.0; reasons.append("Hist=2/4 positive but shrinking")
    else:
        mh = 0.0; reasons.append("Hist=0/4 negative & falling ⚠")

    # (C) Zero Line — max 2 (reduced from 3)
    if mn > 0:
        mz = 2.0; reasons.append("Zero=2/2 above zero ✓")
    elif mn > -0.5 * abs(sn):
        mz = 1.0; reasons.append("Zero=1/2 approaching zero")
    else:
        mz = 0.0; reasons.append("Zero=0/2 below zero ⚠")

    return round(min(mc + mh + mz, 10.0), 2), reasons

def _conv_supertrend(st_green_10_3: bool, st_green_10_1: bool):
    """SuperTrend — max 5 pts: ST(10,3)=3 + ST(10,1)=2."""
    reasons = []
    score = 0.0
    if st_green_10_3:
        score += 3.0; reasons.append("ST(10,3)=3/3 green ✓")
    else:
        reasons.append("ST(10,3)=0/3 red ⚠")
    if st_green_10_1:
        score += 2.0; reasons.append("ST(10,1)=2/2 green ✓")
    else:
        reasons.append("ST(10,1)=0/2 red")
    return round(score, 2), reasons


# ── Column-merge helpers (v2.0.0 enhancement) ────────────────────────────────

def _supertrend_combined_label(st_green_10_3: bool, st_green_10_1: bool) -> str:
    """
    Merge SuperTrend(10,3) and SuperTrend(10,1) into a single descriptive cell value.
    Possible values returned:
      "(10,3) - Yes\n(10,1) - Yes"
      "(10,3) - Yes\n(10,1) - No"
      "(10,3) - No\n(10,1) - Yes"
      "(10,3) - No\n(10,1) - No"
    Uses plain newline so it renders as two lines in Excel (wrap-text)
    and as <br> when converted for HTML.
    """
    line1 = f"(10,3) - {'Yes' if st_green_10_3 else 'No'}"
    line2 = f"(10,1) - {'Yes' if st_green_10_1 else 'No'}"
    return f"{line1}\n{line2}"


def _support_resistance_combined_label(supports: str, resistances: str) -> str:
    """
    Merge the Supports and Resistances strings into one unified cell.
    At least two spaces separate the support block from the resistance block.
    Empty strings are handled gracefully.
    """
    parts = []
    if supports:
        parts.append(supports)
    if resistances:
        parts.append(resistances)
    if not parts:
        return ""
    # Two blank lines between sections when both are present
    return "\n\n".join(parts)


def _format_conviction_breakdown_html(
    pa_s: float, pa_r: list,
    vo_s: float, vo_r: list,
    de_s: float, de_r: list,
    tr_s: float, tr_r: list,
    ri_s: float, ri_r: list,
    ma_s: float, ma_r: list,
    st_s: float, st_r: list,
    total: float,
) -> str:
    """
    Build a rich, multi-line HTML conviction breakdown for display in the HTML report.
    Structure:  section header  →  bullet-point sub-factors  →  summary.

    Color coding:
      🟢 Green  : score >= 70 % of max
      🟡 Yellow : score >= 40 % of max
      🔴 Red    : score <  40 % of max
    """

    def _indicator(score: float, max_pts: float) -> str:
        pct = score / max(max_pts, 0.001)
        if pct >= 0.70:
            return "🟢"
        if pct >= 0.40:
            return "🟡"
        return "🔴"

    def _pct_color(score: float, max_pts: float) -> str:
        pct = score / max(max_pts, 0.001)
        if pct >= 0.70:
            return "#00c853"
        if pct >= 0.40:
            return "#ffd600"
        return "#ff1744"

    def _section(title: str, score: float, max_pts: float, reasons: list) -> str:
        ind = _indicator(score, max_pts)
        pct_val = round(score / max(max_pts, 0.001) * 100)
        color = _pct_color(score, max_pts)
        pct_badge = f' — <span style="color:{color};font-weight:bold;font-size:1.0em;">{pct_val}%</span>'
        header = (
            f'<b style="font-size:1.05em;letter-spacing:0.03em;">' +
            f'{ind}&nbsp;<span style="color:{color};">{title}</span>' +
            f'&nbsp;({score:.0f}/{max_pts:.0f}){pct_badge}</b>'
        )
        lines = [header]
        for r in reasons:
            lines.append(f"&nbsp;&nbsp;&nbsp;• {_humanise_reason(r)}")
        return "<br>".join(lines)

    def _humanise_reason(raw: str) -> str:
        """Convert compact reason tokens into readable English."""
        r = raw.strip()
        mappings = [
            ("KeyLvl=",      "Key Level     :"),
            ("PatQual=",     "Pattern       :"),
            ("CandleStr=",   "Candle        :"),
            ("MktStr=",      "Market Str    :"),
            ("BreakVol=",    "Break Volume  :"),
            ("VolTrend=",    "Volume Trend  :"),
            ("PullVol=",     "Pullback Vol  :"),
            ("DelivPct=",    "Deliv %       :"),
            ("DelivTrend=",  "Deliv Trend   :"),
        ("DelivStrength=","Deliv Strength:"),
            ("PriceConfirm=","Price Confirm :"),
            ("EMA=",         "EMA Alignment :"),
            ("Dow=",         "Dow Theory    :"),
            ("RSIZone=",     "RSI Zone      :"),
            ("RSISlope=",    "RSI Slope     :"),
            ("RSIDiv=",      "RSI Divergence:"),
            ("MACD=",        "MACD Cross    :"),
            ("Hist=",        "MACD Hist     :"),
            ("Zero=",        "MACD Zero Line:"),
            ("ST(10,3)=",    "SuperTrend(10,3):"),
            ("ST(10,1)=",    "SuperTrend(10,1):"),
        ]
        for prefix, label in mappings:
            if r.startswith(prefix):
                rest = r[len(prefix):]
                rest = rest.replace(" ", " → ", 1)
                return f"{label} {rest}"
        return r  # fallback: return as-is (covers ⚠ warning lines)

    # ── Build each section (order: structure → participation → momentum) ─────
    s_pa = _section("PRICE ACTION",  pa_s, 25, pa_r)
    s_tr = _section("TREND",         tr_s, 20, tr_r)
    s_vo = _section("VOLUME",        vo_s, 15, vo_r)
    s_de = _section("DELIVERY",      de_s, 15, de_r)
    s_ri = _section("RSI MOMENTUM",  ri_s, 10, ri_r)
    s_ma = _section("MACD",          ma_s, 10, ma_r)
    s_st = _section("SUPERTREND",    st_s,  5, st_r)

    # ── Overall conviction summary ────────────────────────────────────────────
    overall_ind   = _indicator(total, 100)
    overall_color = _pct_color(total, 100)
    score_html    = f'<span style="color:{overall_color};font-weight:bold;">{total:.1f}/100</span>'

    bias_parts = []
    if tr_s / 20 >= 0.65:
        bias_parts.append("strong trend")
    if ri_s / 10 >= 0.65:
        bias_parts.append("strong momentum")
    if ma_s / 10 >= 0.65:
        bias_parts.append("positive MACD")
    if vo_s / 15 >= 0.65:
        bias_parts.append("healthy volume")
    if de_s / 15 >= 0.65:
        bias_parts.append("strong delivery")

    caution_parts = []
    for r in pa_r:
        if "approaching res" in r.lower() or "near sup" in r.lower():
            caution_parts.append("price is near a key level")
            break
    if pa_s / 25 < 0.50:
        caution_parts.append("price action is mixed")
    if vo_s / 15 < 0.45:
        caution_parts.append("volume is below average")
    if de_s / 15 < 0.40:
        caution_parts.append("delivery is weak")

    if total >= 75:
        overall_text = "High conviction — all major factors aligned bullish."
    elif total >= 55:
        overall_text = "Moderate conviction"
        if bias_parts:
            overall_text += f" with {', '.join(bias_parts)}"
        if caution_parts:
            overall_text += f", but {'  '.join(caution_parts)}"
        overall_text += ". Wait for one more confirmation."
    else:
        overall_text = "Low conviction — watchlist only; do not trade yet."

    summary_block = (
        f"<br><b>{overall_ind} SUMMARY ({score_html})</b><br>"
        f"&nbsp;&nbsp;{overall_text}"
    )

    legend = (
        "<br><small>"
        "🟢 Strong / Bullish &nbsp;&nbsp;"
        "🟡 Neutral &nbsp;&nbsp;"
        "🔴 Weak / Bearish"
        "</small>"
    )

    return (
        s_pa + "<br>" +
        s_tr + "<br>" +
        s_vo + "<br>" +
        s_de + "<br>" +
        s_ri + "<br>" +
        s_ma + "<br>" +
        s_st +
        summary_block +
        legend
    )

def _format_conviction_breakdown_text(
    pa_s: float, pa_r: list,
    vo_s: float, vo_r: list,
    de_s: float, de_r: list,
    tr_s: float, tr_r: list,
    ri_s: float, ri_r: list,
    ma_s: float, ma_r: list,
    st_s: float, st_r: list,
    total: float,
) -> str:
    """
    Plain-text conviction breakdown for Excel cells (no HTML tags).
    Uses Unicode bullets and dashes for readability.
    """
    def _ind(score, max_pts):
        pct = score / max(max_pts, 0.001)
        if pct >= 0.70: return "● Strong"
        if pct >= 0.40: return "◐ Neutral"
        return "○ Weak"

    def _section(title, score, max_pts, reasons):
        header  = f"{_ind(score, max_pts)}  {title} ({score:.0f}/{max_pts:.0f})"
        bullets = "\n".join(f"  • {r}" for r in reasons)
        return header + "\n" + bullets

    lines = [
        _section("PRICE ACTION",  pa_s, 25, pa_r),
        _section("TREND",         tr_s, 20, tr_r),
        _section("VOLUME",        vo_s, 15, vo_r),
        _section("DELIVERY",      de_s, 15, de_r),
        _section("RSI MOMENTUM",  ri_s, 10, ri_r),
        _section("MACD",          ma_s, 10, ma_r),
        _section("SUPERTREND",    st_s,  5, st_r),
        f"\nOVERALL SCORE: {total:.1f}/100",
    ]
    return "\n\n".join(lines)

def compute_conviction_score(
    g: pd.DataFrame,
    dow_label: str,
    rsi_div: str,
    st_green_10_3: bool,
    st_green_10_1: bool,
    deliv_series: "pd.Series | None" = None,
    today_pct: "float | None" = None,
    prev_close: "float | None" = None,
    fallback_used: bool = False,
):
    """
    Compute a 0-100 conviction score for a shortlisted stock.

    Parameters
    ----------
    g             : full OHLCV + indicators DataFrame for the symbol (sorted asc)
    dow_label     : output of dow_theory_label()
    rsi_div       : output of compute_rsi_divergence()
    st_green_10_3 : SuperTrend(10,3) is green (trend==1)
    st_green_10_1 : SuperTrend(10,1) is green (trend==1)
    deliv_series  : pd.Series of last ≤5 delivery % values (oldest→today). None = unavailable.
    today_pct     : today's delivery % float. None = unavailable.
    prev_close    : close price 3 bars prior (used for PriceConfirm direction). None = auto.
    fallback_used : True if delivery file unavailable and cache avg was used instead.

    Returns
    -------
    (score: float, breakdown_html: str, breakdown_text: str)
    score     : 0-100, rounded to 1 decimal
    breakdown : one segment per factor group

    Conviction bands:
        >= 75  High conviction  — consider entry
        55-74  Moderate         — wait for 1 more confirmation
        < 55   Low              — watchlist only, do not trade
    """
    # Auto-derive prev_close if not supplied
    if prev_close is None:
        if len(g) >= 4:
            prev_close = float(g["Close"].iloc[-4])
        else:
            prev_close = float(g["Close"].iloc[0])
    last_close = float(g["Close"].iloc[-1])

    pa_s, pa_r = _conv_price_action(g)
    vo_s, vo_r = _conv_volume(g)
    de_s, de_r = _conv_delivery(
        deliv_series, today_pct, last_close, prev_close, fallback_used
    )
    tr_s, tr_r = _conv_trend(g, dow_label)
    ri_s, ri_r = _conv_rsi(g, rsi_div)
    ma_s, ma_r = _conv_macd(g)
    st_s, st_r = _conv_supertrend(st_green_10_3, st_green_10_1)

    total = round(pa_s + vo_s + de_s + tr_s + ri_s + ma_s + st_s, 1)

    # Fix 6 — Hard cap for stocks in a confirmed bearish primary trend.
    # A Bear: Markdown or Bear: Distribution stock cannot legitimately score
    # above 58; a counter-trend bounce is not an actionable buy signal.
    _cap_applied = False
    if ("Bear: Markdown" in dow_label or "Bear: Distribution" in dow_label) and total > 58.0:
        total = 58.0
        _cap_applied = True

    breakdown_html = _format_conviction_breakdown_html(
        pa_s, pa_r, vo_s, vo_r, de_s, de_r,
        tr_s, tr_r, ri_s, ri_r, ma_s, ma_r, st_s, st_r, total,
    )
    breakdown_text = _format_conviction_breakdown_text(
        pa_s, pa_r, vo_s, vo_r, de_s, de_r,
        tr_s, tr_r, ri_s, ri_r, ma_s, ma_r, st_s, st_r, total,
    )
    if _cap_applied:
        cap_note_html = (
            "<b>⚠ Score capped at 58 — Bearish primary trend "
            "(Bear: Markdown / Distribution).<br>"
            "Do not trade against the primary trend.</b><br><br>"
        )
        cap_note_text = (
            "⚠ Score capped at 58 — Bearish primary trend "
            "(Bear: Markdown/Distribution).\n"
            "Do not trade against the primary trend.\n\n"
        )
        breakdown_html = cap_note_html + breakdown_html
        breakdown_text = cap_note_text + breakdown_text

    return round(total, 1), breakdown_html, breakdown_text, pa_s, vo_s, de_s, tr_s, ri_s, ma_s, st_s

def _build_stock_row(
    symbol: str,
    g: pd.DataFrame,
    latest,
    history: pd.DataFrame,
    cfg: Config,
    s: requests.Session,
    delivery_dict: dict = None,
    deliv_fallback_used: bool = False,
    corp_actions: list = None,
    index_map: dict = None,
) -> dict:
    """Compute all common indicator fields for a qualifying stock row."""
    pattern = detect_candlestick(g)
    vol_above = bool(latest["Volume"] > latest["VolAvg14"]) if pd.notna(latest["VolAvg14"]) else False
    macd_above = bool(latest["MACD"] > latest["MACDSignal"])
    st_green_10_3 = bool(latest["SupertrendGreen_10_3"])
    st_green_10_1 = bool(latest["SupertrendGreen_10_1"])
    momentum = momentum_label(latest)
    dow = dow_theory_label(g)

    # 52-week high / low
    hist_symbol = history[history["Symbol"] == symbol].sort_values("Date")
    if not hist_symbol.empty:
        last_date = hist_symbol["Date"].max()
        cutoff_52 = last_date - pd.Timedelta(days=365)
        last_52w = hist_symbol[hist_symbol["Date"] >= cutoff_52]
    else:
        last_52w = hist_symbol
    if not last_52w.empty:
        hi_52w = float(last_52w["High"].max())
        lo_52w = float(last_52w["Low"].min())
        val_52w = f"H: {hi_52w:.2f}<br>L: {lo_52w:.2f}"
    else:
        val_52w = ""

    rsi_div = compute_rsi_divergence(g["Close"], g["RSI14"])

    # ── Delivery data for conviction scorer ───────────────────────────────────
    _ddict = delivery_dict if delivery_dict is not None else {}
    # RESOLVED_TRADE_DATE is the actual analysis date (e.g. 2026-04-16).
    # Using TODAY_IST here would make the "prior rows" filter include
    # today's cached delivery row AND then re-append today_pct, producing
    # a duplicate last value in the DelivTrend sparkline.
    _deliv_series, _today_pct, _fallback = get_delivery_series(
        symbol, RESOLVED_TRADE_DATE, _ddict, deliv_fallback_used
    )

    conviction_score, conviction_breakdown_html, conviction_breakdown_text, \
        _pa_s, _vo_s, _de_s, _tr_s, _ri_s, _ma_s, _st_s = compute_conviction_score(
        g, dow, rsi_div, st_green_10_3, st_green_10_1,
        deliv_series=_deliv_series,
        today_pct=_today_pct,
        fallback_used=_fallback,
    )
    chart_pattern = detect_chart_patterns(g)
    supports, resistances = compute_pivots_support_resistance(g)

    news_html, news_text = fetch_recent_news(
        symbol,
        latest.get("CompanyName", symbol),
        cfg,
        s,
    )
    # Corporate actions block - prepended to news cell
    ca_html = ""
    ca_text = ""
    _ICON_RED   = chr(0x1F534)  # red circle
    _ICON_AMB   = chr(0x1F7E1)  # yellow circle
    _ICON_GRN   = chr(0x1F7E2)  # green circle
    _ICON_BANK  = chr(0x1F3E6)  # bank building
    _ICON_CHECK = chr(0x2705)   # check mark
    def _ca_urgency(days):
        if days <= 2:  return _ICON_RED
        if days <= 6:  return _ICON_AMB
        return _ICON_GRN
    _CA_COL = {
        "DIVIDEND": "#00c853", "BONUS": "#00c853",
        "SPLIT": "#ffd600",    "BUYBACK": "#ffd600", "RIGHTS": "#ffd600",
        "AGM": "#90a4ae",      "EARNINGS": "#40c4ff",
        "CORPORATE ACTION": "#90a4ae",
    }
    if corp_actions is not None:
        if corp_actions:
            _sca = sorted(corp_actions, key=lambda x: x["days_away"])
            _rows_h, _rows_t = [], []
            for ca in _sca:
                kind = ca.get("kind", "CORPORATE ACTION")
                icon = _ca_urgency(ca["days_away"])
                col  = _CA_COL.get(kind, "#90a4ae")
                dlbl = ("TODAY" if ca["days_away"] == 0
                        else f"in {ca['days_away']} day{'s' if ca['days_away']!=1 else ''}")
                _rows_h.append(
                    '<div style="margin:2px 0;padding:3px 6px;'
                    f'border-left:3px solid {col};font-size:11px;">'
                    f'{icon} <b style="color:{col}">{kind}</b> &#8212; {ca["event"]}<br>'
                    f'&nbsp;&nbsp;Ex-Date: <b>{ca["ex_date"]}</b>'
                    f'&nbsp;&middot;&nbsp;Rec-Date: {ca["rec_date"]}'
                    f'&nbsp;<span style="color:{col}">({dlbl})</span></div>'
                )
                _rows_t.append(
                    f'[{icon} {kind}] {ca["event"]} | '
                    f'Ex: {ca["ex_date"]} | Rec: {ca["rec_date"]} ({dlbl})'
                )
            ca_html = (
                '<div style="background:#1a1a2e;border:1px solid #444;'
                'border-radius:4px;padding:6px 8px;margin-bottom:6px;">'
                f'<div style="font-size:11px;font-weight:bold;color:#fff;'
                f'margin-bottom:4px;">{_ICON_BANK} CORPORATE ACTIONS (upcoming)</div>'
                + "".join(_rows_h) + "</div>"
            )
            ca_text = (
                _ICON_BANK + " CORPORATE ACTIONS\n"
                + "\n".join(_rows_t) + "\n\n"
            )
        else:
            _la = getattr(cfg, "corporate_action_lookahead_days", 14)
            ca_html = (
                '<div style="background:#1a2e1a;border:1px solid #2e7d32;'
                'border-radius:4px;padding:5px 8px;margin-bottom:6px;'
                f'font-size:11px;color:#81c784;">'
                f'{_ICON_CHECK} No corporate actions in the next {_la} days</div>'
            )
            ca_text = f"{_ICON_CHECK} No corporate actions in the next {_la} days\n\n"
    # Prepend CA block as first list item so HTML renderer (<ul><li>) sees a list
    if ca_html:
        news_html = [ca_html] + (news_html if isinstance(news_html, list) else [news_html])
    if ca_text:
        news_text = ca_text + (news_text if isinstance(news_text, str) else " | ".join(news_text))

    return {
        "Date": latest["Date"].date().isoformat(),
        "Symbol": symbol,
        "Company": latest.get("CompanyName", symbol),
        "Index": (index_map.get(symbol.upper(), "Nifty 100")
                  if index_map else "Nifty 100"),
        "Last Close": round(float(latest["Close"]), 2),
        "52WH/52WL": val_52w,
        "EMA9":  round(float(latest.get("EMA9",  float("nan"))), 2) if pd.notna(latest.get("EMA9",  float("nan"))) else float("nan"),
        "EMA10": round(float(latest.get("EMA10", float("nan"))), 2) if pd.notna(latest.get("EMA10", float("nan"))) else float("nan"),
        "EMA20": round(float(latest.get("EMA20", float("nan"))), 2) if pd.notna(latest.get("EMA20", float("nan"))) else float("nan"),
        "EMA21": round(float(latest.get("EMA21", float("nan"))), 2) if pd.notna(latest.get("EMA21", float("nan"))) else float("nan"),
        "EMA50": round(float(latest.get("EMA50", float("nan"))), 2) if pd.notna(latest.get("EMA50", float("nan"))) else float("nan"),
        "EMA200":round(float(latest.get("EMA200",float("nan"))), 2) if pd.notna(latest.get("EMA200",float("nan"))) else float("nan"),
        "Candlestick Pattern": pattern,
        "Volume": int(latest["Volume"]),
        "Volume14Avg": round(float(latest["VolAvg14"]), 2) if pd.notna(latest["VolAvg14"]) else float("nan"),
        "VolumeAbove14Avg": "Yes" if vol_above else "No",
        "RSI14": round(float(latest["RSI14"]), 2) if pd.notna(latest["RSI14"]) else float("nan"),
        "RSI Divergence": rsi_div,
        "MACDAboveSignal": "Yes" if macd_above else "No",
        "SuperTrend(10,3) (10,1) Green": _supertrend_combined_label(st_green_10_3, st_green_10_1),
        "Momentum": momentum,
        "DowTheory": dow,
        "DowTheoryTrend": _dow_trend(dow),
        "DowTheoryPhase": _dow_phase(dow),
        "Chart Pattern": chart_pattern,
        "Support and Resistance": _support_resistance_combined_label(supports, resistances),
        "Conviction Score": conviction_score,
        "Conviction Explanation": conviction_breakdown_html,
        "Conviction Explanation Text": conviction_breakdown_text,
        "DeliveryPct":     round(float(_today_pct), 2) if _today_pct is not None and str(_today_pct) not in ("", "nan") else float("nan"),
        "PAScore":         round(_pa_s, 1),
        "TrendScore":      round(_tr_s, 1),
        "VolumeScore":     round(_vo_s, 1),
        "DeliveryScore":   round(_de_s, 1),
        "RSIScore":        round(_ri_s, 1),
        "MACDScore":       round(_ma_s, 1),
        "STScore":         round(_st_s, 1),
        "RecentNewsHtml":  news_html,
        "RecentNewsText":  news_text,
    }



# ── Pullback-to-EMA scanner ───────────────────────────────────────────────────

def scan_pullback_to_ema(
    g: pd.DataFrame,
    ema_col: str,
    pullback_window: int,
    last_cross_date: Optional[pd.Timestamp] = None,
    max_cross_age_bars: int = 60,
) -> tuple:
    """
    Returns (qualified: bool, touch_price: float, ema_level: float, pullback_days: int)

    Conditions (all must be True):
      1. EMA alignment over the lookback window:
         - EMA9  pullback : EMA9 > EMA21 > EMA50 on every bar in window
         - EMA21 pullback : EMA21 > EMA50 on every bar in window
                            (allows EMA9 < EMA21 OR EMA9 > EMA21)
      2. Clear HH+HL structure in the bars just BEFORE the pullback window.
      3. Today's Close OR Today's Low is within 1% of the target EMA level.
      4. Volume during the pullback period is below the preceding 14-bar average
         (volume contraction confirms a healthy, low-conviction pullback).
      5. Fix 8 — Crossover age limit: the underlying EMA crossover that created
         this pullback opportunity must have occurred within max_cross_age_bars
         bars (default 60). Pullbacks against a stale crossover (e.g. one that
         happened 4+ months ago) are structurally weak and excluded.
    """
    required = [ema_col, "EMA9", "EMA21", "EMA50", "Close", "High", "Low", "Volume"]
    if not all(c in g.columns for c in required):
        return False, 0.0, 0.0, 0
    if len(g) < pullback_window + 20:
        return False, 0.0, 0.0, 0

    # Fix 8 — Crossover age limit.
    # If the caller supplies the date of the most recent crossover for this symbol,
    # discard the pullback signal when that crossover is older than max_cross_age_bars.
    # A pullback to an EMA that crossed over 3–6 months ago is not a fresh signal.
    if last_cross_date is not None and not pd.isnull(last_cross_date):
        dates = g["Date"].values if "Date" in g.columns else pd.Series(g.index)
        latest_date = pd.Timestamp(g.index[-1]) if "Date" not in g.columns                       else pd.Timestamp(g["Date"].iloc[-1])
        cross_ts = pd.Timestamp(last_cross_date)
        if cross_ts <= latest_date:
            # Count trading bars between crossover and today
            if "Date" in g.columns:
                bars_since = int((g["Date"] >= cross_ts).sum())
            else:
                bars_since = int((pd.Series(g.index) >= cross_ts).sum())
            if bars_since > max_cross_age_bars:
                return False, 0.0, 0.0, 0

    # ── 1. EMA alignment ─────────────────────────────────────────────────────
    window = g.tail(pullback_window + 20)
    e9_s  = window["EMA9"].values
    e21_s = window["EMA21"].values
    e50_s = window["EMA50"].values

    if ema_col == "EMA9":
        aligned = all(
            e9_s[i] > e21_s[i] > e50_s[i]
            for i in range(len(e9_s))
            if not (np.isnan(e9_s[i]) or np.isnan(e21_s[i]) or np.isnan(e50_s[i]))
        )
    else:  # EMA21
        aligned = all(
            e21_s[i] > e50_s[i]
            for i in range(len(e21_s))
            if not (np.isnan(e21_s[i]) or np.isnan(e50_s[i]))
        )
    if not aligned:
        return False, 0.0, 0.0, 0

    # ── 2. HH+HL structure in bars before the pullback window ────────────────
    pre_window = g.iloc[-(pullback_window + 20): -pullback_window]
    if len(pre_window) < 10:
        return False, 0.0, 0.0, 0
    highs_pre = pre_window["High"].values
    lows_pre  = pre_window["Low"].values
    sh, sl = [], []
    for i in range(1, len(highs_pre) - 1):
        if highs_pre[i] > highs_pre[i - 1] and highs_pre[i] > highs_pre[i + 1]:
            sh.append(highs_pre[i])
        if lows_pre[i] < lows_pre[i - 1] and lows_pre[i] < lows_pre[i + 1]:
            sl.append(lows_pre[i])
    hh_ok = len(sh) >= 2 and sh[-1] > sh[-2]
    hl_ok = len(sl) >= 2 and sl[-1] > sl[-2]
    if not (hh_ok and hl_ok):
        return False, 0.0, 0.0, 0

    # ── 3. Today's close or low within 1% of target EMA ─────────────────────
    latest     = g.iloc[-1]
    ema_level  = float(latest[ema_col])
    last_close = float(latest["Close"])
    last_low   = float(latest["Low"])
    tol        = ema_level * 0.01
    close_near = abs(last_close - ema_level) <= tol
    low_near   = abs(last_low   - ema_level) <= tol
    if not (close_near or low_near):
        return False, 0.0, 0.0, 0
    touch_price = last_close if close_near else last_low

    # ── 4. Volume contraction during the pullback ─────────────────────────────
    closes   = g["Close"].values
    pb_start = 1
    for k in range(2, min(pullback_window, len(g))):
        if closes[-k] > closes[-k - 1]:
            pb_start = k
    pb_start = max(3, pb_start)

    vol_pb   = float(g["Volume"].iloc[-pb_start:].mean())
    vol_base = float(g["Volume"].iloc[-(pb_start + 14): -pb_start].mean())
    if not ((vol_base > 0) and (vol_pb < vol_base)):
        return False, 0.0, 0.0, 0

    return True, round(touch_price, 2), round(ema_level, 2), pb_start

# ─────────────────────────────────────────────────────────────────────────────
# Signal Tracker — log, update, rotate
# ─────────────────────────────────────────────────────────────────────────────

SIGNAL_LOG_COLS = [
    # Group 1 — Identity
    "signal_id", "signal_date", "signal_type", "symbol", "company",
    # Group 2 — Price on signal day
    "signal_close", "signal_open", "signal_high", "signal_low",
    "signal_volume", "signal_vol_ratio",
    # Group 3 — Technical levels on signal day
    "signal_ema9", "signal_ema21", "signal_ema50", "signal_ema200",
    "signal_rsi", "signal_macd_above", "signal_supertrend",
    "signal_delivery_pct", "signal_dow_phase",
    # Group 4 — Conviction on signal day
    "signal_conviction_score", "signal_conviction_band",
    "signal_pa_score", "signal_trend_score", "signal_volume_score",
    "signal_delivery_score", "signal_rsi_score",
    "signal_macd_score", "signal_st_score",
    # Group 5 — Live tracking (updated each run)
    "last_updated", "days_since_signal", "current_close",
    "pct_return", "abs_return",
    "max_high_since", "max_pct_gain",
    "min_low_since", "max_pct_drawdown",
    # Group 6 — Signal health (updated each run)
    "ema_still_aligned", "cross_back_date", "current_conviction",
    "current_rsi", "current_delivery_pct", "ribbon_active", "status",
    # Group 7 — Risk levels
    "atr_on_signal", "suggested_stop", "suggested_target_1",
    "suggested_target_2", "stop_hit", "target1_hit", "target2_hit",
    # Group 8 — Audit trail (Fix 11)
    "days_aligned", "exit_reason",
]


def _load_signal_log() -> pd.DataFrame:
    """Load signal log CSV or return empty frame with correct columns."""
    if SIGNAL_LOG_PATH.exists():
        try:
            df = pd.read_csv(SIGNAL_LOG_PATH, parse_dates=["signal_date", "last_updated", "cross_back_date"])
            # Ensure all expected columns exist (forward-compat)
            for col in SIGNAL_LOG_COLS:
                if col not in df.columns:
                    df[col] = None
            return df[SIGNAL_LOG_COLS]
        except Exception as exc:
            logger.warning("Could not read signal log: %s", exc)
    return pd.DataFrame(columns=SIGNAL_LOG_COLS)


def rotate_signal_log(cfg: Config) -> None:
    """
    Rotate nifty100_signal_log.csv when it reaches cfg.signal_log_max_mb MB.

    Rotation steps:
      1. Copy active file to data/archives/signal_log_archive_YYYYMMDD_HHMMSS.csv
      2. Compress to .zip and delete the plain CSV copy.
      3. Retain only rows that are Active OR younger than cfg.signal_log_archive_days days
         in the live signal_log.csv.
    """
    import zipfile as _zf

    if not SIGNAL_LOG_PATH.exists():
        return
    size_mb = SIGNAL_LOG_PATH.stat().st_size / (1024 * 1024)
    if size_mb < cfg.signal_log_max_mb:
        return

    logger.info(
        "Signal log size %.1f MB >= %d MB threshold — rotating.",
        size_mb, cfg.signal_log_max_mb,
    )

    SIGNAL_ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    ts        = NOW_IST.strftime("%Y%m%d_%H%M%S")
    arch_csv  = SIGNAL_ARCHIVE_DIR / f"signal_log_archive_{ts}.csv"
    arch_zip  = SIGNAL_ARCHIVE_DIR / f"signal_log_archive_{ts}.zip"

    # Step 1+2: copy → zip → delete plain copy
    import shutil as _sh
    _sh.copy2(SIGNAL_LOG_PATH, arch_csv)
    with _zf.ZipFile(arch_zip, "w", _zf.ZIP_DEFLATED) as zf:
        zf.write(arch_csv, arch_csv.name)
    arch_csv.unlink()
    logger.info("Signal log archived to %s", arch_zip)

    # Step 3: trim live file — keep Active or recent rows
    try:
        df = _load_signal_log()
        if not df.empty:
            cutoff = pd.Timestamp(RESOLVED_TRADE_DATE) - pd.Timedelta(
                days=cfg.signal_log_archive_days
            )
            mask = (df["status"] == "Active") | (
                pd.to_datetime(df["signal_date"]) >= cutoff
            )
            df_keep = df[mask].reset_index(drop=True)
            df_keep.to_csv(SIGNAL_LOG_PATH, index=False)
            logger.info(
                "Signal log trimmed: %d rows retained after rotation.", len(df_keep)
            )
    except Exception as exc:
        logger.warning("Signal log trim error: %s", exc)


def _atr14(g: pd.DataFrame) -> float:
    """Compute ATR(14) for the last bar of g."""
    try:
        hi = g["High"].values.astype(float)
        lo = g["Low"].values.astype(float)
        cl = g["Close"].values.astype(float)
        tr = [max(hi[i] - lo[i],
                  abs(hi[i] - cl[i-1]),
                  abs(lo[i] - cl[i-1]))
              for i in range(1, len(g))]
        if len(tr) >= 14:
            return float(sum(tr[-14:]) / 14)
        return float(sum(tr) / max(len(tr), 1))
    except Exception:
        return 0.0


def log_new_signals(
    rows_21_50:      list,
    rows_9_21:       list,
    rows_pb_ema9:    list,
    rows_pb_ema21:   list,
    rows_ribbon_new: list,
    stock_data: dict,   # {symbol: g_dataframe} — kept for ATR calc only
    cfg: Config,
) -> None:
    """
    Append new EMA crossover / pullback / ribbon signals to nifty100_signal_log.csv.
    Reads all indicator fields directly from the report row dict (which was
    already computed by _build_stock_row) to avoid re-computing on raw history.
    Signal types logged:
        EMA21x50       – EMA 21 crossed above EMA 50
        EMA9x21        – EMA 9 crossed above EMA 21
        PBtoEMA9       – Pullback to (or nearing) EMA 9
        PBtoEMA21      – Pullback to (or nearing) EMA 21
        RibbonNew      – EMA Ribbon new entry (10>20>50>200 turned true today)
    """
    rotate_signal_log(cfg)
    # ── Fix 16A: skip signal insertion on weekends/holidays — no new signals fire ──
    if RESOLVED_TRADE_DATE.weekday() >= 5:
        logger.info(
            "Signal log: weekend run detected (%s). RESOLVED_TRADE_DATE points to "
            "last trading day — proceeding with signal insertion.",
            RESOLVED_TRADE_DATE,
        )
    # Weekend guard (return) removed — sig_id dedup + update_signal_tracker()
    # is_trading_day guard are sufficient. No duplicate risk.
    existing     = _load_signal_log()
    today_str    = RESOLVED_TRADE_DATE.strftime("%Y-%m-%d")
    existing_ids = set(existing["signal_id"].astype(str)) if not existing.empty else set()

    def _safe_float(val, default=float("nan")):
        try:
            f = float(val)
            return f if not (f != f) else default   # NaN check
        except Exception:
            return default

    def _parse_subscore(text: str, label: str) -> float:
        """Extract numeric score from conviction text e.g. 'PRICE ACTION : 14/25' → 14.0"""
        import re as _re
        m = _re.search(rf"{label}\s*[-:|=]?\s*([\d.]+)\s*/", str(text), _re.IGNORECASE)
        return float(m.group(1)) if m else float("nan")

    def _st_green(st_label: str) -> bool:
        """True if SuperTrend(10,3) is green from the combined label string."""
        if not st_label:
            return False
        parts = str(st_label).lower()
        # Label format: "10,3 - Yes / 10,1 - Yes" or "Yes / Yes" etc.
        if "10,3" in parts:
            segment = parts.split("10,3")[1][:20]
            return "yes" in segment
        return "yes" in parts[:10]

    new_rows = []
    for stype, rows in [
        ("EMA21x50",  rows_21_50),
        ("EMA9x21",   rows_9_21),
        ("PBtoEMA9",  rows_pb_ema9),
        ("PBtoEMA21", rows_pb_ema21),
        ("RibbonNew", rows_ribbon_new),
    ]:
        for row in rows:
            sym    = row.get("Symbol", "")
            sig_id = f"{sym}_{stype}_{today_str.replace('-','')}"
            if sig_id in existing_ids:
                continue

            # ── Group 2: OHLCV — read from row (already computed) ───────────
            g = stock_data.get(sym)
            lat = g.iloc[-1] if g is not None and len(g) >= 1 else None

            if lat is not None:
                close = _safe_float(lat.get("Close", 0.0), 0.0)
                open_ = _safe_float(lat.get("Open",  close), close)
                high_ = _safe_float(lat.get("High",  close), close)
                low_  = _safe_float(lat.get("Low",   close), close)
            else:
                close = _safe_float(row.get("Last Close", 0.0), 0.0)
                open_ = high_ = low_ = close
            volume   = int(row.get("Volume", 0) or 0)
            vol14avg = _safe_float(row.get("Volume14Avg", 1), 1) or 1.0
            vol_ratio = round(volume / vol14avg, 2) if vol14avg > 0 else float("nan")
            delivery_pct = _safe_float(row.get("DeliveryPct"))

            # ── Group 3: Technical levels — all from row dict ───────────────
            ema9   = _safe_float(row.get("EMA9"))
            ema21  = _safe_float(row.get("EMA21"))
            ema50  = _safe_float(row.get("EMA50"))
            ema200 = _safe_float(row.get("EMA200"))
            rsi14  = _safe_float(row.get("RSI14"))
            macd_above  = str(row.get("MACDAboveSignal", "")).strip().lower() == "yes"
            st_label    = str(row.get("SuperTrend(10,3) (10,1) Green", ""))
            supertrend  = _st_green(st_label)
            deliv_pct   = _safe_float(row.get("DeliveryPct"))
            dow_phase   = str(row.get("DowTheoryPhase", "") or "")

            # ── Group 4: Conviction score + sub-scores ──────────────────────
            cs   = _safe_float(row.get("Conviction Score", 0), 0)
            band = "High" if cs >= 75 else "Moderate" if cs >= 55 else "Low"
            pa_score   = _safe_float(row.get("PAScore"))
            tr_score   = _safe_float(row.get("TrendScore"))
            vo_score   = _safe_float(row.get("VolumeScore"))
            de_score   = _safe_float(row.get("DeliveryScore"))
            ri_score   = _safe_float(row.get("RSIScore"))
            ma_score   = _safe_float(row.get("MACDScore"))
            st_score_v = _safe_float(row.get("STScore"))

            # ── ATR — still computed from history slice (only use for this) ─
            g   = stock_data.get(sym)
            atr = _atr14(g) if g is not None and len(g) >= 2 else 0.0

            # Fix 12 — anchor stop/targets to nearest S/R zone;
            # ATR multiples used only as fallback when no zone is found.
            import re as _re12
            def _sr_levels_fix12(sr_str, above, ref):
                levels = []
                for _m in _re12.findall(r"\d+\.?\d*", str(sr_str).replace(",", "")):
                    try:
                        _v = float(_m)
                        if above and _v > ref * 1.005:
                            levels.append(_v)
                        elif not above and _v < ref * 0.995:
                            levels.append(_v)
                    except ValueError:
                        pass
                return sorted(levels, key=lambda x: abs(x - ref))

            _sr_str  = str(row.get("Support and Resistance", "") or "")
            _sups    = _sr_levels_fix12(_sr_str, above=False, ref=close)
            _ress    = _sr_levels_fix12(_sr_str, above=True,  ref=close)
            _stop    = round(_sups[0]    if _sups              else close - 1.5 * atr, 2) if (atr or _sups) else float("nan")
            _tgt1    = round(_ress[0]    if _ress              else close * 1.05,       2)
            _tgt2    = round(_ress[1]    if len(_ress) > 1     else close * 1.10,       2)

            new_rows.append({
                # Group 1
                "signal_id":               sig_id,
                "signal_date":             today_str,
                "signal_type":             stype,
                "symbol":                  sym,
                "company":                 row.get("Company", ""),
                # Group 2
                "signal_close":            round(close, 2),
                "signal_open":             round(open_, 2),
                "signal_high":             round(high_, 2),
                "signal_low":              round(low_, 2),
                "signal_volume":           volume,
                "signal_vol_ratio":        vol_ratio,
                # Group 3
                "signal_ema9":             round(ema9, 2)   if ema9 == ema9   else float("nan"),
                "signal_ema21":            round(ema21, 2)  if ema21 == ema21 else float("nan"),
                "signal_ema50":            round(ema50, 2)  if ema50 == ema50 else float("nan"),
                "signal_ema200":           round(ema200, 2) if ema200 == ema200 else float("nan"),
                "signal_rsi":              round(rsi14, 2)  if rsi14 == rsi14 else float("nan"),
                "signal_macd_above":       macd_above,
                "signal_supertrend":       supertrend,
                "signal_delivery_pct":     round(deliv_pct, 2) if deliv_pct == deliv_pct else float("nan"),
                "signal_dow_phase":        dow_phase,
                # Group 4
                "signal_conviction_score": round(cs, 1),
                "signal_conviction_band":  band,
                "signal_pa_score":         round(pa_score, 1)   if pa_score == pa_score   else float("nan"),
                "signal_trend_score":      round(tr_score, 1)   if tr_score == tr_score   else float("nan"),
                "signal_volume_score":     round(vo_score, 1)   if vo_score == vo_score   else float("nan"),
                "signal_delivery_score":   round(de_score, 1)   if de_score == de_score   else float("nan"),
                "signal_rsi_score":        round(ri_score, 1)   if ri_score == ri_score   else float("nan"),
                "signal_macd_score":       round(ma_score, 1)   if ma_score == ma_score   else float("nan"),
                "signal_st_score":         round(st_score_v, 1) if st_score_v == st_score_v else float("nan"),
                # Group 5 — seed with signal-day values
                "last_updated":            today_str,
                "days_since_signal":       0,
                "current_close":           round(close, 2),
                "pct_return":              0.0,
                "abs_return":              0.0,
                "max_high_since":          round(high_, 2),
                "max_pct_gain":            0.0,
                "min_low_since":           round(low_, 2),
                "max_pct_drawdown":        0.0,
                # Group 6 — seed with signal-day values
                "ema_still_aligned":       True,
                "cross_back_date":         None,
                "current_conviction":      round(cs, 1),
                "current_rsi":             round(rsi14, 2) if rsi14 == rsi14 else float("nan"),
                "current_delivery_pct":    round(deliv_pct, 2) if deliv_pct == deliv_pct else float("nan"),
                "ribbon_active":           bool(row.get("RibbonActive", False)),
                "status":                  "Active",
                # Group 7 — S/R-anchored risk levels (Fix 12); ATR fallback
                "atr_on_signal":           round(atr, 2),
                "suggested_stop":          _stop,
                "suggested_target_1":      _tgt1,
                "suggested_target_2":      _tgt2,
                "stop_hit":                False,
                "target1_hit":             False,
                "target2_hit":             False,
                # Group 8 — Audit trail (Fix 11)
                "days_aligned":            0,
                "exit_reason":             "",
            })

    if not new_rows:
        logger.info("Signal log: no new signals to append today.")
        return

    new_df = pd.DataFrame(new_rows, columns=SIGNAL_LOG_COLS)
    if not existing.empty:
        combined = pd.concat([existing, new_df], ignore_index=True)
    else:
        combined = new_df
    combined.to_csv(SIGNAL_LOG_PATH, index=False)
    logger.info("Signal log: appended %d new signals. Total rows: %d",
                len(new_rows), len(combined))
    logger.info("Signal log written/updated: %s", SIGNAL_LOG_PATH)



def update_signal_tracker(history: pd.DataFrame, cfg: Config) -> None:
    """
    Refresh live-tracking fields (Groups 5 & 6) for all Active signals.
    Called once per run after build_analysis().
    """
    if not SIGNAL_LOG_PATH.exists():
        return
    df = _load_signal_log()
    if df.empty:
        return

    # Fix 11 — CSV migration: add new audit columns to existing rows
    for _col, _default in [("days_aligned", 0), ("exit_reason", "")]:
        if _col not in df.columns:
            df[_col] = _default
    df["days_aligned"] = pd.to_numeric(df["days_aligned"], errors="coerce").fillna(0).astype(int)

    active_mask = df["status"] == "Active"
    if not active_mask.any():
        logger.info("Signal tracker: no Active signals to update.")
        return

    today_str    = RESOLVED_TRADE_DATE.strftime("%Y-%m-%d")
    today_ts     = pd.Timestamp(RESOLVED_TRADE_DATE)

    # ── Fix 16B: only increment days_aligned on a new trading day (weekday) ──────
    # Check if we already ran today by looking at last_updated for active rows.
    already_ran_today = (
        df.loc[active_mask, "last_updated"].astype(str).str[:10] == today_str
    ).all() and active_mask.any()
    is_trading_day = RESOLVED_TRADE_DATE.weekday() < 5
    if is_trading_day and not already_ran_today:
        df.loc[active_mask, "days_aligned"] = df.loc[active_mask, "days_aligned"] + 1

    # Build a lookup: symbol -> latest bar from history
    hist_latest = (
        history.sort_values("Date")
               .groupby("Symbol")
               .last()
               .reset_index()
    )
    price_map = hist_latest.set_index("Symbol")["Close"].to_dict()

    updated = 0
    for idx in df[active_mask].index:
        sym          = str(df.at[idx, "symbol"])
        sig_close    = float(df.at[idx, "signal_close"] or 0)
        sig_date     = pd.Timestamp(df.at[idx, "signal_date"])
        sig_type     = str(df.at[idx, "signal_type"])

        cur_close = price_map.get(sym, None)
        if cur_close is None or sig_close == 0:
            continue
        cur_close = float(cur_close)

        # ── Fix 16C: count trading days (Mon-Fri) not calendar days ──────────
        days = int(np.busday_count(sig_date.date(), today_ts.date()))
        days = max(0, days)

        # Return metrics
        pct_ret  = (cur_close - sig_close) / sig_close * 100
        abs_ret  = cur_close - sig_close

        # Running max/min
        prev_max_high = float(df.at[idx, "max_high_since"] or sig_close)
        prev_min_low  = float(df.at[idx, "min_low_since"]  or sig_close)
        new_max_high  = max(prev_max_high, cur_close)
        new_min_low   = min(prev_min_low,  cur_close)
        max_pct_gain  = (new_max_high - sig_close) / sig_close * 100
        max_pct_dd    = (new_min_low  - sig_close) / sig_close * 100  # negative = drawdown

        # EMA alignment check from history
        sym_hist = history[history["Symbol"] == sym].sort_values("Date")
        ema_aligned = True
        ribbon_on   = False
        # Preserve existing RSI/delivery from log — raw history has no computed indicators
        cur_rsi     = float(df.at[idx, "current_rsi"]) if pd.notna(df.at[idx, "current_rsi"]) else float("nan")
        cur_deliv   = float(df.at[idx, "current_delivery_pct"]) if pd.notna(df.at[idx, "current_delivery_pct"]) else float("nan")
        if not sym_hist.empty:
            last = sym_hist.iloc[-1]
            # EMA alignment — only check if indicators were computed (columns present)
            if sig_type == "EMA9x21" and "EMA9" in last.index and "EMA21" in last.index:
                ema_aligned = bool(last["EMA9"] > last["EMA21"])
            elif sig_type == "EMA21x50" and "EMA21" in last.index and "EMA50" in last.index:
                ema_aligned = bool(last["EMA21"] > last["EMA50"])
            # RSI / ribbon — only update if columns present
            if "RSI14" in last.index and pd.notna(last["RSI14"]):
                cur_rsi = float(last["RSI14"])
            if "RibbonActive" in last.index:
                ribbon_on = bool(last["RibbonActive"])

        # Status update
        status = "Active"
        cross_back_date = df.at[idx, "cross_back_date"]
        if not ema_aligned and pd.isna(cross_back_date):
            df.at[idx, "cross_back_date"] = today_str
            status = "Crossed Back"
        elif not ema_aligned:
            status = "Crossed Back"

        stop = float(df.at[idx, "suggested_stop"] or 0)
        tgt1 = float(df.at[idx, "suggested_target_1"] or 0)
        tgt2 = float(df.at[idx, "suggested_target_2"] or 0)
        stop_hit  = bool(df.at[idx, "stop_hit"])  or (stop > 0 and cur_close <= stop)
        tgt1_hit  = bool(df.at[idx, "target1_hit"]) or (tgt1 > 0 and cur_close >= tgt1)
        tgt2_hit  = bool(df.at[idx, "target2_hit"]) or (tgt2 > 0 and cur_close >= tgt2)

        if stop_hit:    status = "Stop Hit"
        if tgt2_hit:    status = "Completed"

        # Archive old closed signals
        cutoff = today_ts - pd.Timedelta(days=cfg.signal_log_archive_days)
        if status != "Active" and sig_date < cutoff:
            status = "Archived"

        # Write back
        df.at[idx, "last_updated"]         = today_str
        df.at[idx, "days_since_signal"]    = days
        df.at[idx, "current_close"]        = round(cur_close, 2)
        df.at[idx, "pct_return"]           = round(pct_ret, 2)
        df.at[idx, "abs_return"]           = round(abs_ret, 2)
        df.at[idx, "max_high_since"]       = round(new_max_high, 2)
        df.at[idx, "max_pct_gain"]         = round(max_pct_gain, 2)
        df.at[idx, "min_low_since"]        = round(new_min_low, 2)
        df.at[idx, "max_pct_drawdown"]     = round(max_pct_dd, 2)
        df.at[idx, "ema_still_aligned"]    = ema_aligned
        df.at[idx, "current_rsi"]          = round(cur_rsi, 2)   if not pd.isna(cur_rsi)   else float("nan")
        df.at[idx, "current_delivery_pct"]  = round(cur_deliv, 2) if not pd.isna(cur_deliv) else float("nan")
        df.at[idx, "ribbon_active"]        = ribbon_on
        df.at[idx, "stop_hit"]             = stop_hit
        df.at[idx, "target1_hit"]          = tgt1_hit
        df.at[idx, "target2_hit"]          = tgt2_hit
        df.at[idx, "status"]               = status
        # Fix 11 — write exit_reason when signal closes
        prev_status = str(df.at[idx, "exit_reason"])
        if status != "Active" and not prev_status.strip():
            if status in ("Crossed Back",):
                df.at[idx, "exit_reason"] = "EMA alignment lost"
            elif status == "Stop Hit":
                df.at[idx, "exit_reason"] = "Price crossed below stop level"
            elif status in ("Completed", "Target Hit"):
                df.at[idx, "exit_reason"] = "Price reached target level"
            elif status == "Archived":
                df.at[idx, "exit_reason"] = "Archived after retention period"
        updated += 1

    df.to_csv(SIGNAL_LOG_PATH, index=False)
    logger.info("Signal tracker: updated %d Active signal rows.", updated)


def build_analysis(
    history: pd.DataFrame,
    universe: pd.DataFrame,
    cfg: Config,
    s: requests.Session,
) -> dict:
    """
    Returns a dict with three keys, each a pd.DataFrame:
      'ema_21_50'  – EMA 21 crossed above EMA 50 today  (original report)
      'ema_9_21'   – EMA 9  crossed above EMA 21 today  (Report 1 new)
      'ema_ribbon' – EMA10 > EMA20 > EMA50 > EMA200 today (Report 2 new)
    """
    merged = history.merge(universe, on="Symbol", how="left")
    merged = merged.sort_values(["Symbol", "Date"]).reset_index(drop=True)

    res_21_50      = []
    res_9_21       = []
    res_ribbon_new = []
    res_ribbon     = []
    res_pb_ema9    = []
    res_pb_ema21   = []

    # ── Delivery data: backfill cache if needed, then fetch latest ─────────
    backfill_delivery_history(s, cfg)
    delivery_dict, deliv_fallback_used = fetch_delivery_data(s, RESOLVED_TRADE_DATE)
    logger.info(
        "Delivery data: %d symbols | fallback=%s",
        len(delivery_dict), deliv_fallback_used,
    )

    # Corporate actions (non-critical; fails gracefully to {})
    corporate_actions_dict: dict = {}
    try:
        corporate_actions_dict = fetch_corporate_actions(
            s,
            list(universe["Symbol"].str.upper()),
            lookahead_days=cfg.corporate_action_lookahead_days,
        )
    except Exception as _ca_err:
        logger.warning("Corporate actions fetch skipped: %s", _ca_err)

    # Index membership (cached; non-critical)
    index_map: dict = {}
    try:
        index_map = fetch_index_membership(s)
    except Exception as _im_err:
        logger.warning("Index membership fetch skipped: %s", _im_err)

    # Fix 8 — Build crossover lookup dict: {symbol → most recent crossover date}
    # Used by scan_pullback_to_ema() to reject stale pullback signals.
    _cross_lookup: dict = {}
    if SIGNAL_LOG_PATH.exists():
        try:
            _slog = pd.read_csv(SIGNAL_LOG_PATH, parse_dates=["signal_date"],
                                usecols=["symbol", "signal_type", "signal_date"])
            _cross_types = {"EMA9x21", "EMA21x50"}
            _cross_rows  = _slog[_slog["signal_type"].isin(_cross_types)]
            if not _cross_rows.empty:
                _cross_lookup = (
                    _cross_rows.sort_values("signal_date")
                               .groupby("symbol")["signal_date"]
                               .max()
                               .to_dict()
                )
        except Exception as _e:
            logger.warning("Fix 8: could not build crossover lookup — %s", _e)

    for symbol, g in merged.groupby("Symbol", sort=True):
        g = g.sort_values("Date").copy()
        # Fix 14 — Minimum bars guard.
        # 220 = 200 bars for EMA-200 + 20 warm-up buffer. Stocks with fewer
        # bars produce unreliable EMA-200, SuperTrend, and S/R values; they
        # are skipped and logged so the operator knows to extend the backfill.
        MIN_BARS_REQUIRED = 220
        if len(g) < MIN_BARS_REQUIRED:
            logger.warning(
                "Skipping %s — only %d bars available (need %d). "
                "Extend history_backfill_days in config.json to include this symbol.",
                symbol, len(g), MIN_BARS_REQUIRED,
            )
            continue

        # ── Compute all EMAs needed across all three reports ──────────────────
        # Fix 13 — Discard the first `span` warm-up bars of each EMA.
        # ewm() seeds from bar 0 using an SMA estimate; those early values are
        # mathematically unreliable until the window has fully converged.
        # We NaN them out so crossover/ribbon logic never fires on warm-up artefacts.
        def _ema_warmed(series: pd.Series, span: int) -> pd.Series:
            result = ema(series, span)
            if len(result) > span:
                result.iloc[:span] = np.nan
            return result

        g["EMA9"]   = _ema_warmed(g["Close"],   9)
        g["EMA10"]  = _ema_warmed(g["Close"],  10)
        g["EMA20"]  = _ema_warmed(g["Close"],  20)
        g["EMA21"]  = _ema_warmed(g["Close"],  21)
        g["EMA50"]  = _ema_warmed(g["Close"],  50)
        g["EMA200"] = _ema_warmed(g["Close"], 200)

        g["RSI14"] = rsi(g["Close"], 14)
        g["MACD"], g["MACDSignal"], g["MACDHist"] = macd(g["Close"])
        g["VolAvg14"] = g["Volume"].rolling(14).mean()

        g["SupertrendTrend_10_3"], g["SupertrendValue_10_3"] = supertrend(
            g[["High", "Low", "Close"]].copy(), 10, 3.0
        )
        g["SupertrendTrend_10_1"], g["SupertrendValue_10_1"] = supertrend(
            g[["High", "Low", "Close"]].copy(), 10, 1.0
        )
        g["SupertrendGreen_10_3"] = g["SupertrendTrend_10_3"] == 1
        g["SupertrendGreen_10_1"] = g["SupertrendTrend_10_1"] == 1

        latest = g.iloc[-1]

        # ── Report 1 (original): EMA 21 × EMA 50 bullish crossover ───────────
        cross_21_50 = (
            (g["EMA21"] > g["EMA50"]) &
            (g["EMA21"].shift(1) <= g["EMA50"].shift(1))
        )
        if bool(cross_21_50.iloc[-1]):
            res_21_50.append(_build_stock_row(symbol, g, latest, history, cfg, s,
                                  delivery_dict=delivery_dict,
                                  deliv_fallback_used=deliv_fallback_used,
                                  corp_actions=corporate_actions_dict.get(symbol.upper(), None),
                                  index_map=index_map))

        # ── Report 2: EMA 9 × EMA 21 bullish crossover ───────────────────────
        cross_9_21 = (
            (g["EMA9"] > g["EMA21"]) &
            (g["EMA9"].shift(1) <= g["EMA21"].shift(1))
        )
        if bool(cross_9_21.iloc[-1]):
            res_9_21.append(_build_stock_row(symbol, g, latest, history, cfg, s,
                                  delivery_dict=delivery_dict,
                                  deliv_fallback_used=deliv_fallback_used,
                                  corp_actions=corporate_actions_dict.get(symbol.upper(), None),
                                  index_map=index_map))

        # ── Report 3: EMA Ribbon NEW ENTRY – ribbon turned true on latest bar only ──
        g["RibbonActive"] = (
            (g["EMA10"] > g["EMA20"]) &
            (g["EMA20"] > g["EMA50"]) &
            (g["EMA50"] > g["EMA200"])
        )
        ribbon_new_entry = bool(g["RibbonActive"].iloc[-1]) and not bool(g["RibbonActive"].iloc[-2])
        if ribbon_new_entry:
            res_ribbon_new.append(_build_stock_row(symbol, g, latest, history, cfg, s,
                                  delivery_dict=delivery_dict,
                                  deliv_fallback_used=deliv_fallback_used,
                                  corp_actions=corporate_actions_dict.get(symbol.upper(), None),
                                  index_map=index_map))

        # ── Report 4: EMA Ribbon ACTIVE – EMA10 > EMA20 > EMA50 > EMA200 ────
        if bool(g["RibbonActive"].iloc[-1]):
            res_ribbon.append(_build_stock_row(symbol, g, latest, history, cfg, s,
                                  delivery_dict=delivery_dict,
                                  deliv_fallback_used=deliv_fallback_used,
                                  corp_actions=corporate_actions_dict.get(symbol.upper(), None),
                                  index_map=index_map))


        # ── Report 5: Pullback to (or nearing) EMA 9 ─────────────────────────
        pb9_ok, pb9_touch, pb9_level, pb9_days = scan_pullback_to_ema(
            g, "EMA9", pullback_window=15,
            last_cross_date=_cross_lookup.get(symbol),  # Fix 8
        )
        if pb9_ok:
            row = _build_stock_row(symbol, g, latest, history, cfg, s,
                                  delivery_dict=delivery_dict,
                                  deliv_fallback_used=deliv_fallback_used,
                                  corp_actions=corporate_actions_dict.get(symbol.upper(), None),
                                  index_map=index_map)
            row["PullbackDays"]  = pb9_days
            row["PullbackToEMA"] = f"EMA9 @ {pb9_level:.2f}"
            row["TouchPrice"]    = pb9_touch
            res_pb_ema9.append(row)

        # ── Report 6: Pullback to (or nearing) EMA 21 ────────────────────────
        pb21_ok, pb21_touch, pb21_level, pb21_days = scan_pullback_to_ema(
            g, "EMA21", pullback_window=30,
            last_cross_date=_cross_lookup.get(symbol),  # Fix 8
        )
        if pb21_ok:
            row = _build_stock_row(symbol, g, latest, history, cfg, s,
                                  delivery_dict=delivery_dict,
                                  deliv_fallback_used=deliv_fallback_used,
                                  corp_actions=corporate_actions_dict.get(symbol.upper(), None),
                                  index_map=index_map)
            row["PullbackDays"]  = pb21_days
            row["PullbackToEMA"] = f"EMA21 @ {pb21_level:.2f}"
            row["TouchPrice"]    = pb21_touch
            res_pb_ema21.append(row)

        def _to_df(results: list, empty_msg: str) -> pd.DataFrame:
            if not results:
                return pd.DataFrame([{"Message": empty_msg}])
            df = pd.DataFrame(results)
            if "Conviction Score" in df.columns:
                df = df.sort_values("Conviction Score", ascending=False).reset_index(drop=True)
            return df

    return {
        "ema_21_50": _to_df(
            res_21_50,
            "No stock met the rule: EMA 21 crossed above EMA 50 on the latest trading day.",
        ),
        "ema_9_21": _to_df(
            res_9_21,
            "No stock met the rule: EMA 9 crossed above EMA 21 on the latest trading day.",
        ),
        "ema_ribbon_new": _to_df(
            res_ribbon_new,
            "No stock met the rule: EMA 10 > EMA 20 > EMA 50 > EMA 200 (new entry on latest trading day).",
        ),
        "ema_ribbon": _to_df(
            res_ribbon,
            "No stock met the rule: EMA 10 > EMA 20 > EMA 50 > EMA 200 on the latest trading day.",
        ),
        "pb_ema9": _to_df(
            res_pb_ema9,
            "No stock met the pullback-to-EMA-9 conditions on the latest trading day.",
        ),
        "pb_ema21": _to_df(
            res_pb_ema21,
            "No stock met the pullback-to-EMA-21 conditions on the latest trading day.",
        ),
    }


# ── Dow Theory cell splitter helpers ───────────────────────────────────────
def _dow_trend(label: str) -> str:
    """Extract 'Primary / Secondary / Minor' part from DowTheory label."""
    parts = [p.strip() for p in label.split('|') if p.strip()]
    trend_parts = [p for p in parts if p.startswith(('Primary', 'Secondary', 'Minor'))]
    return ' | '.join(trend_parts) if trend_parts else label

def _dow_phase(label: str) -> str:
    """Extract 'Phase' part from DowTheory label."""
    for p in label.split('|'):
        p = p.strip()
        if p.startswith('Phase:'):
            return p.replace('Phase:', '').strip()
    return ''


_section_counter = 0  # gives each table a unique id for the JS sorter

def _render_report_section(
    report_df: pd.DataFrame,
    heading: str,
    empty_rule: str,
) -> str:
    """Return an HTML string for one report section (heading + sortable table or empty notice)."""
    global _section_counter
    _section_counter += 1
    tbl_id = f"tbl{_section_counter}"
    sec_id = f"section-{_section_counter}"

    section = (
        f"<div id='{sec_id}' class='section-block'>"
        f"<div class='section-head'>"
        f"<h2>{heading}</h2>"
        f"<span class='section-badge'>— signals</span>"
        f"<button class='section-toggle'>&#9650; Collapse</button>"
        f"</div>"
        f"<div class='section-body'>\n"
    )

    if "Message" in report_df.columns:
        section += f"<div class='empty'><p>{html.escape(empty_rule)}</p></div>\n</div></div>\n"
        return section

    def _index_badge(val: str) -> str:
        if val == "NIFTY 50":
            return ('<span style="background:#1565c0;color:#fff;padding:2px 7px;'
                    'border-radius:3px;font-size:10px;font-weight:bold;'
                    'white-space:nowrap;">N50</span>')
        if val == "Nifty Next 50":
            return ('<span style="background:#e65100;color:#fff;padding:2px 7px;'
                    'border-radius:3px;font-size:10px;font-weight:bold;'
                    'white-space:nowrap;">NN50</span>')
        return '<span style="color:#c8dcf4;font-size:10px;">N100</span>'

    tr = []
    for _, r in report_df.iterrows():
        news_items = r.get("RecentNewsHtml", [])
        if not isinstance(news_items, list):
            news_items = [news_items] if news_items else []
        # ── RSI colour ──────────────────────────────────────────────────────
        try:
            _rsi_val = float(str(r["RSI14"]).replace("%","").strip())
            _rsi_col = ("#4ecb82" if _rsi_val >= 60 else
                        "#f5a623" if _rsi_val >= 40 else "#f06464")
            _rsi_td  = (f"<td style='text-align:right;font-weight:600;white-space:nowrap;"
                        f"color:{_rsi_col}'>{r['RSI14']}</td>")
        except Exception:
            _rsi_td  = f"<td style='white-space:nowrap'>{r['RSI14']}</td>"
        # ── 52WH/52WL colour ────────────────────────────────────────────────
        _whl = str(r.get("52WH/52WL",""))
        _whl_col = ("#4ecb82" if "52WH" in _whl else
                    "#f06464" if "52WL" in _whl else "inherit")
        _whl_td  = (f"<td style='font-size:.86rem;white-space:nowrap;color:{_whl_col}'>{_whl}</td>")
        # ── Last Close: right-align monospace ───────────────────────────────
        _lc_td   = (f"<td style='text-align:right;"
                    f"font-size:.94rem;white-space:nowrap'>{r['Last Close']}</td>")
        # ── Conviction score ─────────────────────────────────────────────────
        _cv      = r.get("Conviction Score","")
        _cv_val  = float(_cv) if str(_cv).replace('.','',1).isdigit() else 0
        _cv_col  = '#4ecb82' if _cv_val >= 70 else ('#e8c300' if _cv_val >= 40 else '#f06464')
        _cv_td   = f"<td style='text-align:center;font-weight:700;color:{_cv_col}'>{_cv}</td>"
        _sym_id = r.get('Symbol', '').replace(' ', '_')
        tr.append(
            f"<tr id='row-{_sym_id}'>"
            f"<td style='font-weight:700;white-space:nowrap;color:#dce8f8;font-weight:700'>{r['Symbol']}</td>"
            "<td style='text-align:center;padding:7px 8px'>" + _index_badge(r.get("Index","Nifty 100")) + "</td>"
            + f"<td style='white-space:normal;word-break:normal;overflow-wrap:break-word'>{r['Company']}</td>"
            + f"<td style='white-space:nowrap;color:#c8dcf4'>{r['Date']}</td>"

            + _lc_td
            + _whl_td
            + f"<td style='font-size:.87rem'>{r['Candlestick Pattern']}</td>"
            + f"<td style='text-align:center'>{r['VolumeAbove14Avg']}</td>"
            + _rsi_td
            + f"<td style='font-size:.87rem'>{r['RSI Divergence']}</td>"
            + f"<td style='text-align:center'>{r['MACDAboveSignal']}</td>"
            + f"<td style='white-space:pre-line;font-size:.87rem'>{r.get('SuperTrend(10,3) (10,1) Green','')}</td>"
            + f"<td style='font-size:.87rem'>{r['Momentum']}</td>"
            + f"<td style='font-size:.87rem'>{_dow_trend(r['DowTheory'])}</td>"
            + f"<td style='font-size:.87rem'>{_dow_phase(r['DowTheory'])}</td>"
            + f"<td style='font-size:.87rem'>{r['Chart Pattern']}</td>"
            + f"<td style='white-space:pre-line;font-size:.87rem'>{r.get('Support and Resistance','')}</td>"
            + _cv_td
            + f"<td style='font-size:.85rem;line-height:1.55;color:#dce8f8'>{r.get('Conviction Explanation','')}</td>"
            + f"<td style='font-size:.84rem'><ul>{''.join(f'<li>{item}</li>' for item in news_items)}</ul></td>"
            + f"</tr>"
        )

    # Column index constants (0-based, matching td order above)
    # Conviction Score is col 18; used as the default sort arrow indicator
    CONVICTION_COL = 17

    def _th(label: str, col: int, numeric: bool = False) -> str:
        cls = "th-sort th-num" if numeric else "th-sort"
        # Mark the default sort column (Conviction Score desc) with the down arrow
        arrow = " &#9660;" if col == CONVICTION_COL else " <span class='sort-arrow'></span>"
        data = f"data-col='{col}' data-numeric='{'1' if numeric else '0'}' data-dir='{'desc' if col == CONVICTION_COL else 'none'}'"
        return f"<th class='{cls}' {data}>{label}{arrow}</th>"

    header_row = (
        _th("Symbol", 0)
        + _th("Index", 1)
        + _th("Company", 2)
        + _th("Date", 3)
        + _th("Last Close", 4, numeric=True)
        + _th("52W H/L", 5)
        + _th("Candle", 6)
        + _th("Vol>14D", 7)
        + _th("RSI", 8, numeric=True)
        + _th("RSI Div", 9)
        + _th("MACD > Sig", 10)
        + _th("ST(10,3)/(10,1)", 11)
        + _th("Momentum", 12)
        + _th("DT Trend", 13)
        + _th("DT Phase", 14)
        + _th("Chart Pat.", 15)
        + _th("Sup/Res", 16)
        + _th("Conv Score", 17, numeric=True)
        + _th("Conv Explanation", 18)
        + _th("News", 19)
    )

    section += (
        f"<table id='{tbl_id}'><thead><tr>"
        + header_row
        + "</tr></thead><tbody>"
        + "".join(tr)
        + "</tbody></table>\n"
        + f"<div style='text-align:right;padding:4px 8px 8px 0;font-size:0.72rem;'>"
          f"<a href='#{sec_id}' style='color:var(--accent);text-decoration:none;"
          f"opacity:0.70;' onmouseover=\"this.style.opacity='1'\""
          f" onmouseout=\"this.style.opacity='0.70'\">"
          f"&#8593;&nbsp;Back to top of this section</a></div>\n"
        + "</div></div>\n"
    )
    return section



def _render_ribbon_tiered(report_df: pd.DataFrame) -> str:
    """
    Render the EMA Ribbon Cumulative section as three collapsible tiers:
      Strong   (Conviction >= 70) — expanded by default
      Moderate (Conviction 50-69) — collapsed by default
      Weak     (Conviction  < 50) — collapsed by default
    Reuses the same row/header HTML as _render_report_section so all
    columns, colours and sort behaviour are identical.
    """
    global _section_counter
    _section_counter += 1
    outer_id = f"rbn{_section_counter}"

    TIER_DEFS = [
        # (key, emoji+label, header-bg, lo_inc, hi_exc, expanded)
        ("strong",   "🟢 Strong",   "#1a472a", 70,  101, True),
        ("moderate", "🟡 Moderate", "#5a4000", 50,   70, False),
        ("weak",     "🔴 Weak",     "#5a1a1a",  0,   50, False),
    ]

    # ── helpers (mirrors _render_report_section exactly) ────────────────────
    def _index_badge(val):
        if val == "NIFTY 50":
            return ('<span style="background:#1565c0;color:#fff;padding:2px 7px;'
                    'border-radius:3px;font-size:10px;font-weight:bold;'
                    'white-space:nowrap;">N50</span>')
        if val == "Nifty Next 50":
            return ('<span style="background:#e65100;color:#fff;padding:2px 7px;'
                    'border-radius:3px;font-size:10px;font-weight:bold;'
                    'white-space:nowrap;">NN50</span>')
        return '<span style="color:#c8dcf4;font-size:10px;">N100</span>'

    def _build_tr(r):
        news_items = r.get("RecentNewsHtml", [])
        if not isinstance(news_items, list):
            news_items = [news_items] if news_items else []
        try:
            _rsi_val = float(str(r["RSI14"]).replace("%", "").strip())
            _rsi_col = ("#4ecb82" if _rsi_val >= 60 else
                        "#f5a623" if _rsi_val >= 40 else "#f06464")
            _rsi_td  = (f"<td style='text-align:right;font-weight:600;white-space:nowrap;"
                        f"color:{_rsi_col}'>{r['RSI14']}</td>")
        except Exception:
            _rsi_td  = f"<td style='white-space:nowrap'>{r['RSI14']}</td>"
        _whl     = str(r.get("52WH/52WL", ""))
        _whl_col = ("#4ecb82" if "52WH" in _whl else
                    "#f06464" if "52WL" in _whl else "inherit")
        _whl_td  = f"<td style='font-size:.86rem;white-space:nowrap;color:{_whl_col}'>{_whl}</td>"
        _lc_td   = (f"<td style='text-align:right;font-size:.94rem;white-space:nowrap'>"
                    f"{r['Last Close']}</td>")
        _cv      = r.get("Conviction Score", "")
        _cv_val  = float(_cv) if str(_cv).replace(".", "", 1).isdigit() else 0
        _cv_col  = "#4ecb82" if _cv_val >= 70 else ("#e8c300" if _cv_val >= 40 else "#f06464")
        _cv_td   = f"<td style='text-align:center;font-weight:700;color:{_cv_col}'>{_cv}</td>"
        _sym_id  = r.get("Symbol", "").replace(" ", "_")
        return (
            f"<tr id='row-{_sym_id}'>"
            f"<td style='font-weight:700;white-space:nowrap;color:#dce8f8'>{r['Symbol']}</td>"
            "<td style='text-align:center;padding:7px 8px'>" + _index_badge(r.get("Index", "Nifty 100")) + "</td>"
            + f"<td style='white-space:normal;word-break:normal;overflow-wrap:break-word'>{r['Company']}</td>"
            + f"<td style='white-space:nowrap;color:#c8dcf4'>{r['Date']}</td>"
            + _lc_td + _whl_td
            + f"<td style='font-size:.87rem'>{r['Candlestick Pattern']}</td>"
            + f"<td style='text-align:center'>{r['VolumeAbove14Avg']}</td>"
            + _rsi_td
            + f"<td style='font-size:.87rem'>{r['RSI Divergence']}</td>"
            + f"<td style='text-align:center'>{r['MACDAboveSignal']}</td>"
            + f"<td style='white-space:pre-line;font-size:.87rem'>{r.get('SuperTrend(10,3) (10,1) Green', '')}</td>"
            + f"<td style='font-size:.87rem'>{r['Momentum']}</td>"
            + f"<td style='font-size:.87rem'>{_dow_trend(r['DowTheory'])}</td>"
            + f"<td style='font-size:.87rem'>{_dow_phase(r['DowTheory'])}</td>"
            + f"<td style='font-size:.87rem'>{r['Chart Pattern']}</td>"
            + f"<td style='white-space:pre-line;font-size:.87rem'>{r.get('Support and Resistance', '')}</td>"
            + _cv_td
            + f"<td style='font-size:.85rem;line-height:1.55;color:#dce8f8'>{r.get('Conviction Explanation', '')}</td>"
            + f"<td style='font-size:.84rem'><ul>{''.join(f'<li>{item}</li>' for item in news_items)}</ul></td>"
            + "</tr>"
        )

    def _th(label, col, numeric=False):
        cls   = "th-sort th-num" if numeric else "th-sort"
        arrow = " &#9660;" if col == 17 else " <span class='sort-arrow'></span>"
        ddir  = "desc" if col == 17 else "none"
        data  = f"data-col='{col}' data-numeric='{'1' if numeric else '0'}' data-dir='{ddir}'"
        return f"<th class='{cls}' {data}>{label}{arrow}</th>"

    header_row = (
        _th("Symbol", 0) + _th("Index", 1) + _th("Company", 2) + _th("Date", 3)
        + _th("Last Close", 4, True) + _th("52W H/L", 5) + _th("Candle", 6)
        + _th("Vol>14D", 7) + _th("RSI", 8, True) + _th("RSI Div", 9)
        + _th("MACD > Sig", 10) + _th("ST(10,3)/(10,1)", 11)
        + _th("Momentum", 12) + _th("DT Trend", 13) + _th("DT Phase", 14)
        + _th("Chart Pat.", 15) + _th("Sup/Res", 16)
        + _th("Conv Score", 17, True) + _th("Conv Explanation", 18) + _th("News", 19)
    )

    # ── empty state ──────────────────────────────────────────────────────────
    if "Message" in report_df.columns:
        return (
            "<div class='section-block'>"
            "<div class='section-head'>"
            "<h2>EMA Ribbon &#8211; EMA 10 &gt; EMA 20 &gt; EMA 50 &gt; EMA 200 "
            "Bullish Report &#8211; Cumulative (Tiered)</h2>"
            "<span class='section-badge'>0 signals</span>"
            "<button class='section-toggle'>&#9650; Collapse</button>"
            "</div>"
            "<div class='section-body'>"
            "<div class='empty'><p>No stock currently has EMA 10 &gt; EMA 20 &gt; "
            "EMA 50 &gt; EMA 200 alignment.</p></div>"
            "</div></div>\n"
        )

    total = len(report_df)

    # ── outer section wrapper ────────────────────────────────────────────────
    out = (
        "<div class='section-block'>"
        "<div class='section-head'>"
        "<h2>EMA Ribbon &#8211; EMA 10 &gt; EMA 20 &gt; EMA 50 &gt; EMA 200 "
        "Bullish Report &#8211; Cumulative (Tiered)</h2>"
        f"<span class='section-badge'>{total} signal{'s' if total != 1 else ''}</span>"
        "<button class='section-toggle'>&#9650; Collapse</button>"
        "</div>"
        "<div class='section-body'>"
    )

    # ── summary bar (always visible, shows all 3 tier counts at a glance) ───
    def _tier_count(lo, hi):
        return report_df["Conviction Score"].apply(
            lambda x: lo <= (float(x) if str(x).replace(".", "", 1).isdigit() else 0) < hi
        ).sum()

    s_cnt = _tier_count(70, 101)
    m_cnt = _tier_count(50,  70)
    w_cnt = _tier_count( 0,  50)
    out += (
        "<div style='display:flex;flex-wrap:wrap;gap:12px;align-items:center;"
        "margin-bottom:16px;padding:12px 16px;background:#0a1628;"
        "border:1px solid #1a3560;border-radius:8px;'>"
        f"<span style='font-size:.87rem;color:#8aafd4'>Ribbon-aligned stocks today: "
        f"<b style='color:#dce8f8;font-size:1rem'>{total}</b></span>"
        "<span style='color:#1a3560;font-size:1.2rem'>|</span>"
        f"<span style='font-size:.87rem'>🟢 Strong <b style='color:#4ecb82'>{s_cnt}</b></span>"
        f"<span style='font-size:.87rem'>🟡 Moderate <b style='color:#e8c300'>{m_cnt}</b></span>"
        f"<span style='font-size:.87rem'>🔴 Weak <b style='color:#f06464'>{w_cnt}</b></span>"
        "<span style='margin-left:auto;font-size:.78rem;color:#6a8cb8;font-style:italic'>"
        "Strong expanded by default · click tier header to toggle</span>"
        "</div>"
    )

    # ── render each tier ─────────────────────────────────────────────────────
    for tier_key, tier_label, hdr_bg, lo, hi, expanded in TIER_DEFS:
        mask = report_df["Conviction Score"].apply(
            lambda x: lo <= (float(x) if str(x).replace(".", "", 1).isdigit() else 0) < hi
        )
        tier_df  = report_df[mask].copy()
        cnt      = len(tier_df)
        body_id  = f"tier-{tier_key}-{outer_id}"

        _section_counter += 1
        tbl_id = f"tbl{_section_counter}"

        toggle_txt = "▲ Collapse" if expanded else f"▼ Expand ({cnt})"
        disp_style = "" if expanded else "display:none"

        out += (
            f"<div style='margin-bottom:14px;border:1px solid #1a3560;border-radius:8px;overflow:hidden'>"
            # tier header — full-width clickable strip
            f"<div style='background:{hdr_bg};padding:9px 16px;display:flex;"
            f"align-items:center;gap:10px;cursor:pointer;user-select:none'"
            f" onclick=\"var b=document.getElementById('{body_id}');"
            f"var btn=document.getElementById('btn-{body_id}');"
            f"if(b.style.display==='none'){{b.style.display='';btn.textContent='▲ Collapse';}}"
            f"else{{b.style.display='none';btn.textContent='▼ Expand ({cnt})';}}\">"
            f"<span style='font-size:.95rem;font-weight:700;color:#fff'>{tier_label}</span>"
            f"<span style='background:rgba(255,255,255,.2);color:#fff;font-size:.78rem;"
            f"padding:2px 9px;border-radius:12px;font-weight:600'>"
            f"{cnt} stock{'s' if cnt != 1 else ''}</span>"
            f"<button id='btn-{body_id}' style='margin-left:auto;background:rgba(255,255,255,.15);"
            f"border:1px solid rgba(255,255,255,.35);color:#fff;border-radius:5px;"
            f"padding:2px 10px;font-size:.78rem;cursor:pointer;pointer-events:none'>"
            f"{toggle_txt}</button>"
            f"</div>"
            # tier body
            f"<div id='{body_id}' style='{disp_style}'>"
        )

        if cnt == 0:
            out += (
                "<div style='padding:12px 16px;color:#6a8cb4;font-size:.87rem;background:#0d1c30'>"
                "No stocks in this tier today.</div>"
            )
        else:
            rows_html = "".join(_build_tr(r) for _, r in tier_df.iterrows())
            out += (
                f"<table id='{tbl_id}'><thead><tr>"
                + header_row
                + "</tr></thead><tbody>"
                + rows_html
                + "</tbody></table>"
            )

        out += "</div></div>"   # close tier-body div + tier wrapper div

    out += "</div></div>\n"     # close section-body + section-block
    return out


def render_html(reports: dict, history: pd.DataFrame, market_health: dict = None) -> Path:
    fname = REPORTS_DIR / f"nifty100_daily_report_{TODAY_IST.strftime('%Y%m%d')}.html"
    last_date = history["Date"].max().strftime("%Y-%m-%d") if not history.empty else "NA"

    _mh_html = render_market_health_html(market_health) if market_health else ""
    _hc_html = render_high_conviction_cards(reports, threshold=75.0)

    section_21_50 = _render_report_section(
        reports["ema_21_50"],
        "Nifty 100 Daily EMA 21/50 Bullish Cross Report",
        "No stock met the rule: EMA 21 crossed above EMA 50 on the latest trading day.",
    )
    section_9_21 = _render_report_section(
        reports["ema_9_21"],
        "Nifty 100 Daily EMA 9/21 Bullish Cross Report",
        "No stock met the rule: EMA 9 crossed above EMA 21 on the latest trading day.",
    )
    section_ribbon_new = _render_report_section(
        reports["ema_ribbon_new"],
        "EMA Ribbon &#8211; EMA 10 &gt; EMA 20 &gt; EMA 50 &gt; EMA 200 Bullish Report &#8211; Today (New Entry)",
        "No stock met the rule: EMA 10 > EMA 20 > EMA 50 > EMA 200 (new entry on latest trading day).",
    )
    section_ribbon = _render_ribbon_tiered(reports["ema_ribbon"])
    section_pb_ema9 = _render_report_section(
        reports["pb_ema9"],
        "Pullback to (or Nearing) EMA 9 Report",
        "No stock met the pullback-to-EMA-9 conditions on the latest trading day.",
    )
    section_pb_ema21 = _render_report_section(
        reports["pb_ema21"],
        "Pullback to (or Nearing) EMA 21 Report",
        "No stock met the pullback-to-EMA-21 conditions on the latest trading day.",
    )

    page = f"""<!doctype html>
<html lang='en'>
<head>
<meta charset='utf-8'>
<meta name='viewport' content='width=device-width, initial-scale=1'>
<title>Nifty 100 Daily Report &middot; {TODAY_IST.strftime('%d %b %Y')}</title>
<link rel='preconnect' href='https://fonts.googleapis.com'>
<link href='https://fonts.googleapis.com/css2?family=Inter:ital,wght@0,400;0,500;0,600;0,700;1,400&display=swap' rel='stylesheet'>
<style>
/* ── Reset ── */
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}

/* ── Design tokens ── */
:root{{
  --bg:          #0b1623;
  --bg-card:     #0f1f35;
  --bg-table:    #0d1c30;
  --bg-thead:    #112240;
  --bg-even:     #0f1f35;
  --bg-hover:    #1f4070;
  --border:      #1e3a5f;
  --border-soft: #172e50;
  --text:        #e2ecff;
  --text-dim:    #c8dcf4;
  --text-muted:  #4a6a8a;
  --accent:      #f5c842;
  --accent2:     #5aabff;
  --green:       #4ecb82;
  --red:         #f06464;
  --orange:      #f5a623;
  --shadow-sm:   0 1px 4px rgba(0,0,0,.4);
  --shadow-md:   0 4px 16px rgba(0,0,0,.55);
  --radius:      10px;
}}

/* ── Base ── */
html{{scroll-behavior:smooth;font-size:16px}}
body{{
  font-family:'Inter',system-ui,sans-serif;
  background:var(--bg);color:var(--text);
  line-height:1.6;min-height:100vh;
}}

/* ── Sticky top header ── */
.site-header{{
  background:linear-gradient(135deg,#0a1e3d 0%,#0d2454 100%);
  border-bottom:1px solid var(--border);
  padding:14px 12px 10px;
  position:sticky;top:0;z-index:200;
  box-shadow:var(--shadow-md);
}}
.site-header h1{{
  font-size:1.1rem;font-weight:700;
  color:#fff;letter-spacing:.4px;margin-bottom:4px;
}}
.meta-bar{{font-size:.83rem;color:var(--text-dim);line-height:1.5}}
.meta-bar strong{{color:var(--accent);font-weight:600}}
.hint-pill{{
  display:inline-block;margin-top:4px;
  background:#112240;border:1px solid var(--border-soft);
  border-radius:20px;padding:2px 10px;
  font-size:.78rem;color:var(--text-muted);
}}

/* ── Page wrapper ── */
.page-wrap{{width:100%;max-width:100%;margin:0;padding:20px 12px 80px;box-sizing:border-box}}

/* ── Section blocks ── */
.section-block{{margin-bottom:6px}}
.section-head{{
  display:flex;align-items:center;gap:10px;
  margin:32px 0 8px;padding-bottom:8px;
  border-bottom:2px solid var(--border);
}}
.section-head h2{{
  font-size:.95rem;font-weight:600;
  color:var(--accent);letter-spacing:.3px;
}}
.section-badge{{
  background:#132444;color:var(--accent2);
  font-size:.78rem;font-weight:700;
  padding:2px 9px;border-radius:20px;
  border:1px solid #1a3560;
}}
.section-toggle{{
  margin-left:auto;background:none;
  border:1px solid var(--border);color:var(--text-dim);
  border-radius:6px;padding:3px 10px;font-size:.80rem;
  cursor:pointer;transition:background .15s,color .15s;
}}
.section-toggle:hover{{background:var(--bg-hover);color:var(--text)}}

/* ── Empty state ── */
.empty{{
  background:var(--bg-card);border:1px solid var(--border-soft);
  border-radius:var(--radius);padding:14px 20px;margin-bottom:12px;
  color:var(--text-dim);font-size:.94rem;
}}

/* ── Table scroll wrapper ── */
.tbl-wrap{{
  overflow:visible;border-radius:var(--radius);
  border:1px solid var(--border);
  box-shadow:var(--shadow-md);margin-bottom:8px;
}}
.tbl-wrap::-webkit-scrollbar{{height:5px}}
.tbl-wrap::-webkit-scrollbar-track{{background:var(--bg-card)}}
.tbl-wrap::-webkit-scrollbar-thumb{{background:#2a4a80;border-radius:3px}}

/* ── Table ── */
table{{
  width:100%;border-collapse:collapse;
  background:var(--bg-table);font-size:.90rem;
}}
thead{{position:sticky;top:55px;z-index:10}}
th{{
  background:var(--bg-thead);
  color:#dce8f8;font-weight:700;font-size:.84rem;
  letter-spacing:.4px;text-transform:uppercase;
  white-space:nowrap;padding:9px 12px;
  border-bottom:2px solid var(--border);
  border-right:1px solid var(--border-soft);
}}
td{{
  padding:8px 12px;vertical-align:top;
  border-bottom:1px solid var(--border);
  border-right:1px solid #1e3a5f;
  color:var(--text);line-height:1.5;
}}
tbody tr:nth-child(even) td{{background:var(--bg-even)}}
tbody tr:hover td{{
  background:var(--bg-hover);
  transition:background .08s;
}}

/* ── Sort headers ── */
th.th-sort{{
  cursor:pointer;user-select:none;
  padding-right:22px;position:relative;
}}
th.th-sort::after{{
  content:"⇅";position:absolute;right:6px;
  top:50%;transform:translateY(-50%);
  opacity:.25;font-size:.65rem;
}}
th.th-sort:hover{{background:#163060;color:#c8dcf4}}
th.th-sort.sorted-asc::after{{content:" ▲";opacity:1;color:var(--accent)}}
th.th-sort.sorted-desc::after{{content:" ▼";opacity:1;color:var(--accent)}}

/* ── Links ── */
a{{color:var(--accent2);text-decoration:none}}
a:hover{{color:#8ccfff;text-decoration:underline}}
small{{color:var(--text-dim);font-size:.84rem}}

/* ── Back to top ── */
#back-top{{
  position:fixed;bottom:24px;right:24px;z-index:300;
  background:#112240;border:1px solid var(--border);
  color:var(--text-dim);border-radius:50%;
  width:36px;height:36px;
  display:flex;align-items:center;justify-content:center;
  font-size:1rem;text-decoration:none;
  box-shadow:var(--shadow-md);opacity:.75;transition:all .2s;
}}
#back-top:hover{{opacity:1;background:var(--bg-hover);color:var(--text)}}

/* ── Disclaimer ── */
.disclaimer{{
  margin-top:40px;padding:12px 18px;
  border-top:1px solid var(--border-soft);
  color:var(--text-muted);font-size:.80rem;line-height:1.6;
}}

/* ── News list ── */
td ul{{padding-left:16px;margin:0}}
td ul li{{margin-bottom:3px;font-size:.86rem;color:var(--text-dim)}}

/* ── Conviction score highlight ── */
td.conv-high{{color:var(--green);font-weight:700}}
td.conv-mid {{color:#e8c300;font-weight:600}}
td.conv-low {{color:var(--red);font-weight:500}}



/* ── Report tables: auto layout — browser allocates by content width ── */
.section-body table {{
  table-layout: auto;
  width: 100%;
  word-break: normal;
}}
.section-body th {{
  white-space: normal;
  word-break: normal;
  overflow-wrap: break-word;
}}
.section-body td {{
  white-space: normal;
  word-break: normal;
  overflow-wrap: break-word;
}}
/* Nudge: give Conviction Explanation a generous min-width */
.section-body th:nth-child(19),
.section-body td:nth-child(19) {{
  min-width: 220px;
}}
/* Nudge: keep narrow cols from bloating */
.section-body th:nth-child(2),
.section-body td:nth-child(2) {{
  max-width: 60px;
}}
.section-body th:nth-child(4),
.section-body td:nth-child(4) {{
  white-space: nowrap;
}}
.section-body th:nth-child(9),
.section-body td:nth-child(9) {{
  min-width: 42px;
  white-space: nowrap;
}}
.section-body th:nth-child(11),
.section-body td:nth-child(11) {{
  max-width: 60px;
}}

/* ── No hover colour change on report tables ── */
.section-body tbody tr:hover td {{
  background:inherit;
  transition:none;
}}

/* ── Report section font: -7% from current ── */
.section-body table {{ font-size:0.84rem; }}
.section-body th    {{ font-size:0.78rem; }}
.section-body td    {{ font-size:0.82rem; }}

/* ── Report section: left align all content ── */
.section-body td {{ text-align:left; }}
/* keep numeric/centre overrides for specific cells */
.section-body td.td-center {{ text-align:center; }}
.section-body td.td-right  {{ text-align:right; }}


/* ── High Conviction Cards ── */
.hc-section{{
  margin:28px 0 20px;
}}
.hc-grid{{
  display:flex;
  flex-wrap:wrap;
  gap:12px;
  margin-top:12px;
}}
.hc-card{{
  background:#0d1e38;
  border:1px solid #1e3a5f;
  border-radius:10px;
  padding:14px 16px;
  min-width:200px;
  max-width:240px;
  flex:1 1 200px;
  box-shadow:0 2px 10px rgba(0,0,0,.45);
  cursor:pointer;
  transition:border-color .15s, box-shadow .15s;
  text-decoration:none;
  display:block;
  position:relative;
}}
.hc-card:hover{{
  border-color:#4ecb82;
  box-shadow:0 4px 18px rgba(78,203,130,.18);
}}
.hc-card-symbol{{
  font-size:.95rem;
  font-weight:700;
  color:#dce8f8;
  letter-spacing:.5px;
  margin-bottom:6px;
}}
.hc-card-score{{
  font-size:1.1rem;
  font-weight:700;
  color:#4ecb82;
  margin-bottom:8px;
}}
.hc-card-score.mid{{color:#f5a623}}
.hc-card-score.low{{color:#f06464}}
.hc-card-tag{{
  display:inline-block;
  font-size:.72rem;
  padding:2px 7px;
  border-radius:12px;
  margin:2px 2px 2px 0;
  font-weight:600;
}}
.tag-trend{{background:#12304a;color:#5aabff}}
.tag-rsi{{background:#1a2f1a;color:#4ecb82}}
.tag-vol{{background:#2d1f0a;color:#f5a623}}
.tag-idx{{background:#1e1240;color:#b388ff}}
.hc-card-arrow{{
  position:absolute;
  bottom:10px;right:14px;
  font-size:.75rem;
  color:#c8dcf4;
  opacity:.65;
}}
.hc-card:hover .hc-card-arrow{{opacity:1;color:#4ecb82}}

</style>

<script>
(function(){{
  function cellVal(td,num){{
    var t=td?td.textContent.trim():'';
    if(num){{var n=parseFloat(t.replace(/[^0-9.\x2D]/g,''));return isNaN(n)?-Infinity:n;}}
    return t.toLowerCase();
  }}
  function sortTable(th){{
    var tbl=th.closest('table'),col=parseInt(th.dataset.col,10),
        num=th.dataset.numeric==='1',dir=th.dataset.dir;
    dir=(dir==='desc')?'asc':'desc';
    tbl.querySelectorAll('th.th-sort').forEach(function(h){{h.classList.remove('sorted-asc','sorted-desc');h.dataset.dir='none';}});
    th.dataset.dir=dir;th.classList.add(dir==='asc'?'sorted-asc':'sorted-desc');
    var tbody=tbl.querySelector('tbody'),rows=Array.from(tbody.querySelectorAll('tr'));
    rows.sort(function(a,b){{
      var va=cellVal(a.querySelectorAll('td')[col],num),vb=cellVal(b.querySelectorAll('td')[col],num);
      return va<vb?(dir==='asc'?-1:1):va>vb?(dir==='asc'?1:-1):0;
    }});
    rows.forEach(function(r){{tbody.appendChild(r);}});
  }}
  function toggleSection(btn){{
    var body=btn.closest('.section-block').querySelector('.section-body');
    var hide=body.style.display==='none';
    body.style.display=hide?'':'none';
    btn.textContent=hide?'▲ Collapse':'▼ Expand';
  }}
  document.addEventListener('DOMContentLoaded',function(){{
    // Sortable headers
    document.querySelectorAll('th.th-sort').forEach(function(th){{
      th.addEventListener('click',function(){{sortTable(th);}});
    }});
    document.querySelectorAll('th.th-sort[data-dir="desc"]').forEach(function(th){{
      th.classList.add('sorted-desc');
    }});
    // Collapse toggles
    document.querySelectorAll('.section-toggle').forEach(function(btn){{
      btn.addEventListener('click',function(){{toggleSection(btn);}});
    }});
    // Wrap tables in .tbl-wrap
    document.querySelectorAll('table').forEach(function(t){{
      if(!t.parentElement.classList.contains('tbl-wrap')){{
        var w=document.createElement('div');w.className='tbl-wrap';
        t.parentNode.insertBefore(w,t);w.appendChild(t);
      }}
    }});
    // Auto signal count badges
    document.querySelectorAll('.section-block').forEach(function(sec){{
      var badge=sec.querySelector('.section-badge');
      if(!badge)return;
      var n=sec.querySelectorAll('tbody tr').length;
      badge.textContent=n+(n===1?' signal':' signals');
    }});
    // Colour conviction score cells
    document.querySelectorAll('tbody tr').forEach(function(row){{
      var tds=row.querySelectorAll('td');
      if(tds.length<18)return;
      var td=tds[17],v=parseFloat(td.textContent);
      if(!isNaN(v)){{
        td.classList.add(v>=70?'conv-high':v>=40?'conv-mid':'conv-low');
      }}
    }});
  }});
}})();
</script>
</head>
<body>
<header class='site-header' id='top'>
  <h1>&#9642;&nbsp; Nifty 100 &mdash; Daily EMA Report</h1>
  <div class='meta-bar'>
    <strong>{TODAY_IST.strftime('%A, %d %b %Y &nbsp;&middot;&nbsp; %I:%M %p IST')}</strong>
    &nbsp;&middot;&nbsp; v{APP_VERSION}
    &nbsp;&middot;&nbsp; {len(history):,} history rows
    &nbsp;&middot;&nbsp; last: {last_date}
    &nbsp;&middot;&nbsp; price: yfinance (adj.)
    <br><span class='hint-pill'>&#x21C5; click column header to sort &nbsp;&middot;&nbsp; click section heading to collapse</span>
  </div>
</header>
<div class='page-wrap'>
{_mh_html}
{_hc_html}
{section_21_50}
{section_9_21}
{section_pb_ema9}
{section_pb_ema21}
{section_ribbon_new}
{section_ribbon}
<div class='disclaimer'>
  &#9888;&nbsp; Automatically generated for informational purposes only.
  Validate all signals on your broker or charting platform before acting.
  Past performance is not indicative of future results.
</div>
</div>
<a href='#top' id='back-top' title='Back to top'>&#8593;</a>
</body></html>"""

    fname.write_text(page, encoding="utf-8")
    logger.info("HTML report written: %s", fname)
    return fname


def _write_excel_sheet(wb: Workbook, sheet_title: str, report_df: pd.DataFrame) -> None:
    """Add one worksheet to *wb* for the given report DataFrame."""
    ws = wb.create_sheet(title=sheet_title)

    # Columns to export: drop HTML-only columns; rename text-breakdown for clean header
    _drop_cols = ["RecentNewsHtml", "RecentNewsText", "Conviction Explanation"]
    excel_df = report_df.drop(columns=_drop_cols, errors="ignore").copy()
    if "Conviction Explanation Text" in excel_df.columns:
        excel_df.rename(columns={"Conviction Explanation Text": "Conviction Explanation"}, inplace=True)
    # Convert newlines in merged columns to readable separators for Excel cells
    if "SuperTrend(10,3) (10,1) Green" in excel_df.columns:
        excel_df["SuperTrend(10,3) (10,1) Green"] = (
            excel_df["SuperTrend(10,3) (10,1) Green"].astype(str).str.replace("\n", "  /  ", regex=False)
        )
    if "Support and Resistance" in excel_df.columns:
        excel_df["Support and Resistance"] = (
            excel_df["Support and Resistance"].astype(str).str.replace("\n\n", "    |    ", regex=False)
        )

    headers = list(excel_df.columns) if not excel_df.empty else ["Message"]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    if "Message" in excel_df.columns:
        ws.append([excel_df.iloc[0]["Message"]])
    else:
        for row in excel_df.itertuples(index=False):
            ws.append(list(row))

    for col in ws.columns:
        width = max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2
        _c0 = next((c for c in col if hasattr(c, "column_letter")), None)
        if _c0 is not None:
            ws.column_dimensions[_c0.column_letter].width = min(width, 40)


def _write_signal_tracker_sheet(wb) -> None:
    """Add Signal Tracker sheet to the daily Excel workbook."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    sig_df = _load_signal_log()
    ws     = wb.create_sheet("Signal Tracker")

    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    SUM_FILL  = PatternFill("solid", fgColor="2E75B6")
    ACT_FILL  = PatternFill("solid", fgColor="E2EFDA")
    XBCK_FILL = PatternFill("solid", fgColor="FCE4D6")
    STOP_FILL = PatternFill("solid", fgColor="FFCCCC")
    COMP_FILL = PatternFill("solid", fgColor="DDEBF7")
    ARCH_FILL = PatternFill("solid", fgColor="F2F2F2")
    POS_FILL  = PatternFill("solid", fgColor="C6EFCE")
    NEG_FILL  = PatternFill("solid", fgColor="FFC7CE")
    YES_FILL  = PatternFill("solid", fgColor="C6EFCE")
    NO_FILL   = PatternFill("solid", fgColor="FFC7CE")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
    SUM_FONT  = Font(bold=True, color="FFFFFF", size=10)
    BODY_FONT = Font(size=9)
    thin      = Side(style="thin", color="BFBFBF")
    BRD       = Border(left=thin, right=thin, top=thin, bottom=thin)

    display_cols = [
        ("Signal Date",  "signal_date"),
        ("Type",         "signal_type"),
        ("Symbol",       "symbol"),
        ("Company",      "company"),
        ("Entry",        "signal_close"),
        ("Curr Close",   "current_close"),
        ("Ret %",        "pct_return"),
        ("Ret INR",      "abs_return"),
        ("Max Gain%",    "max_pct_gain"),
        ("Max DD%",      "max_pct_drawdown"),
        ("Days",         "days_since_signal"),
        ("EMA OK",       "ema_still_aligned"),
        ("Ribbon",       "ribbon_active"),
        ("RSI",          "current_rsi"),
        ("Deliv%",       "current_delivery_pct"),
        ("Stop",         "suggested_stop"),
        ("Tgt1",         "suggested_target_1"),
        ("Tgt2",         "suggested_target_2"),
        ("Stop Hit",     "stop_hit"),
        ("Tgt1 Hit",     "target1_hit"),
        ("Tgt2 Hit",     "target2_hit"),
        ("Status",       "status"),
        ("Updated",      "last_updated"),
        ("Conv",         "signal_conviction_score"),
    ]
    headers = [d[0] for d in display_cols]
    fields  = [d[1] for d in display_cols]
    ncols   = len(headers)

    total    = len(sig_df) if not sig_df.empty else 0
    act_df   = sig_df[sig_df["status"] == "Active"] if not sig_df.empty else pd.DataFrame()
    n_act    = len(act_df)
    n_prof   = int((act_df["pct_return"] > 0).sum()) if not act_df.empty else 0
    avg_ret  = round(float(act_df["pct_return"].mean()), 2) if not act_df.empty else 0.0
    best_lbl = ""
    if not act_df.empty:
        bi      = act_df["pct_return"].idxmax()
        best_lbl = f'{act_df.at[bi,"symbol"]} ({act_df.at[bi,"pct_return"]:+.1f}%)'

    summary = (
        f"Signal Tracker  |  Total: {total}  |  Active: {n_act}  |  "
        f"Profitable: {n_prof}/{n_act}  |  Avg Ret: {avg_ret:+.2f}%  |  "
        f"Best: {best_lbl}  |  As of: {RESOLVED_TRADE_DATE}"
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    sc = ws.cell(1, 1, summary)
    sc.fill = SUM_FILL; sc.font = SUM_FONT
    sc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    for ci, hdr in enumerate(headers, 1):
        c = ws.cell(2, ci, hdr)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.border = BRD
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "A3"

    if sig_df.empty:
        ws.cell(3, 1, "No signals logged yet.")
        return

    STATUS_ORDER = {"Active": 0, "Crossed Back": 1, "Stop Hit": 2, "Completed": 3, "Archived": 4}
    sig_df["_so"] = sig_df["status"].map(STATUS_ORDER).fillna(9)
    sig_df = sig_df.sort_values(["_so", "pct_return"], ascending=[True, False]).drop(columns=["_so"])

    SFILL = {"Active": ACT_FILL, "Crossed Back": XBCK_FILL,
             "Stop Hit": STOP_FILL, "Completed": COMP_FILL, "Archived": ARCH_FILL}

    for ri, (_, row) in enumerate(sig_df.iterrows(), start=3):
        rfill = SFILL.get(str(row.get("status", "")), ARCH_FILL)
        for ci, field in enumerate(fields, 1):
            val = row.get(field, "")
            if field in ("signal_date", "last_updated"):
                try:   val = pd.Timestamp(val).strftime("%Y-%m-%d") if pd.notna(val) else ""
                except: val = str(val) if val else ""
            elif field in ("ema_still_aligned", "ribbon_active",
                           "stop_hit", "target1_hit", "target2_hit"):
                val = "YES" if val else "NO"
            elif field in ("pct_return", "max_pct_gain", "max_pct_drawdown",
                           "signal_close", "current_close", "abs_return",
                           "suggested_stop", "suggested_target_1", "suggested_target_2",
                           "signal_conviction_score", "current_rsi", "current_delivery_pct"):
                try:   val = round(float(val), 2) if pd.notna(val) else ""
                except: val = ""

            c = ws.cell(ri, ci, val)
            c.font = BODY_FONT; c.border = BRD; c.fill = rfill
            c.alignment = Alignment(horizontal="center", vertical="center")

            if field == "pct_return":
                try:    c.fill = POS_FILL if float(row.get("pct_return", 0)) >= 0 else NEG_FILL
                except: pass
            if field == "ema_still_aligned":
                c.fill = YES_FILL if row.get("ema_still_aligned") else NO_FILL
            if field == "status":
                c.font = Font(bold=True, size=9)

    cw = {"Signal Date": 13, "Type": 11, "Symbol": 10, "Company": 26,
          "Entry": 9, "Curr Close": 11, "Ret %": 8, "Ret INR": 9,
          "Max Gain%": 10, "Max DD%": 9, "Days": 6, "EMA OK": 7,
          "Ribbon": 7, "RSI": 7, "Deliv%": 8, "Stop": 9, "Tgt1": 9,
          "Tgt2": 9, "Stop Hit": 9, "Tgt1 Hit": 9, "Tgt2 Hit": 9,
          "Status": 13, "Updated": 13, "Conv": 8}
    for ci, hdr in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = cw.get(hdr, 11)


def render_excel(reports: dict, market_health: dict = None) -> Path:
    fname = REPORTS_DIR / f"nifty100_daily_report_{TODAY_IST.strftime('%Y%m%d')}.xlsx"
    wb = Workbook()

    # ── Market Health sheet (Sheet 1 — inserted first) ──────────────────────
    if market_health:
        try:
            render_market_health_excel(wb, market_health)
        except Exception as _mhe:
            logger.warning("Market Health Excel sheet skipped: %s", _mhe)

    # ── Tab 1: original EMA 21/50 report (rename default sheet) ──────────────
    ws_default = wb.active
    ws_default.title = "EMA 21_50 Crossover"
    _drop_cols_e = ["RecentNewsHtml", "RecentNewsText", "Conviction Explanation"]
    df_21_50 = reports["ema_21_50"].drop(columns=_drop_cols_e, errors="ignore").copy()
    if "Conviction Explanation Text" in df_21_50.columns:
        df_21_50.rename(columns={"Conviction Explanation Text": "Conviction Explanation"}, inplace=True)
    if "SuperTrend(10,3) (10,1) Green" in df_21_50.columns:
        df_21_50["SuperTrend(10,3) (10,1) Green"] = (
            df_21_50["SuperTrend(10,3) (10,1) Green"].astype(str).str.replace("\n", "  /  ", regex=False)
        )
    if "Support and Resistance" in df_21_50.columns:
        df_21_50["Support and Resistance"] = (
            df_21_50["Support and Resistance"].astype(str).str.replace("\n\n", "    |    ", regex=False)
        )
    headers = list(df_21_50.columns) if not df_21_50.empty else ["Message"]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    ws_default.append(headers)
    for cell in ws_default[1]:
        cell.fill = header_fill
        cell.font = header_font
    if "Message" in df_21_50.columns:
        ws_default.append([df_21_50.iloc[0]["Message"]])
    else:
        for row in df_21_50.itertuples(index=False):
            ws_default.append(list(row))
    for col in ws_default.columns:
        width = max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2
        _col0 = next((c for c in col if hasattr(c, "column_letter")), None)
        if _col0 is not None:
            ws_default.column_dimensions[_col0.column_letter].width = min(width, 40)

    # ── Tab 2: EMA 9/21 crossover ─────────────────────────────────────────────
    _write_excel_sheet(wb, "EMA 9_21 Crossover", reports["ema_9_21"])

    # ── Tab 3: EMA Ribbon ─────────────────────────────────────────────────────
    _write_excel_sheet(wb, "EMA Ribbon New Entry",  reports["ema_ribbon_new"])
    _write_excel_sheet(wb, "EMA ribbon Crossover",  reports["ema_ribbon"])
    _write_excel_sheet(wb, "Pullback to EMA 9",  reports["pb_ema9"])
    _write_excel_sheet(wb, "Pullback to EMA 21", reports["pb_ema21"])

    # ── Signal Tracker tab ────────────────────────────────────────────────────
    _write_signal_tracker_sheet(wb)

    wb.save(fname)
    logger.info("Excel report written: %s", fname)
    return fname


def get_gmail_service() -> "googleapiclient.discovery.Resource":
    token_path = BASE_DIR / "token.json"
    creds = None

    if token_path.exists():
        logger.info("Loading Gmail tokens from token.json")
        creds = Credentials.from_authorized_user_file(str(token_path), GMAIL_SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            logger.info("Refreshing Gmail API token...")
            creds.refresh(Request())
        else:
            logger.info("Starting Gmail OAuth flow in browser...")
            flow = InstalledAppFlow.from_client_secrets_file(
                str(BASE_DIR / "credentials.json"), GMAIL_SCOPES
            )
            creds = flow.run_local_server(port=0)
        token_path.write_text(creds.to_json(), encoding="utf-8")
        logger.info("Gmail token saved to %s", token_path)

    service = build("gmail", "v1", credentials=creds)
    return service

def send_email_with_reports_via_gmail_api(cfg: Config, html_path: Path, xlsx_path: Path):
    if not cfg.email_enabled:
        logger.info("Email disabled in config; skipping Gmail API send.")
        return
    if not cfg.mail_to:
        logger.warning("No mail_to recipients configured; skipping Gmail API send.")
        return

    subject = cfg.__dict__.get(
        "subject",
        f"Nifty 100 EMA Bullish Cross Report - Local Execution - {TODAY_IST.strftime('%Y-%m-%d')}",
    )
    body_text = (
        "Attached is today's Nifty 100 EMA crossover report.\n\n"
        "This email was sent automatically by the Python scanner via Gmail API."
    )

    msg = MIMEMultipart()
    msg["From"] = cfg.mail_from
    msg["To"] = ", ".join(cfg.mail_to)
    msg["Subject"] = subject
    msg.attach(MIMEText(body_text, "plain"))

    # Attach HTML report
    with open(html_path, "rb") as f:
        part = MIMEBase("text", "html")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={html_path.name}")
    msg.attach(part)

    # Attach Excel report
    with open(xlsx_path, "rb") as f:
        part2 = MIMEBase(
            "application",
            "vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        part2.set_payload(f.read())
    encoders.encode_base64(part2)
    part2.add_header("Content-Disposition", f"attachment; filename={xlsx_path.name}")
    msg.attach(part2)

    # Build raw message for Gmail API
    import base64
    raw_message = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")

    service = get_gmail_service()
    logger.info("Sending email via Gmail API to %s", cfg.mail_to)

    from googleapiclient.errors import HttpError
    try:
        service.users().messages().send(
            userId="me",
            body={"raw": raw_message},
        ).execute()
        logger.info("Email sent successfully via Gmail API.")
    except HttpError as e:
        logger.exception("Failed to send email via Gmail API: %s", e)



def send_email_with_reports(cfg: Config, html_path: Path, xlsx_path: Path):
    if not cfg.email_enabled:
        logger.info("Email disabled in config; skipping email send.")
        return
    if not cfg.mail_to:
        logger.warning("No mail_to recipients configured; skipping email send.")
        return

    subject = cfg.__dict__.get(
        "subject",
        f"Nifty 100 EMA 21/50 Bullish Cross Report - {TODAY_IST.strftime('%Y-%m-%d')}",
    )
    body_text = (
        "Attached is today's Nifty 100 EMA 21/50 crossover report.\n\n"
        "This email was sent automatically by the Python scanner."
    )

    msg = MIMEMultipart()
    msg["From"] = cfg.mail_from
    msg["To"] = ", ".join(cfg.mail_to)
    msg["Subject"] = subject
    msg.attach(MIMEText(body_text, "plain"))

    # Attach HTML report
    with open(html_path, "rb") as f:
        part = MIMEBase("text", "html")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={html_path.name}")
    msg.attach(part)

    # Attach Excel report
    with open(xlsx_path, "rb") as f:
        part2 = MIMEBase(
            "application",
            "vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        part2.set_payload(f.read())
    encoders.encode_base64(part2)
    part2.add_header("Content-Disposition", f"attachment; filename={xlsx_path.name}")
    msg.attach(part2)

    logger.info(
        "Connecting to SMTP server %s:%s as %s",
        cfg.smtp_host,
        cfg.smtp_port,
        cfg.smtp_username,
    )

    import smtplib
    try:
        with smtplib.SMTP(cfg.smtp_host, cfg.smtp_port, timeout=20) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.login(cfg.smtp_username, cfg.smtp_password)
            server.sendmail(cfg.mail_from, cfg.mail_to, msg.as_string())
        logger.info("Email sent successfully to %s", cfg.mail_to)
    except Exception as e:
        logger.exception("Failed to send email via SMTP: %s", e)

def main():
    logger.info("%s | Version %s | Released %s", APP_NAME, APP_VERSION, APP_RELEASE_DATE)
    logger.info("Base directory: %s", BASE_DIR)
    cfg = load_config()
    s = session_with_headers()
    bootstrap_nse_session(s)
    universe = fetch_universe(s)
    history = update_history(s, universe, cfg)

    # ── Data currency log lines ───────────────────────────────────────────────
    _ohlcv_date = history["Date"].max().strftime("%Y-%m-%d") if not history.empty else "N/A"
    logger.info("Latest OHLCV data is from %s", _ohlcv_date)
    try:
        import pandas as _pd_chk
        if DELIVERY_HISTORY_PATH.exists():
            _dh = _pd_chk.read_csv(DELIVERY_HISTORY_PATH, parse_dates=["Date"])
            _deliv_date = _dh["Date"].max().strftime("%Y-%m-%d")
        else:
            _deliv_date = "N/A (delivery cache not yet built)"
    except Exception:
        _deliv_date = "N/A (error reading delivery cache)"
    logger.info("Latest Delivery data is from %s", _deliv_date)

    reports = build_analysis(history, universe, cfg, s)

    # ── Market Health Dashboard ───────────────────────────────────────────
    market_health = {}
    try:
        _mh_index_map = fetch_index_membership(s)
        market_health = compute_market_health(s, history, _mh_index_map)
    except Exception as _mh_err:
        logger.warning("Market Health computation skipped: %s", _mh_err)

    # ── Signal Tracker ────────────────────────────────────────────────────────
    def _df_to_rows(val) -> list:
        """Convert a build_analysis report value (DataFrame or list) to list of dicts."""
        if val is None:
            return []
        if hasattr(val, "to_dict"):           # it's a DataFrame
            rows = val.to_dict(orient="records")
            # drop the empty-message sentinel rows
            return [r for r in rows if isinstance(r, dict) and r.get("Symbol")]
        if isinstance(val, list):
            return [r for r in val if isinstance(r, dict) and r.get("Symbol")]
        return []

    _rows_21_50      = _df_to_rows(reports.get("ema_21_50"))
    _rows_9_21       = _df_to_rows(reports.get("ema_9_21"))
    _rows_pb_ema9    = _df_to_rows(reports.get("pb_ema9"))
    _rows_pb_ema21   = _df_to_rows(reports.get("pb_ema21"))
    _rows_ribbon_new = _df_to_rows(reports.get("ema_ribbon_new"))
    _signal_syms = {
        r["Symbol"]
        for r in _rows_21_50 + _rows_9_21 + _rows_pb_ema9 + _rows_pb_ema21 + _rows_ribbon_new
    }
    _stock_data  = {
        sym: history[history["Symbol"] == sym]
        for sym in _signal_syms
    }
    log_new_signals(
        rows_21_50      = _rows_21_50,
        rows_9_21       = _rows_9_21,
        rows_pb_ema9    = _rows_pb_ema9,
        rows_pb_ema21   = _rows_pb_ema21,
        rows_ribbon_new = _rows_ribbon_new,
        stock_data      = _stock_data,
        cfg             = cfg,
    )
    update_signal_tracker(history, cfg)

    html_path = render_html(reports, history, market_health=market_health)
    xlsx_path = render_excel(reports, market_health=market_health)
    #send_email_with_reports(cfg, html_path, xlsx_path)
    if os.environ.get("GITHUB_ACTIONS") == "true":
        logger.info("GitHub Actions detected — skipping Gmail API send (workflow handles email).")
    else:
        send_email_with_reports_via_gmail_api(cfg, html_path, xlsx_path)
    logger.info("Run completed successfully.")
    logger.info("Artifacts: %s | %s | log: %s", html_path, xlsx_path, LOG_FILE)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.exception("Run failed: %s", e)
        print(f"\nERROR: {e}\nCheck log file: {LOG_FILE}")
        sys.exit(1)
    sys.exit(0)
